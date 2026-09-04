#!/usr/bin/env python3
"""Export disclosure↔AMFI mapping Excel snapshot to exports/.

  .venv/bin/python3 matching/export_mapping_xlsx.py
"""
from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]


def load(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def style_header(ws, ncols: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 0
        for cell in col[:80]:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = max(10, width + 2)


def main() -> None:
    disc = load(ROOT / "data/sources/disclosure_to_amfi_global_mapping.json")
    nav = load(ROOT / "data/sources/amfi_navall_to_disclosure_global_mapping.json")
    active_path = ROOT / "data/sources/amfi_schemes_active_aug2026.json"
    active = load(active_path)["schemes"] if active_path.exists() else []

    stamp = datetime.now().strftime("%Y-%m-%d")
    out = ROOT / "exports" / f"amfi_disclosure_complete_mapping_{stamp}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    orange = PatternFill("solid", fgColor="FCE4D6")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    green = PatternFill("solid", fgColor="C6EFCE")

    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "AMFI ↔ Disclosure complete mapping"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated: {datetime.now(timezone.utc).isoformat()}"
    ws["A3"] = "Sources: data/sources/*global_mapping.json + registry/shortcode map"
    notes = [
        ("disclosure_to_parent", "Disclosure fund → AMFI parent"),
        ("parent_summary_aug_nav", "Active parents coverage"),
        ("navall_plan_to_disclosure", "NAVAll plan → disclosure"),
        ("pending_and_specials", "Unmapped / alias / segregated notes"),
    ]
    for i, (a, b) in enumerate(notes, start=5):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)
    autosize(ws)

    headers = [
        "amc_name",
        "disclosure_fund_name",
        "disclosure_fund_shortname",
        "fund_amfi_scheme_id",
        "parent_fund_name",
        "mapped",
        "match_how",
        "is_segregated_disclosure",
        "fund_amfi_plan_codes",
        "fix_note",
    ]
    ws = wb.create_sheet("disclosure_to_parent")
    ws.append(headers)
    style_header(ws, len(headers))
    for r in sorted(
        disc["mappings"],
        key=lambda x: ((x.get("amc_name") or ""), (x.get("disclosure_fund_name") or "")),
    ):
        row = [
            r.get("amc_name"),
            r.get("disclosure_fund_name"),
            r.get("disclosure_fund_shortname"),
            r.get("fund_amfi_scheme_id"),
            r.get("parent_fund_name"),
            bool(r.get("mapped")),
            r.get("match_how"),
            bool(r.get("is_segregated_disclosure")),
            ",".join(str(x) for x in (r.get("fund_amfi_plan_codes") or [])),
            r.get("fix_note"),
        ]
        ws.append(row)
        if not row[5] or not row[3]:
            for c in range(1, len(headers) + 1):
                ws.cell(ws.max_row, c).fill = orange
    autosize(ws)

    aug = defaultdict(lambda: {"plans": 0, "with_disc": 0, "names": set(), "parent": None})
    for r in nav["mappings"]:
        sid = str(r.get("fund_amfi_scheme_id") or "")
        if not sid or sid == "None":
            continue
        b = aug[sid]
        b["plans"] += 1
        b["parent"] = r.get("parent_fund_name")
        if r.get("disclosure_fund_name"):
            b["with_disc"] += 1
            b["names"].add(r["disclosure_fund_name"])

    headers2 = [
        "fund_amfi_scheme_id",
        "parent_fund_name",
        "nav_plans",
        "plans_with_disclosure",
        "disclosure_names",
        "coverage_ok",
    ]
    ws = wb.create_sheet("parent_summary_aug_nav")
    ws.append(headers2)
    style_header(ws, len(headers2))
    for s in sorted(active, key=lambda x: x.get("scheme_name") or ""):
        sid = str(s["scheme_id"])
        b = aug.get(sid) or {"plans": 0, "with_disc": 0, "names": set(), "parent": s["scheme_name"]}
        disc_names = {
            r.get("disclosure_fund_name")
            for r in disc["mappings"]
            if str(r.get("fund_amfi_scheme_id")) == sid and r.get("mapped")
        }
        names = sorted((b["names"] | disc_names) - {None, ""})
        plans = b["plans"]
        with_disc = b["with_disc"]
        ok = plans == 0 or with_disc == plans
        ws.append([sid, b["parent"] or s["scheme_name"], plans, with_disc, " | ".join(names)[:320], ok])
        if not ok:
            for c in range(1, len(headers2) + 1):
                ws.cell(ws.max_row, c).fill = yellow
    autosize(ws)

    headers3 = [
        "fund_amfi_plan_code",
        "plan_name",
        "fund_amfi_scheme_id",
        "parent_fund_name",
        "disclosure_fund_name",
        "disclosure_fund_shortname",
        "mapped",
        "last_known_nav_date",
    ]
    ws = wb.create_sheet("navall_plan_to_disclosure")
    ws.append(headers3)
    style_header(ws, len(headers3))
    for r in nav["mappings"]:
        ws.append(
            [
                r.get("fund_amfi_plan_code"),
                r.get("plan_name"),
                r.get("fund_amfi_scheme_id"),
                r.get("parent_fund_name"),
                r.get("disclosure_fund_name"),
                r.get("disclosure_fund_shortname"),
                bool(r.get("mapped")),
                r.get("last_known_nav_date"),
            ]
        )
    autosize(ws, max_width=40)

    headers4 = [
        "kind",
        "amc_name",
        "disclosure_fund_name",
        "shortname",
        "fund_amfi_scheme_id",
        "parent_fund_name",
        "mapped",
        "match_how",
        "fix_note",
    ]
    ws = wb.create_sheet("pending_and_specials")
    ws.append(headers4)
    style_header(ws, len(headers4))
    for r in disc["mappings"]:
        kind = None
        if not r.get("mapped") or not r.get("fund_amfi_scheme_id"):
            kind = "unmapped_disclosure"
        elif r.get("is_segregated_disclosure"):
            kind = "segregated_disclosure"
        elif r.get("fix_note") or (
            r.get("match_how")
            and any(k in (r.get("match_how") or "") for k in ("manual", "alias", "renamed", "incremental"))
        ):
            kind = "alias_or_manual_note"
        if not kind:
            continue
        ws.append(
            [
                kind,
                r.get("amc_name"),
                r.get("disclosure_fund_name"),
                r.get("disclosure_fund_shortname"),
                r.get("fund_amfi_scheme_id"),
                r.get("parent_fund_name"),
                bool(r.get("mapped")),
                r.get("match_how"),
                r.get("fix_note"),
            ]
        )
        fill = orange if kind == "unmapped_disclosure" else green if kind == "alias_or_manual_note" else yellow
        for c in range(1, len(headers4) + 1):
            ws.cell(ws.max_row, c).fill = fill
    autosize(ws)

    wb.save(out)
    meta = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "excel": str(out.relative_to(ROOT)),
        "disclosure_rows": len(disc["mappings"]),
        "disclosure_mapped": sum(
            1 for r in disc["mappings"] if r.get("mapped") and r.get("fund_amfi_scheme_id")
        ),
        "nav_rows": len(nav["mappings"]),
        "active_parents": len(active),
    }
    (ROOT / "exports" / "amfi_disclosure_complete_mapping_meta.json").write_text(
        json.dumps(meta, indent=2) + "\n", encoding="utf-8"
    )
    print(f"Wrote {out}")
    print(json.dumps(meta, indent=2))


if __name__ == "__main__":
    main()
