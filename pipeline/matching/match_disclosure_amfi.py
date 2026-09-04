#!/usr/bin/env python3
"""
Match disclosure scheme names/shortcodes to AMFI base funds.

Strategy:
  1. Map AMC registry id → AMFI AMC name (fuzzy on AMC label).
  2. Collect disclosure scheme labels per AMC (Index tables preferred,
     then sheet titles / cleaned schemes.json). Capture durable pack
     shortcodes from filenames / sheets whenever present.
  3. Resolve AMFI via durable shortcode map first (amc_id+shortcode);
     fuzzy name-match only for packs without a known shortcode binding.
  4. Persist new confirmed shortcode→AMFI bindings so later months reuse them.
  5. Write per-AMC matched vs unmatched matrices.

Portfolio disclosures are scheme-level (not Direct/Regular plans), so matching
is against collapsed AMFI funds (data/amfi/funds.json).
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import tempfile
import zipfile
import warnings
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from rapidfuzz import fuzz, process

warnings.filterwarnings("ignore")

try:
    from amfi_navall import norm_text  # type: ignore
except ImportError:
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from amfi_navall import norm_text

JUNK_SCHEME_RE = re.compile(
    r"(?ix)^(index|contents|cover|notes?|disclaimer|summary|risk.?o.?meter|"
    r"annexure|instruction|read.?me|legend|glossary|overview|"
    r"counter[-\s]?party|derivative|dividend\s+history|debt\s+replication|"
    r"total|sub[\s\-]?total|grand\s+total|"
    r"\(a\)|\(b\)|\(c\)|"
    r"listed\b|privately\s+placed|unlisted|"
    r"equity\s*&\s*equity|debt\s+instruments|money\s+market|"
    r"cash\b|tri[- ]?party|treps|net\s+receivables|"
    r"others?\b|classification)$"
)
JUNK_CONTAINS_RE = re.compile(
    r"(?ix)awaiting\s+listing|privately\s+placed|risk.?o.?meter|"
    r"industry\s+classification|as\s+recommended\s+by\s+amfi|"
    r"^0[_\s].*strips|t[-\s]?bill\b"
)
# Leading digits allowed (Mirae 1DGROWTH / 200EWETF).
SHORTCODE_RE = re.compile(r"^[A-Za-z0-9]{2,20}$")
SHORTCODE_BLOCKLIST = {
    "industry",
    "quantity",
    "isin",
    "rating",
    "sector",
    "scheme",
    "index",
    "total",
    "equity",
    "debt",
    "value",
    "market",
    "assets",
    "coupon",
    "yield",
    "maturity",
    "underlying",
    "long",
    "short",
    "sheet",
    "sheet1",
    "click",
    "notes",
    "disclosure",
    "portfolio",
    "monthly",
    "fortnight",
    "fortnightly",
    "report",
    "contents",
    "cover",
}
_MONTH_TOKEN = (
    r"(?:jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|"
    r"jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)"
)
_DESC_FILENAME_RE = re.compile(
    r"(?i)\b(fund|etf|index|fof|hybrid|liquid|debt|equity|bond|duration|"
    r"arbitrage|gilt|silver|gold|mid|cap|large|small|flexi|multi|asset|"
    r"overnight|money|portfolio)\b"
)
_AMC_BRAND_CODES = {
    "SBI",
    "HDFC",
    "HSBC",
    "ICICI",
    "UTI",
    "DSP",
    "AXIS",
    "NIPPON",
    "KOTAK",
    "BANDHAN",
    "BAJAJ",
    "GROWW",
    "NAVI",
    "PGIM",
    "TATA",
    "FRANKLIN",
    "EDELWEISS",
    "MIRAE",
    "LIC",
    "UNION",
    "QUANT",
    "QUANTUM",
    "INVESCO",
    "CANARA",
    "ROBECO",
    "MOTILAL",
    "OSWAL",
    "WHITEOAK",
    "BARODA",
    "ADITYA",
    "BIRLA",
    "HELIO",
    "HELIOS",
    "ZERODHA",
    "ANGEL",
    "TAURUS",
    "SAMCO",
    "UNIFI",
    "PPFAS",
    "CHOICE",
}


def normalize_shortcode(label: str | None) -> str | None:
    """Compact ticker form (alphanumeric upper). Not for full sheet titles."""
    s = re.sub(r"[^A-Za-z0-9]", "", (label or "").strip())
    if not s:
        return None
    return s.upper()


def durable_disclosure_key(label: str | None) -> str | None:
    """Stable disclosure identity for monthly reuse.

    Tickers → compact uppercase (``1DGROWTH``). Sheet titles → as published
    (strip only), e.g. ``360 ONE Balanced Hybrid Fund``.
    """
    s = (label or "").strip()
    if not s or s.lower() in SHORTCODE_BLOCKLIST:
        return None
    if is_junk_label(s):
        return None
    # Compact tickers / pack codes
    if " " not in s and is_shortcode(s):
        return normalize_shortcode(s)
    # Sheet / Index labels that identify a fund — keep as-is
    if looks_like_fund_name(s) or len(s) >= 8:
        return s
    if is_shortcode(s):
        return normalize_shortcode(s)
    return None


def is_shortcode(label: str) -> bool:
    s = (label or "").strip()
    if not s or s.lower() in SHORTCODE_BLOCKLIST:
        return False
    if not SHORTCODE_RE.match(s):
        return False
    if s.upper() in _AMC_BRAND_CODES and not re.search(r"\d", s):
        return False
    # Prefer codes that look like AMC tickers (has digit or ALLCAPS ≥2), not plain English words
    if re.search(r"\d", s):
        return True
    if s.isupper() and len(s) >= 2:
        return True
    if re.search(r"[a-z]", s) and re.search(r"[A-Z]", s) and len(s) <= 12:
        return True  # CamelCase tickers
    return False


def pack_code_from_filename(stem: str) -> str | None:
    """Durable AMC pack/ticker from a disclosure filename stem (month-agnostic)."""
    s = (stem or "").strip()
    if not s:
        return None

    # Prefer explicit CODE_Portfolio / CODE_Monthly / CODE_<hash> (Capitalmind, Old Bridge)
    m = re.match(
        r"^([A-Za-z][A-Za-z0-9]{1,15})_(?:portfolio|monthly|disclosure|[0-9a-f]{6,})",
        s,
        re.I,
    )
    if m and is_shortcode(m.group(1)):
        return normalize_shortcode(m.group(1))

    # NJ-MF-Monthly-Portfolio-NJABF-June-2026-…
    m = re.search(rf"(?i)portfolio[-_]([A-Za-z][A-Za-z0-9]{{1,15}})[-_]{_MONTH_TOKEN}", s)
    if m and is_shortcode(m.group(1)):
        return normalize_shortcode(m.group(1))

    # ZNFTY - Monthly Portfolio June 2026
    m = re.match(
        rf"^([A-Za-z][A-Za-z0-9]{{1,15}})\s*[-–—]\s*(?:monthly|portfolio|disclosure)",
        s,
        re.I,
    )
    if m and is_shortcode(m.group(1)):
        return normalize_shortcode(m.group(1))

    # Canara MD-Canara-… / OF---Canara-… / FR-CRSF-July-2026 / MI-CRCHF-July-2026
    m = re.match(rf"^([A-Za-z]{{1,6}}\d{{0,3}})[-_]+(?:canara|robeco)\b", s, re.I)
    if m and is_shortcode(m.group(1)):
        return normalize_shortcode(m.group(1))
    m = re.match(rf"^([A-Za-z]{{1,4}})[-_]([A-Za-z0-9]{{2,12}})[-_]{_MONTH_TOKEN}", s, re.I)
    if m and not _DESC_FILENAME_RE.search(m.group(2)):
        # Prefer longer trailing ticker (CRSF / CRCHF) when present
        code = m.group(2) if len(m.group(2)) >= len(m.group(1)) else m.group(1)
        if is_shortcode(code):
            return normalize_shortcode(code)

    # LIC LEMETF08-07-2026-09_34_56 → LEMETF ; LET10008-07-2026 → LET100
    m = re.match(r"^([A-Za-z]{2,10}\d{0,3})(\d{2})[-_]\d{2}[-_]\d{4}", s)
    if m and is_shortcode(m.group(1)):
        return normalize_shortcode(m.group(1))

    # Mirae-style compact ticker-month: 1dgrowth-june2026 / man1dltf-june2026
    m = re.match(rf"^([A-Za-z0-9]{{2,20}})[-_]{_MONTH_TOKEN}\d{{0,4}}$", s, re.I)
    if m and not _DESC_FILENAME_RE.search(m.group(1)) and is_shortcode(m.group(1)):
        return normalize_shortcode(m.group(1))

    return None


def load_shortcode_map(path: Path) -> dict[str, dict[str, Any]]:
    """Return map keyed by ``amc_id::durable_key`` (ticker or sheet title as-is)."""
    if not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    out: dict[str, dict[str, Any]] = {}
    for row in data.get("entries") or []:
        amc_id = (row.get("amc_id") or "").strip()
        raw = (row.get("shortcode") or "").strip()
        if not amc_id or not raw:
            continue
        # Tickers stay compact; sheet titles stay as stored
        code = normalize_shortcode(raw) if (" " not in raw and is_shortcode(raw)) else raw
        amfi = str(row.get("canonical_amfi_code") or "").strip()
        if not code or not amfi:
            continue
        key = f"{amc_id}::{code}"
        out[key] = dict(row)
        out[key]["shortcode"] = code
        out[key]["canonical_amfi_code"] = amfi
        # casefold alias for sheet-title lookups
        if " " in code:
            out.setdefault(f"{amc_id}::{code.casefold()}", out[key])
        for alias in row.get("aliases") or []:
            ac_raw = (alias or "").strip()
            if not ac_raw:
                continue
            ac = normalize_shortcode(ac_raw) if (" " not in ac_raw and is_shortcode(ac_raw)) else ac_raw
            if ac and f"{amc_id}::{ac}" not in out:
                out[f"{amc_id}::{ac}"] = out[key]
    return out


def save_shortcode_map(path: Path, entries_by_key: dict[str, dict[str, Any]]) -> None:
    """Write durable shortcode registry (deduped by amc_id+key)."""
    primary: dict[str, dict[str, Any]] = {}
    for key, row in entries_by_key.items():
        amc_id = (row.get("amc_id") or key.split("::", 1)[0]).strip()
        raw = (row.get("shortcode") or key.split("::", 1)[-1] or "").strip()
        if " " in raw or not is_shortcode(raw):
            # Skip casefold shadow keys when saving
            if raw != raw.casefold() and key.endswith("::" + raw.casefold()):
                continue
            code = raw
        else:
            code = normalize_shortcode(raw)
        amfi = str(row.get("canonical_amfi_code") or "").strip()
        if not amc_id or not code or not amfi:
            continue
        # Drop pure casefold duplicates of a sheet title
        if " " in code and code == code.casefold() and any(
            (r.get("shortcode") or "").casefold() == code
            and (r.get("shortcode") or "") != code
            for r in entries_by_key.values()
            if (r.get("amc_id") or "") == amc_id
        ):
            # Prefer the cased sheet title entry if present elsewhere
            continue
        pk = f"{amc_id}::{code}"
        prev = primary.get(pk)
        if prev and prev.get("canonical_amfi_code") != amfi:
            # Keep confirmed/manual over auto; otherwise keep existing
            if (prev.get("confidence") or "") in {"confirmed", "manual"}:
                continue
            if (row.get("confidence") or "") not in {"confirmed", "manual"}:
                continue
        aliases = []
        for a in list(row.get("aliases") or []) + list((prev or {}).get("aliases") or []):
            ar = (a or "").strip()
            if not ar or ar == code:
                continue
            aliases.append(ar)
        aliases = sorted(set(aliases))
        primary[pk] = {
            "amc_id": amc_id,
            "shortcode": code,
            "aliases": aliases,
            "canonical_amfi_code": amfi,
            "amfi_base_name": row.get("amfi_base_name") or (prev or {}).get("amfi_base_name"),
            "disclosure_label": row.get("disclosure_label") or (prev or {}).get("disclosure_label"),
            "confidence": row.get("confidence") or (prev or {}).get("confidence") or "seeded",
            "source": row.get("source") or (prev or {}).get("source") or "match",
            "first_seen_period": (prev or {}).get("first_seen_period")
            or row.get("first_seen_period"),
            "last_verified_period": row.get("last_verified_period")
            or (prev or {}).get("last_verified_period"),
        }

    payload = {
        "version": 1,
        "note": (
            "Durable disclosure key → AMFI fund map. Key is a pack ticker "
            "(e.g. 1DGROWTH) or sheet title as-is (e.g. '360 ONE Balanced Hybrid Fund'). "
            "Matcher resolves these before name fuzzy-match."
        ),
        "entries": sorted(primary.values(), key=lambda r: (r["amc_id"], r["shortcode"])),
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    csv_path = path.with_suffix(".csv")
    import csv

    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "amc_id",
                "shortcode",
                "aliases",
                "canonical_amfi_code",
                "amfi_base_name",
                "disclosure_label",
                "confidence",
                "source",
                "first_seen_period",
                "last_verified_period",
            ],
        )
        w.writeheader()
        for row in payload["entries"]:
            w.writerow({**row, "aliases": "|".join(row.get("aliases") or [])})


@dataclass
class DisclosureScheme:
    amc_id: str
    label: str
    label_norm: str
    base_name: str
    base_name_norm: str
    shortcode: str | None = None
    source: str = "unknown"  # index|sheet|file|parsed
    rows: int | None = None


@dataclass
class MatchRow:
    disclosure: dict[str, Any]
    amfi: dict[str, Any] | None
    score: float
    status: str  # matched|unmatched_disclosure


def load_amc_registry(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    return list(data.get("amcs") or [])


def map_amcs_to_amfi(
    registry: list[dict[str, Any]],
    amfi_amc_names: list[str],
    *,
    cutoff: float = 86.0,
) -> dict[str, dict[str, Any]]:
    """amc_id → {amfi_amc_name, score, registry_name}"""

    def amc_core(s: str) -> str:
        # Drop shared suffixes so IL&FS ≠ NJ (both end with "Mutual Fund")
        t = norm_text(s) or ""
        t = re.sub(
            r"\b(mutual\s+fund|asset\s+management(?:\s+company)?|limited|ltd|pvt|private|company|amc|idf)\b",
            " ",
            t,
        )
        return re.sub(r"\s+", " ", t).strip()

    choices = {amc_core(n): n for n in amfi_amc_names if amc_core(n)}
    choice_keys = list(choices.keys())
    out: dict[str, dict[str, Any]] = {}
    for amc in registry:
        amc_id = amc["id"]
        # Prefer display "name" which usually ends with Mutual Fund
        candidates = [
            amc.get("name") or "",
            amc.get("amc_name") or "",
            amc_id.replace("-", " "),
        ]
        best = None
        for cand in candidates:
            key = amc_core(cand)
            if not key:
                continue
            # exact / contains helpers
            if key in choices:
                best = (100.0, choices[key], cand)
                break
            hit = process.extractOne(key, choice_keys, scorer=fuzz.token_sort_ratio)
            if hit and (best is None or hit[1] > best[0]):
                best = (float(hit[1]), choices[hit[0]], cand)
        if best and best[0] >= cutoff:
            out[amc_id] = {
                "amfi_amc_name": best[1],
                "score": best[0],
                "registry_name": amc.get("name"),
                "matched_via": best[2],
            }
        else:
            out[amc_id] = {
                "amfi_amc_name": None,
                "score": best[0] if best else 0.0,
                "registry_name": amc.get("name"),
                "matched_via": best[2] if best else None,
                "note": "no AMFI AMC match",
            }
    return out


def prepare_workbook(path: Path, tmp: Path) -> Path | None:
    try:
        head = path.read_bytes()[:8]
    except OSError:
        return None
    if head.startswith(b"<") or head.startswith(b"<!DOC") or head.startswith(b"<!doc"):
        return None
    if head.startswith(b"PK") and path.suffix.lower() == ".xls":
        dest = tmp / (path.stem[:100] + ".xlsx")
        if not dest.exists():
            shutil.copyfile(path, dest)
        return dest
    return path


def iter_workbooks(amc_dir: Path, tmp: Path) -> list[Path]:
    files: list[Path] = []
    for p in sorted(amc_dir.iterdir()):
        if not p.is_file():
            continue
        suf = p.suffix.lower()
        if suf in {".xlsx", ".xls", ".xlsm"}:
            files.append(p)
        elif suf == ".zip":
            dest = tmp / "zips" / amc_dir.name / p.stem
            dest.mkdir(parents=True, exist_ok=True)
            try:
                with zipfile.ZipFile(p) as zf:
                    for info in zf.infolist():
                        if info.is_dir():
                            continue
                        name = Path(info.filename).name
                        if re.search(r"\.(xlsx|xls|xlsm)$", name, re.I):
                            t = dest / name
                            if not t.exists():
                                with zf.open(info) as src, t.open("wb") as out:
                                    shutil.copyfileobj(src, out)
                            files.append(t)
            except Exception:
                continue
    return files


def read_sheets(path: Path, tmp: Path, max_rows: int = 120) -> list[tuple[str, list[list[str]]]]:
    work = prepare_workbook(path, tmp)
    if work is None:
        return []
    try:
        head = work.read_bytes()[:4]
        if head.startswith(b"PK") or work.suffix.lower() in {".xlsx", ".xlsm"}:
            wb = load_workbook(work, data_only=True, read_only=True)
            sheets = []
            for ws in wb.worksheets:
                rows = []
                for i, r in enumerate(ws.iter_rows(values_only=True)):
                    if i >= max_rows:
                        break
                    rows.append([("" if c is None else str(c).strip()) for c in (list(r[:12]) if r else [])])
                sheets.append((ws.title or "", rows))
            wb.close()
            return sheets
        import xlrd

        book = xlrd.open_workbook(str(work))
        sheets = []
        for ws in book.sheets():
            rows = [
                [str(ws.cell_value(r, c)).strip() for c in range(min(ws.ncols, 12))]
                for r in range(min(ws.nrows, max_rows))
            ]
            sheets.append((ws.name or "", rows))
        return sheets
    except Exception:
        return []


def parse_index_rows(rows: list[list[str]]) -> list[tuple[str, str]]:
    """Return (code, name) pairs from Index-like tables."""
    hdr = None
    code_j = name_j = short_j = None
    for i, row in enumerate(rows[:40]):
        low = [c.lower().strip() for c in row]
        for cand in ("fund code", "scheme code", "fund id", "scheme id", "code"):
            if cand in low:
                code_j = low.index(cand)
                break
        for cand in ("scheme short code", "short name", "short code", "fund short name"):
            if cand in low:
                short_j = low.index(cand)
                break
        for cand in (
            "scheme name",
            "scheme names",
            "fund name",
            "scheme full name",
            "fund desc",
            "fund description",
            "name of the scheme",
            "name of scheme",
        ):
            if cand in low:
                name_j = low.index(cand)
                break
        # ITI: Short Name + Scheme Name without separate code col
        if name_j is not None and code_j is None and short_j is not None:
            code_j = short_j
        if code_j is not None and name_j is not None:
            hdr = i
            break
    if hdr is None:
        return []
    out: list[tuple[str, str]] = []
    for row in rows[hdr + 1 :]:
        if code_j >= len(row) or name_j >= len(row):
            continue
        code = row[code_j].strip()
        name = re.sub(r"\s+", " ", row[name_j]).strip()
        if short_j is not None and short_j < len(row) and row[short_j].strip():
            short = row[short_j].strip()
        else:
            short = code
        if not code or not name:
            continue
        if code.lower() in {"scheme code", "fund code", "code", "classification", "fund id", "short name"}:
            continue
        if name.lower() in {"scheme name", "fund name", "click", "fund desc", "fund description"}:
            continue
        if re.fullmatch(r"\d{2,6}", code) and short and short != code:
            out.append((short, name))
        else:
            out.append((code, name))
    return out


def strip_scheme_description(label: str) -> str:
    """Remove trailing scheme-category / formerly-known / statement boilerplate."""
    s = re.sub(r"\s+", " ", (label or "")).strip()
    # "MONTHLY PORTFOLIO STATEMENT OF X AS ON …" → X
    # Kotak: "Portfolio of X as on 31 - Jul - 2026" (no "statement")
    m = re.match(
        r"(?i)^(?:monthly\s+|fortnightly\s+)?portfolio(?:\s+statement)?\s+of\s+(.+?)(?:\s+as\s+on\b.*)?$",
        s,
    )
    if m:
        s = m.group(1).strip(" -")
    else:
        # Bare trailing as-of dates left on otherwise clean titles
        s = re.sub(
            r"(?i)\s+as\s+on\s+\d{1,2}\s*[-/]\s*[A-Za-z]{3,9}\s*[-/]\s*\d{2,4}\s*$",
            "",
            s,
        ).strip()
        s = re.sub(
            r"(?i)\s+as\s+on\s+\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\s*$",
            "",
            s,
        ).strip()
    # Zerodha-style trailing period: "… FOR JUNE 2026" / "… JUNE 2026"
    s = re.sub(
        r"(?i)\s+(?:for\s+)?(?:january|february|march|april|may|june|july|august|september|october|november|december)\s+20\d{2}\s*$",
        "",
        s,
    ).strip()
    # dash then open-ended / risk text (360 ONE style)
    s = re.sub(
        r"(?i)\s*[-–—]\s*an?\s+open[-\s]?ended\b.*$",
        "",
        s,
    ).strip()
    s = re.sub(r"(?i)\s*[-–—]\s*a\s+relatively\b.*$", "", s).strip()
    # formerly / erstwhile known (often truncated / misspelled "Know as")
    s = re.sub(
        r"(?i)\s*\(?\s*(?:formerly|erstwhile)\s+know(?:n)?(?:\s+as)?\b.*$",
        "",
        s,
    ).strip()
    s = re.sub(r"(?i)\s*\(?\s*erstwhile\b.*$", "", s).strip()
    # Groww-style shortcode prefixes: "IB01 - Groww Large Cap Fund"
    s = re.sub(r"(?i)^[A-Z]{1,4}\d{0,4}\s*[-–—:]\s*", "", s).strip()
    # truncated ticker tails: "(MOFUSTF" / "(anything without closing paren)
    s = re.sub(r"\s*\([A-Z0-9][^)]*$", "", s).strip()
    # Tata-style: "(The scheme has 1 segregated portfolio…)"
    s = re.sub(r"(?i)\s*\(\s*the\s+scheme\s+has\b.*$", "", s).strip()
    # peel nested/trailing open-ended category parentheses
    for _ in range(3):
        nxt = re.sub(
            r"\s*\(+[^)]*(?:open[-\s]?ended|close[-\s]?ended|exchange\s+traded)[^)]*\)+\s*$",
            "",
            s,
            flags=re.I,
        ).strip()
        nxt2 = re.sub(
            r"\s*\(+[^)]*(?:open[-\s]?ended|close[-\s]?ended).*$",
            "",
            nxt,
            flags=re.I,
        ).strip()
        if nxt2 == s:
            break
        s = nxt2
    s = re.sub(r"\s*\(\s*An?\s+[^)]{8,}\)\s*$", "", s, flags=re.I).strip()
    # Acronym expansions: "(The Infrastructure Growth and Economic Reforms Fund)"
    s = re.sub(r"\s*\(\s*The\s+[^)]{8,}\)\s*$", "", s, flags=re.I).strip()
    # Tenure tails: "(1879 DAYS)" / "(3652 days)" — keep series letter, drop day count paren
    s = re.sub(r"\s*\(\s*\d{2,5}\s*days?\s*\)\s*$", "", s, flags=re.I).strip()
    s = re.sub(r"\s*\(\s*A\s+\d+[^)]*\)\s*$", "", s, flags=re.I).strip()
    # Bare trailing day counts without paren (after other peels)
    s = re.sub(r"(?i)\s+\d{2,5}\s*days?\s*$", "", s).strip()
    if re.match(r"(?i)^name of mutual fund\s*:", s):
        return ""
    if re.fullmatch(r"(?i)scheme\s*name\s*:?", s):
        return ""
    return s


def looks_like_amc_header(label: str) -> bool:
    s = re.sub(r"\s+", " ", (label or "")).strip()
    if re.fullmatch(r"(?i).+\s+mutual\s+fund", s):
        return True
    if re.match(r"(?i)^name of mutual fund\b", s):
        return True
    return False


def is_junk_label(label: str) -> bool:
    s = (label or "").strip()
    if not s or len(s) < 2:
        return True
    if looks_like_amc_header(s):
        return True
    if JUNK_SCHEME_RE.search(s):
        return True
    if JUNK_CONTAINS_RE.search(s):
        return True
    if re.fullmatch(r"(?i)scheme\s*name\s*:?", s):
        return True
    if re.fullmatch(r"IN[A-Z0-9]{10}", s):  # ISIN / G-sec codes
        return True
    if re.fullmatch(r"(?i)(?:domestic\s+)?mutual\s+fund\s+units?", s):
        return True
    if re.fullmatch(r"(?i)(?:back\s+to\s+index|sovereign|nimf)", s):
        return True
    if re.match(r"(?i)^foreign\s+securities\b", s):
        return True
    if re.fullmatch(r"(?i)(?:fund\s*code|fund\s*name|scheme\s*code\d*starts?|moderate|high|low|construction)", s):
        return True
    if re.match(r"(?i)^derivative(?:s)?(?:\s*disclosure)?\b", s):
        return True
    if re.match(r"(?i)^aggregate\s+value\s+of\s+investments\b", s):
        return True
    # Industry / sector classification rows (not scheme titles)
    if re.fullmatch(
        r"(?i)(?:construction|banks| pharmaceuticals|it\s*-\s*software|automobiles|"
        r"finance|telecommunications|power|oil|gas|chemicals|cement|metals)",
        s,
    ):
        return True
    if re.search(r"\b(ltd|limited|reit|corporation|industries|bank)\b", s, re.I) and not re.search(
        r"\b(fund|etf|fof|scheme|plan|index|nifty|sensex|portfolio|income|segregated)\b", s, re.I
    ):
        return True
    if s.lower() in {"underlying", "long", "scheme", "industry", "quantity", "isin"}:
        return True
    return False


def looks_like_fund_name(label: str) -> bool:
    s = (label or "").strip()
    if not s:
        return False
    if looks_like_amc_header(s):
        return False
    if re.match(r"(?i)^nse\s*symbol\b|^bse\s*scrip\b", s):
        return False
    if is_shortcode(s):
        return True
    # Navi / index names sometimes omit the word "Fund"
    if re.search(
        r"\b(fund|etf|fof|index|scheme|plan|momentum|quality|elss|fmp|tax\s*saver|"
        r"infrastructure|hybrid|flexi\s*cap|large\s*cap|mid\s*cap|small\s*cap|"
        r"multi\s*cap|focused|consumption|arbitrage|overnight|liquid|gilt|"
        r"income|accrual|duration|bond|debt|segregated|credit\s*risk)\b",
        s,
        re.I,
    ):
        return True
    # Bare AMC scheme titles like "CANARA ROBECO INFRASTRUCTURE"
    if re.match(r"(?i)^(?:canara\s+robeco|hdfc|uti|sbi|icici|nippon|edelweiss|mirae)\b", s) and len(s) >= 12:
        return True
    # Nifty/Sensex product names (but not "NSE Symbol" / "BSE Scrip Code" lines)
    if re.search(r"\b(nifty|sensex)\b", s, re.I) and not re.search(r"(?i)symbol|scrip\s*code", s):
        return True
    return False


def extract_sheet_title(rows: list[list[str]], sheet_name: str) -> str | None:
    """Best-effort scheme title from portfolio sheet preamble."""
    for r in rows[:15]:
        for j, c in enumerate(r):
            # PORTFOLIO STATEMENT OF {FUND} AS ON …
            # Kotak: PORTFOLIO OF {FUND} AS ON …
            m_ps = re.match(
                r"(?i)^(?:monthly\s+|fortnightly\s+)?portfolio(?:\s+statement)?\s+of\s+(.+?)(?:\s+as\s+on\b.*)?$",
                c.strip(),
            )
            if m_ps:
                title = strip_scheme_description(m_ps.group(1))
                if title and looks_like_fund_name(title):
                    return title
            m = re.match(r"(?i)^scheme\s*name\s*:?\s*(.*)$", c.strip())
            if m:
                title = strip_scheme_description(m.group(1))
                if title and looks_like_fund_name(title):
                    return title
                if j + 1 < len(r) and r[j + 1].strip():
                    title = strip_scheme_description(r[j + 1].split("\n")[0])
                    if title and looks_like_fund_name(title):
                        return title
            cl = re.sub(r"\s+", " ", c.lower()).strip().rstrip(":").strip()
            if cl in {"scheme name", "name of the scheme", "name of scheme"}:
                if j + 1 < len(r) and r[j + 1].strip():
                    title = strip_scheme_description(r[j + 1].split("\n")[0])
                    if title and looks_like_fund_name(title):
                        return title
    candidates: list[str] = []
    for r in rows[:12]:
        for c in r:
            # HSBC merges AMC + scheme with newlines in one cell
            parts = [p.strip() for p in re.split(r"[\n\r]+", c) if p.strip()]
            for part in parts or [c]:
                if len(part) < 8:
                    continue
                low = part.lower()
                if any(
                    x in low
                    for x in (
                        "registered office",
                        "asset management",
                        "cin:",
                        "toll free",
                        "risk-o-meter",
                        "product labelling",
                        "name of mutual fund",
                        "name of the instrument",
                        "inception",
                    )
                ):
                    continue
                if looks_like_amc_header(part):
                    continue
                if "portfolio statement" in low and not re.search(r"(?i)portfolio\s+statement\s+of\b", low):
                    continue
                title = strip_scheme_description(part)
                if title and looks_like_fund_name(title) and not is_junk_label(title):
                    candidates.append(title)
    return candidates[0] if candidates else None


def fund_name_from_filename(stem: str) -> str | None:
    """Helios/Union/Samco filenames → clean fund names."""
    s = (stem or "").replace("_", " ").replace("-", " ")
    s = re.sub(r"(?i)\bmonthly[-\s]*portfolio[-\s]*report\b", " ", s)
    s = re.sub(
        r"(?i)\b(monthly|portfolio|disclosure|statement|as\s*on|as\s*of|dated|report)\b",
        " ",
        s,
    )
    s = re.sub(r"(?i)\b\d{1,2}(st|nd|rd|th)?\b", " ", s)
    s = re.sub(
        r"(?i)\b(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|"
        r"jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\b",
        " ",
        s,
    )
    s = re.sub(r"\b20\d{2}\b", " ", s)
    s = re.sub(r"\b[a-f0-9]{8,}\b", " ", s)
    # camelCase → spaced (SamcoMultiCapFund)
    s = re.sub(r"([a-z])([A-Z])", r"\1 \2", s)
    s = re.sub(r"\s+", " ", s).strip(" .")
    if looks_like_fund_name(s) and not is_junk_label(s):
        return s
    return None


def is_secondary_amfi_fund(name: str) -> bool:
    """Unclaimed / Principal Units / Defunct / side plans — not expected in monthly packs.

    Also covers AMFI NAVAll quirks where IDCW tails like "Payout & Reinvestment"
    survived base_fund collapse and show up as separate "funds".
    """
    return bool(
        re.search(
            r"(?i)\bunclaimed\b|\bprincipal\s+units?\b|\bsuper\s+institutional\b|"
            r"\binvestor\s+education\b|\bie\.?f\.?\b|(?<![a-z])premium(?![a-z])|"
            r"\bdefunct\b|\babove\s*3\s*yrs?\b|\bupto\s*3\s*yrs?\b|\bup\s*to\s*3\s*yrs?\b|"
            r"\bc\s+redemption\b|\bfortnigh?tly\s+dividend\b|"
            r"\bfmp\s+series\b|\bfixed\s+maturity\b|\bfmp\b|\bfixed\s+term\s+series\b|"
            # Kotak/AMFI IDCW tails & truncated payout labels
            r"\bpayout\s*(?:and|&|/)?\s*re(?:invest(?:ment)?)?\b|"
            r"\bpayout\s*(?:and|&|/)?\s*investment\b|"
            r"\bpayout\s+of\b|\bpayouout\b|"
            r"\bnormal\s+dividend\b|"
            r"\binvestment\s+provident\s+fund\b|"
            r"\s[-–—]\s*standard\s*$|"
            # SBI / Nippon / UTI PF, segregated, discontinued, legacy
            r"\bsegregated\b|"
            r"\bdiscontinued\b|"
            r"\bpf\s*\(|\b-\s*p\s*f\s*-|\bp\s*f\s*-\s*(?:fixed|automatic|defined)|"
            r"\bfixed\s+period\b|"
            r"\bincome\s+cum\s+distribution\b|"
            r"\bshort\s+horizon\s+debt\b|"
            r"\binterval\s+fund\b|"
            r"\shybrd\b|\bpayment\s*$|"
            r"\s[-–—]\s*(?:gr(?:owth)?|cumulative|periodic)\s*$|"
            r"\bfof\s*[-–—]\s*[a-z]\b|"
            r"\bgreater\s+than\s*3\s*years?\b|"
            r"\bno\s+lock[\s\-]?in\b|"
            r"\babove\s*3\s*yrs?\b|"
            r"\bc\s*[-–—]\s*greater\b|"
            # Dirty/truncated AMFI NAVAll duplicates of an otherwise-primary fund
            r"\bformerly\s+know(?:n)?\b|"
            r"\berstwhile\b|"
            r"\([A-Z0-9][^)]*$",
            name or "",
        )
    )


DEBT_LIKE_RE = re.compile(
    r"(?i)\b(?:liquid|overnight|money\s*market|ultra\s*short|low\s*duration|"
    r"short\s*(?:term|duration|tem)|medium\s*(?:term|duration)|long\s*duration|"
    r"dynamic\s*(?:bond|accrual)|corporate\s*bond|credit\s*risk|banking\s*(?:and|&)?\s*psu|"
    r"gilt|g[\s-]?sec|floater|floating\s*rate|debt|income|bond|accrual|arbitrage|savings|"
    r"segregated\s*portfolio|"
    r"conservative\s*hybrid|equity\s*savings|fixed\s*(?:maturity|term)|fmp|"
    r"target\s*maturity|crisil\s*ibx|nifty.*(?:sdl|gilt|bond)|"
    r"retirement.*debt|children.*debt)\b"
)


def is_debt_like_amfi_fund(name: str) -> bool:
    """Fortnightly SEBI disclosures cover debt / money-market style schemes."""
    return bool(DEBT_LIKE_RE.search(name or ""))


def compact_name(s: str) -> str:
    """Space-insensitive root key so Midcap ≡ Mid Cap."""
    t = norm_text(s) or ""
    # AMFI typo / rename normalizations
    t = t.replace("glit", "gilt")
    t = t.replace("reality", "realty")
    t = t.replace("financials", "financial")
    t = t.replace("consumer trends", "consumption")
    t = t.replace("hybrd", "hybrid")
    t = t.replace("hyrbrid", "hybrid")
    t = t.replace("govenment", "government")
    t = t.replace("duratio ", "duration ")
    t = t.replace("duratiofund", "durationfund")  # after space strip path: apply before
    t = t.replace("owsal", "oswal")
    t = t.replace("midsmall ", "midsmallcap ")
    t = t.replace("largecap", "large cap")
    t = t.replace("smallcap", "small cap")
    t = t.replace("multicap", "multi cap")
    t = t.replace("midcap", "mid cap")
    t = t.replace("flexicap", "flexi cap")
    t = t.replace("emerging market fund", "emerging markets fund")
    t = t.replace("fund of funds", "fund of fund")
    t = t.replace("fund of fund", "fof")
    t = t.replace("asia pacific yield", "asia pacific ex japan dividend yield")
    # AMFI typos / SEBI rename aliases
    t = t.replace("dynamic term fund", "dynamic bond fund")
    t = t.replace("ultra short term fund", "ultra short duration fund")
    t = t.replace("omni fund of fund", "active fund of fund")
    t = t.replace("payouout", "payout")  # Kotak NAVAll typo
    # Nippon duplicate "Interval Fund Interval Fund Series II"
    t = t.replace("interval fund interval fund", "interval fund")
    # Edelweiss BHARAT: AMFI says "Bond ETF FOF", packs say "Bond FOF"
    t = t.replace("bharat bond etf fof", "bharat bond fof")
    t = re.sub(r"\bmmf\b", "money market fund", t)
    t = re.sub(r"\s+", " ", t).strip()
    # Duratio without trailing space (end of string)
    t = re.sub(r"duratio$", "duration", t)
    t = t.replace(" ", "")
    # post-compact fixes
    t = t.replace("duratio", "duration")
    t = t.replace("midsmallcapcap", "midsmallcap")
    t = t.replace("intervalfundintervalfund", "intervalfund")
    t = t.replace("smallcap250", "smallcap250")
    return t


def expand_disclosure_abbrev(label: str) -> str:
    """Expand common AMC short labels (Edelweiss index packs etc.)."""
    s = re.sub(r"\s+", " ", (label or "")).strip()
    if not s:
        return s
    reps = [
        (r"(?i)\bEdel\b", "Edelweiss"),
        (r"(?i)\bNY\b", "Nifty"),
        (r"(?i)\bLMcap", "LargeMidcap"),
        (r"(?i)\bG[\s\-]?S\b", "G-Sec"),
        (r"(?i)\bIDX\b", "Index Fund"),
        (r"(?i)\bPl\b", "Plus"),
        (r"(?i)\bWD\b", "World"),
        (r"(?i)\bDM\b", "Domestic"),
        (r"(?i)\bHC\b", "Healthcare"),
        (r"(?i)\bID\s+Fund\b", "Index Fund"),
        (r"(?i)\(\s*I\s*\)", "India"),
        (r"(?i)\bMMF\b", "Money Market Fund"),
        (r"(?i)\bLargeMidcap\s*250\b", "LargeMidcap250"),
        (r"(?i)\bfixed\s+term\s+plan\b", "Fixed Term"),
        (r"(?i)\bunit\s+linked\s+insurance\s+scheme\b", "ULIS"),
        (r"(?i)\bQlty\b", "Quality"),
        (r"(?i)\bquarterly\s+interval\s+fund\b", "Interval Fund"),
        (r"(?i)\bseries\s*[-–—]?\s*2\b", "Series II"),
        (r"(?i)\bseries\s*[-–—]?\s*ii\b", "Series II"),
        (r"(?i)\bfund\s+of\s+funds?\s*[-–—:]?\s*conservative\b", "FOF - C"),
        (r"(?i)\bfund\s+of\s+funds?\s*[-–—:]?\s*aggressive\b", "FOF - Aggressive"),
        (r"(?i)\bmulti\s*[- ]?\s*asset\s+allocation\s+fund\b", "Multi Asset Fund"),
        (r"(?i)\bsmallcap\s*250\b", "Smallcap 250"),
        (r"(?i)\bmomentum\s+quality\s*(?:100|index)?\b", "Momentum quality index"),
    ]
    for pat, repl in reps:
        s = re.sub(pat, repl, s)
    return re.sub(r"\s+", " ", s).strip()


YEAR_TOKEN_RE = re.compile(r"(?:19|20)\d{2}")
MATURITY_IN_NAME_RE = re.compile(
    r"(?i)\b((?:jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|"
    r"jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)"
    r"\s*-?\s*(?:19|20)\d{2})\b"
)


def year_tokens(s: str) -> set[str]:
    return set(YEAR_TOKEN_RE.findall(s or ""))


def year_conflict(a: str, b: str) -> bool:
    """True when both sides cite maturity/calendar years and they disagree."""
    ya, yb = year_tokens(a), year_tokens(b)
    return bool(ya and yb and ya.isdisjoint(yb))


def structure_conflict(a: str, b: str) -> bool:
    """FOF vs ETF-only (or ETF bit mismatch) — BHARAT Bond FOF must not match Bond ETF."""
    ca, cb = compact_name(a), compact_name(b)
    if not ca or not cb:
        return False
    a_fof = "fof" in ca or "fundoffund" in ca
    b_fof = "fof" in cb or "fundoffund" in cb
    a_etf = "etf" in ca
    b_etf = "etf" in cb
    if a_fof != b_fof:
        return True
    if a_etf != b_etf and not (a_fof and b_fof):
        return True
    return False


def containment_score(disc: str, amfi: str) -> float:
    """Disclosure root is present in AMFI root (after compacting). Midcap ⊂ Mid Cap → match.

    Also accepts the reverse when a longer disclosure title embeds the AMFI root
    (e.g. ``T.I.G.E.R. Fund (The Infrastructure Growth…)`` ⊃ ``T.I.G.E.R. Fund``).
    """
    if not disc or not amfi:
        return 0.0
    if is_secondary_amfi_fund(amfi) and not is_secondary_amfi_fund(disc):
        return 0.0
    if year_conflict(disc, amfi) or structure_conflict(disc, amfi):
        return 0.0
    dc, ac = compact_name(disc), compact_name(amfi)
    if not dc or not ac:
        return 0.0
    if dc == ac:
        return 100.0
    if dc in ac:
        # Exact root → 100; proper substring → prefer tighter (shorter) AMFI names via score
        # Penalize long leftovers so "X Fund" doesn't prefer "X Fund Segregated …"
        extra = len(ac) - len(dc)
        return max(90.0, 99.0 - min(extra, 20) * 0.3)
    if ac in dc and len(ac) >= 10:
        extra = len(dc) - len(ac)
        # Require leftovers not introduce a conflicting second product token
        return max(90.0, 99.0 - min(extra, 40) * 0.2)
    return 0.0


def name_match_score(a: str, b: str) -> float:
    """Prefer containment of disclosure root in AMFI root; fuzzy only as fallback."""
    if not a or not b:
        return 0.0
    if year_conflict(a, b) or structure_conflict(a, b):
        return 0.0
    contained = containment_score(a, b)
    if contained >= 90.0:
        return contained
    b_pen = 0.0
    if "unclaimed" in b and "unclaimed" not in a:
        b_pen += 12.0
    if is_secondary_amfi_fund(b) and not is_secondary_amfi_fund(a):
        b_pen += 8.0
    sort_s = float(fuzz.token_sort_ratio(a, b))
    set_s = float(fuzz.token_set_ratio(a, b))
    partial = float(fuzz.partial_ratio(a, b))
    la, lb = len(a), len(b)
    ratio = min(la, lb) / max(la, lb)
    raw = 0.55 * sort_s + 0.25 * set_s + 0.20 * partial
    if ratio < 0.55 and set_s > sort_s + 15:
        raw *= 0.75
    return max(0.0, raw - b_pen)


def extract_disclosure_schemes(
    amc_id: str,
    disclosure_dir: Path,
    parsed_dir: Path,
    tmp: Path,
    *,
    disclosure_type: str = "monthly",
) -> list[DisclosureScheme]:
    found: dict[str, DisclosureScheme] = {}

    def add(label: str, *, shortcode: str | None, source: str, rows: int | None = None) -> None:
        # Disclosure packs are never duplicates — do NOT peel Growth/IDCW/Direct/Regular
        # via base_fund_name (that collapses Mirae/SBI/ICICI liquid-ETF series into one).
        label = strip_scheme_description(re.sub(r"\s+", " ", (label or "")).strip())
        label = expand_disclosure_abbrev(label)
        if not label:
            return
        if is_junk_label(label) and not shortcode:
            return
        # Identity = cleaned label as published on the pack (keep Growth vs IDCW etc.)
        base = label
        if is_junk_label(base) and not (shortcode and is_shortcode(shortcode)):
            return
        key = norm_text(base) or (shortcode or "").lower()
        if not key:
            return
        # Prefer index sources over parsed junk; never overwrite a distinct prior pack
        prev = found.get(key)
        rank = {"index": 3, "sheet": 2, "file": 2, "parsed": 1}.get(source, 0)
        prev_rank = {"index": 3, "sheet": 2, "file": 2, "parsed": 1}.get(prev.source, 0) if prev else -1
        if prev and prev_rank > rank:
            return
        found[key] = DisclosureScheme(
            amc_id=amc_id,
            label=label,
            label_norm=norm_text(label),
            base_name=base,
            base_name_norm=norm_text(base),
            shortcode=shortcode,
            source=source,
            rows=rows,
        )

    # 1) Index / sheet titles from workbooks
    if disclosure_dir.is_dir():
        for path in iter_workbooks(disclosure_dir, tmp):
            # When matching monthly, skip fortnightly/mid-month packs.
            # When matching fortnightly, keep those (and still allow month-end debt).
            if disclosure_type == "monthly" and re.search(
                r"(?i)fortnight|june[_\s\-]*15|15[_\s\-]*june|mid[_\s\-]?month",
                path.name,
            ):
                continue
            sheets = read_sheets(path, tmp)
            if not sheets:
                continue
            multi = len([n for n, _ in sheets if not is_junk_label(n)]) > 1
            file_title = fund_name_from_filename(path.stem)
            # Filename pack/ticker is the durable monthly identity for single-fund packs.
            file_pack = pack_code_from_filename(path.stem) if not multi else None
            for name, rows in sheets:
                if name.strip().lower() == "index" or any(
                    "scheme code" in " ".join(r).lower() or "fund code" in " ".join(r).lower() for r in rows[:15]
                ):
                    for code, scheme_name in parse_index_rows(rows):
                        add(scheme_name, shortcode=normalize_shortcode(code) or code, source="index")
                if is_junk_label(name):
                    continue
                title = extract_sheet_title(rows, name)
                # When the sheet tab / title is only a ticker (NBANKETF) but the
                # filename carries the full fund name, prefer the filename title.
                if title and file_title and (
                    is_shortcode(title.strip()) or compact_name(title) == compact_name(name)
                ):
                    if looks_like_fund_name(file_title) and not is_shortcode(file_title):
                        title = file_title
                # Durable key: pack ticker, else sheet name as-is, else scheme title
                # when the tab is junk (e.g. Navi "Sheet1").
                sheet_key = durable_disclosure_key(name)
                title_key = durable_disclosure_key(title) if title else None
                short = file_pack or sheet_key or title_key
                if title and not is_junk_label(title):
                    add(
                        title,
                        shortcode=short,
                        source=(
                            "file"
                            if file_pack or (file_title and title == file_title)
                            else ("sheet" if sheet_key else ("title" if title_key else "sheet"))
                        ),
                    )
                elif file_title:
                    add(
                        file_title,
                        shortcode=short or durable_disclosure_key(file_title),
                        source="file",
                    )
                elif multi and short:
                    # last resort: shortcode alone (weak)
                    pass
                elif not multi and file_title:
                    add(
                        file_title,
                        shortcode=short or durable_disclosure_key(file_title),
                        source="file",
                    )
                elif not multi and short and not file_title:
                    add(short, shortcode=short, source="file")
                elif multi and sheet_key and not title:
                    # Consolidated book: sheet title absent but sheet name identifies the fund
                    add(name.strip(), shortcode=sheet_key, source="sheet")

    # 2) parsed schemes.json ONLY if we found almost nothing from workbooks
    if len(found) < 2:
        schemes_path = parsed_dir / "schemes.json"
        if schemes_path.exists():
            try:
                parsed_rows = json.loads(schemes_path.read_text(encoding="utf-8"))
            except Exception:
                parsed_rows = []
            for row in parsed_rows:
                label = row.get("scheme") or row.get("folder") or ""
                if is_junk_label(label) or not looks_like_fund_name(label):
                    continue
                if re.search(r"\d{2}\.\d{2}\.\d{2}|^\d+\s", label):
                    continue
                if re.search(r"\b\d+(\.\d+)?%\b", label) and "fund" not in label.lower():
                    continue
                sc = label.strip() if is_shortcode(label.strip()) else None
                add(label, shortcode=sc, source="parsed", rows=row.get("rows"))

    return list(found.values())


VARIANT_TAIL_RE = re.compile(
    r"(?ix)\s*[-–—(:]?\s*(?:"
    r"unclaimed\b.*|"
    r"cumulative\b.*|"
    r"retail\b.*|"
    r"(?:super\s+)?institutional\b.*|"
    r"defunct\b.*|"
    r"discontinued\b.*|"
    r"prescribed\s+date\b.*|"
    r"segregated\b.*|"
    r"principal\s+units?\b.*|"
    r"\beco\b.*|"
    # trailing IDCW/Dividend *plan* labels only — not product names like
    # "BSE 500 Dividend Leaders 50 ETF"
    r"dividends?\s*(?:option|plan|payout|reinvest(?:ment)?)?\s*$|"
    r"payout\b.*|"
    r"payouout\b.*|"  # Kotak NAVAll typo twin of payout
    r"reinvest(?:ment)?\b.*|"
    r"formerly\b.*|"
    r"erstwhile\b.*|"
    r"normal\s+dividend\b.*|"
    r"standard\b.*|"
    r"payment\b.*|"
    r"no\s+lock[\s\-]?in\b.*|"
    r"greater\s+than\s*3\b.*|"
    r"above\s*3\s*yrs?\b.*|"
    r"upto\s*3\s*yrs?\b.*|"
    r"up\s*to\s*3\s*yrs?\b.*|"
    r"below\s*3\b.*|"
    r"icdw\b.*|"
    r"income\s+cum\s+distribution\b.*|"
    # NOTE: do NOT peel bare idcw/growth here — listed liquid-ETF Growth vs IDCW
    # series are distinct AMFI products (e.g. Mirae 153045 vs 151908).
    r"idwc\b.*|"
    r"regulats?\b.*|"
    r"\bbi\b.*|"
    r"\bhalf\b.*|"
    r"periodics?\b.*|"
    r"\bp\s*f\b.*|"
    r"provident\b.*|"
    r"investment\s+provident\b.*"
    r")$"
)

CLOSED_END_RE = re.compile(
    r"(?i)\bfmp\b|\bfixed\s+maturity\b|\bfixed\s+term\s+series\b"
)


def strip_variant_tails(name: str) -> str:
    """Peel NAVAll plan/side labels so Unclaimed/Retail/etc. collapse to primary."""
    s = strip_scheme_description(name or "")
    for _ in range(4):
        nxt = VARIANT_TAIL_RE.sub("", s).strip(" -–—:(")
        if nxt == s:
            break
        s = nxt
    return s


INVESTED_IN_RE = re.compile(
    r"(?i)(?:\(|\b)invested\s+in\s+(.+?)(?:\s*\)\s*)?$"
)


def invested_in_fund(name: str) -> str | None:
    """AMFI Unclaimed rows often read '... (Invested in DSP Overnight Fund'."""
    m = INVESTED_IN_RE.search((name or "").strip())
    if not m:
        return None
    host = strip_scheme_description(m.group(1)).strip(" -–—:(")
    return host or None


def looks_like_closed_end_fmp(name: str) -> bool:
    return bool(CLOSED_END_RE.search(name or ""))


def amfi_match_key(f: dict[str, Any]) -> str:
    """Cleaned matching key for an AMFI fund row (shared by variants)."""
    raw = f.get("base_name") or ""
    cleaned = strip_scheme_description(raw) or raw
    cleaned = strip_variant_tails(cleaned) or cleaned
    canon = f.get("canonical_name") or ""
    if not year_tokens(cleaned):
        m = MATURITY_IN_NAME_RE.search(canon)
        if m:
            cleaned = f"{cleaned} - {m.group(1)}"
    if canon:
        peeled = re.sub(
            r"(?i)\s*[-–—]?\s*(?:direct|regular)\b.*$",
            "",
            canon,
        ).strip()
        peeled = re.sub(
            r"(?i)\s*[-–—]?\s*(?:growth|idcw|dividend|payout|reinvestment)\b.*$",
            "",
            peeled,
        ).strip()
        peeled = strip_scheme_description(peeled)
        peeled = strip_variant_tails(peeled) or peeled
        cc, cp = compact_name(cleaned), compact_name(peeled)
        if cp and cc and cp.startswith(cc) and len(cp) >= len(cc) + 3:
            cleaned = peeled
    return norm_text(cleaned) or f.get("base_name_norm") or ""


def match_within_amc(
    disclosures: list[DisclosureScheme],
    amfi_funds: list[dict[str, Any]],
    *,
    cutoff: float = 84.0,
    match_candidates: list[dict[str, Any]] | None = None,
    shortcode_lookup: dict[str, str] | None = None,
) -> tuple[list[MatchRow], list[dict[str, Any]], list[dict[str, Any]]]:
    """Match disclosures to AMFI; return (rows, unmatched_amfi, matched_amfi_meta).

    ``amfi_funds`` is the full AMC catalog (always listed in outputs).
    ``match_candidates`` optionally restricts which AMFI names disclosures can
    bind to directly (e.g. debt-like only on fortnightly), without shrinking
    the output universe.
    ``shortcode_lookup`` maps normalized SHORTCODE → canonical_amfi_code for
    this AMC; checked before fuzzy name matching.
    """
    fund_by_norm: dict[str, dict[str, Any]] = {
        f["base_name_norm"]: f for f in amfi_funds if f.get("base_name_norm")
    }
    fund_by_code: dict[str, dict[str, Any]] = {}
    for f in amfi_funds:
        code = str(f.get("canonical_amfi_code") or "").strip()
        if code and code not in fund_by_code:
            fund_by_code[code] = f
        for c in f.get("amfi_codes") or []:
            cs = str(c).strip()
            if cs and cs not in fund_by_code:
                fund_by_code[cs] = f
    fund_match_key: dict[str, str] = {
        f["base_name_norm"]: amfi_match_key(f) for f in amfi_funds if f.get("base_name_norm")
    }

    candidates = match_candidates if match_candidates is not None else amfi_funds
    amfi_by_key: dict[str, dict[str, Any]] = {}
    keys: list[str] = []
    for f in candidates:
        k = fund_match_key.get(f["base_name_norm"]) or amfi_match_key(f)
        prev = amfi_by_key.get(k)
        if prev is None or len(f.get("base_name") or "") < len(prev.get("base_name") or ""):
            amfi_by_key[k] = f
        keys.append(k)
    keys = list(dict.fromkeys(keys))

    matched_amfi: set[str] = set()  # base_name_norm
    matched_meta: dict[str, dict[str, Any]] = {}  # base_name_norm -> meta
    matched_keys: set[str] = set()
    rows: list[MatchRow] = []
    sc_lookup = shortcode_lookup or {}

    def mark_matched(norm: str, *, via: str, score: float, sibling: str | None = None) -> None:
        if not norm or norm in matched_amfi:
            # Prefer first non-empty sibling annotation
            if norm in matched_meta and sibling and not matched_meta[norm].get("sibling"):
                matched_meta[norm]["sibling"] = sibling
            return
        matched_amfi.add(norm)
        fund = fund_by_norm.get(norm) or {}
        matched_meta[norm] = {
            "base_name": fund.get("base_name"),
            "canonical_amfi_code": fund.get("canonical_amfi_code"),
            "plan_count": fund.get("plan_count"),
            "via": via,
            "score": score,
            "sibling": sibling,
            "variant": is_secondary_amfi_fund(fund.get("base_name") or ""),
        }

    for d in disclosures:
        sc_raw = (d.shortcode or "").strip()
        sc = sc_raw
        if sc and " " not in sc and is_shortcode(sc):
            sc = normalize_shortcode(sc) or sc
        fund = None
        if sc and sc in sc_lookup:
            fund = fund_by_code.get(str(sc_lookup[sc]))
        elif sc_raw and sc_raw.casefold() in sc_lookup:
            fund = fund_by_code.get(str(sc_lookup[sc_raw.casefold()]))
        if fund:
            mark_matched(fund["base_name_norm"], via="shortcode", score=100.0)
            matched_keys.add(fund_match_key.get(fund["base_name_norm"]) or amfi_match_key(fund))
            disc = asdict(d)
            disc["shortcode"] = sc_raw or sc
            disc["match_via"] = "shortcode"
            rows.append(MatchRow(disc, fund, 100.0, "matched"))
            continue

        # Match keys are variant-stripped (Discontinued / Segregated / Retail…).
        # Normalize disclosure labels the same way so "…Segregated Portfolio 3 –
        # Yes Bank…" binds to the AMFI segregat row whose key peeled to the root.
        query = (
            norm_text(strip_variant_tails(d.base_name or d.label or ""))
            or d.base_name_norm
            or d.label_norm
        )
        if not query or not keys:
            rows.append(MatchRow(asdict(d), None, 0.0, "unmatched_disclosure"))
            continue

        contained: list[tuple[float, str]] = []
        for k in keys:
            score = containment_score(query, k)
            if score >= 90.0:
                contained.append((score, k))
        if contained:
            def _rank(item: tuple[float, str]) -> tuple:
                score, key = item
                etf_pen = 1 if ("etf" in key and "etf" not in query) else 0
                return (etf_pen, -score, -len(compact_name(key)), key)

            contained.sort(key=_rank)
            best_key, best_score = contained[0][1], contained[0][0]
        else:
            ranked = process.extract(
                query,
                keys,
                scorer=lambda a, b, **kwargs: name_match_score(a, b),
                limit=3,
            )
            best_key, best_score = (ranked[0][0], float(ranked[0][1])) if ranked else (None, 0.0)

        fund = amfi_by_key.get(best_key) if best_key else None
        if fund and best_score >= cutoff:
            mark_matched(fund["base_name_norm"], via="direct", score=best_score)
            matched_keys.add(best_key)
            disc = asdict(d)
            disc["match_via"] = "direct"
            rows.append(MatchRow(disc, fund, best_score, "matched"))
        else:
            rows.append(MatchRow(asdict(d), fund, best_score, "unmatched_disclosure"))

    # Shared cleaned keys (dirty typo / formerly-known duplicates of a direct match).
    for f in amfi_funds:
        mk = fund_match_key.get(f["base_name_norm"]) or ""
        if mk and mk in matched_keys:
            mark_matched(
                f["base_name_norm"],
                via="key_share",
                score=100.0,
                sibling=(amfi_by_key.get(mk) or {}).get("base_name"),
            )

    def _compact_fund_soft(s: str) -> str:
        """Treat optional trailing 'Fund' as soft: Kotak Bond Short Term ≡ … Fund."""
        c = compact_name(s)
        return c[:-4] if c.endswith("fund") else c

    # Compact / containment near-duplicates of already-matched funds.
    if matched_amfi:
        matched_list = list(matched_amfi)
        for f in amfi_funds:
            k = f["base_name_norm"]
            if not k or k in matched_amfi:
                continue
            k_clean = fund_match_key.get(k) or k
            hit_sibling = None
            for mk in matched_list:
                mk_clean = fund_match_key.get(mk) or mk
                if compact_name(k_clean) == compact_name(mk_clean):
                    hit_sibling = fund_by_norm.get(mk, {}).get("base_name")
                    break
                if _compact_fund_soft(k_clean) == _compact_fund_soft(mk_clean):
                    hit_sibling = fund_by_norm.get(mk, {}).get("base_name")
                    break
                if containment_score(k_clean, mk_clean) >= 95.0 or containment_score(mk_clean, k_clean) >= 95.0:
                    hit_sibling = fund_by_norm.get(mk, {}).get("base_name")
                    break
                if fuzz.token_sort_ratio(k_clean, mk_clean) >= 94.0:
                    la, lb = len(compact_name(k_clean)), len(compact_name(mk_clean))
                    if min(la, lb) / max(la, lb) >= 0.92:
                        hit_sibling = fund_by_norm.get(mk, {}).get("base_name")
                        break
            if hit_sibling:
                mark_matched(k, via="compact", score=95.0, sibling=hit_sibling)

    # Variant tails (Unclaimed / Retail / Cumulative / …) → sibling of matched primary.
    if matched_amfi:
        matched_compact: dict[str, str] = {}
        matched_stripped: dict[str, str] = {}
        matched_fund_soft: dict[str, str] = {}
        for mk in list(matched_amfi):
            fund = fund_by_norm.get(mk) or {}
            name = fund.get("base_name") or mk
            matched_compact[compact_name(name)] = name
            matched_stripped[compact_name(strip_variant_tails(name))] = name
            matched_fund_soft[_compact_fund_soft(name)] = name
            matched_fund_soft[_compact_fund_soft(strip_variant_tails(name))] = name

        for f in amfi_funds:
            k = f["base_name_norm"]
            if not k or k in matched_amfi:
                continue
            name = f.get("base_name") or ""
            if looks_like_closed_end_fmp(name):
                continue  # true distinct product — do not sibling onto an open-end fund

            # Unclaimed / IE-pool style: "... (Invested in <host fund>"
            # Host fund's disclosure (e.g. DSP Overnight → AMFI 146062) covers these rows.
            host = invested_in_fund(name)
            if host:
                host_c = compact_name(host)
                sibling = matched_compact.get(host_c)
                if not sibling:
                    for mk in matched_amfi:
                        mname = (fund_by_norm.get(mk) or {}).get("base_name") or ""
                        if compact_name(mname) == host_c or containment_score(
                            norm_text(host), norm_text(mname)
                        ) >= 95.0:
                            sibling = mname
                            break
                if sibling:
                    mark_matched(k, via="sibling", score=100.0, sibling=sibling)
                    continue

            stripped = strip_variant_tails(name)
            sc = compact_name(stripped)
            sibling = (
                matched_stripped.get(sc)
                or matched_compact.get(sc)
                or matched_fund_soft.get(_compact_fund_soft(stripped))
            )
            if sibling:
                mark_matched(k, via="sibling", score=100.0, sibling=sibling)
                continue
            # Fuzzy sibling against stripped forms of matched primaries
            best_name, best_sc = None, 0.0
            for mk in matched_amfi:
                mname = (fund_by_norm.get(mk) or {}).get("base_name") or ""
                mstrip = strip_variant_tails(mname)
                ts = float(fuzz.token_sort_ratio(stripped, mstrip))
                if ts > best_sc:
                    best_sc = ts
                    best_name = mname
            if best_name and best_sc >= 95.0:
                la, lb = len(compact_name(stripped)), len(compact_name(strip_variant_tails(best_name)))
                if la and lb and min(la, lb) / max(la, lb) >= 0.92:
                    mark_matched(k, via="sibling", score=best_sc, sibling=best_name)

    unmatched_amfi = [f for f in amfi_funds if f["base_name_norm"] not in matched_amfi]
    matched_amfi_list = [
        matched_meta[n]
        for n in sorted(matched_amfi, key=lambda x: (fund_by_norm.get(x) or {}).get("base_name") or x)
        if n in matched_meta
    ]
    return rows, unmatched_amfi, matched_amfi_list


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--period", default="2026-06")
    ap.add_argument("--type", default="monthly", choices=("monthly", "fortnightly"))
    ap.add_argument("--amc", action="append", default=[])
    ap.add_argument(
        "--registry",
        default="registry/amcs.json"
        if Path("registry/amcs.json").exists()
        else "data/sources/amcs.json",
    )
    ap.add_argument("--amfi-funds", default="data/amfi/funds_asof_2026-07-31.json")
    ap.add_argument("--amfi-schemes", default="data/amfi/schemes.json")
    ap.add_argument("--disclosures-root", default="")
    ap.add_argument("--parsed-root", default="")
    ap.add_argument("--cutoff", type=float, default=84.0)
    ap.add_argument(
        "--shortcode-map",
        default="registry/disclosure_shortcode_map.json"
        if Path("registry/disclosure_shortcode_map.json").exists()
        else "data/sources/disclosure_shortcode_map.json",
        help="Durable amc_id+shortcode → AMFI code registry",
    )
    ap.add_argument(
        "--write-shortcode-map",
        action="store_true",
        default=True,
        help="Merge newly matched shortcodes into the durable registry (default: on)",
    )
    ap.add_argument(
        "--no-write-shortcode-map",
        action="store_false",
        dest="write_shortcode_map",
        help="Do not update the durable shortcode registry",
    )
    args = ap.parse_args()
    if not args.disclosures_root:
        args.disclosures_root = f"data/disclosures/{args.type}"
    if not args.parsed_root:
        args.parsed_root = f"data/parsed/{args.type}"

    registry = load_amc_registry(Path(args.registry))
    funds = json.loads(Path(args.amfi_funds).read_text(encoding="utf-8"))
    amfi_by_amc: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for f in funds:
        amfi_by_amc[f["amc_name"]].append(f)

    amc_map = map_amcs_to_amfi(registry, sorted(amfi_by_amc.keys()))
    shortcode_map_path = Path(args.shortcode_map)
    shortcode_map = load_shortcode_map(shortcode_map_path)

    period = args.period
    disc_root = Path(args.disclosures_root) / period
    parsed_root = Path(args.parsed_root) / period
    out_root = parsed_root / "_matching"
    out_root.mkdir(parents=True, exist_ok=True)

    # Only AMCs that have disclosure/parsed folders (or requested)
    amc_ids = args.amc or sorted(
        {p.name for p in disc_root.iterdir() if p.is_dir()}
        | {p.name for p in parsed_root.iterdir() if p.is_dir() and not p.name.startswith("_")}
    )

    tmp = Path(tempfile.mkdtemp(prefix="match-amfi-"))
    summary_rows = []
    shortcode_hits = 0
    name_hits = 0
    try:
        for amc_id in amc_ids:
            meta = amc_map.get(amc_id) or {"amfi_amc_name": None, "score": 0}
            amfi_name = meta.get("amfi_amc_name")
            disclosures = extract_disclosure_schemes(
                amc_id,
                disc_root / amc_id,
                parsed_root / amc_id,
                tmp,
                disclosure_type=args.type,
            )
            amfi_funds = amfi_by_amc.get(amfi_name, []) if amfi_name else []
            # Full catalog always in outputs. Fortnightly disclosures only bind to
            # debt-like names so equity does not false-match — equity stays unmatched
            # until a monthly pack is present (never silently dropped from universe).
            match_candidates = None
            if args.type == "fortnightly" and amfi_funds:
                debt = [f for f in amfi_funds if is_debt_like_amfi_fund(f.get("base_name") or "")]
                match_candidates = debt if debt else amfi_funds

            amc_lookup: dict[str, str] = {}
            for k, v in shortcode_map.items():
                if not k.startswith(f"{amc_id}::"):
                    continue
                token = k.split("::", 1)[1]
                code = str(v["canonical_amfi_code"])
                amc_lookup[token] = code
                amc_lookup[token.casefold()] = code
            matches, unmatched_amfi, matched_amfi_meta = match_within_amc(
                disclosures,
                amfi_funds,
                cutoff=args.cutoff,
                match_candidates=match_candidates,
                shortcode_lookup=amc_lookup,
            )
            matched = [m for m in matches if m.status == "matched"]
            unmatched_disc = [m for m in matches if m.status == "unmatched_disclosure"]
            # Tag helpers only — every fund stays in matched_amfi_meta or unmatched_amfi
            unmatched_variant = [f for f in unmatched_amfi if is_secondary_amfi_fund(f["base_name"])]
            unmatched_core = [f for f in unmatched_amfi if not is_secondary_amfi_fund(f["base_name"])]

            via_by_norm = {
                (mmeta.get("base_name") or ""): mmeta.get("via") or "direct"
                for mmeta in matched_amfi_meta
            }

            def _disc_via(m: MatchRow) -> str:
                return m.disclosure.get("match_via") or via_by_norm.get(
                    (m.amfi or {}).get("base_name") or "", "direct"
                )

            def _persist_key(raw: str | None) -> str | None:
                s = (raw or "").strip()
                if not s:
                    return None
                if " " not in s and is_shortcode(s):
                    return normalize_shortcode(s)
                return s

            for m in matched:
                via = _disc_via(m)
                sc = _persist_key((m.disclosure or {}).get("shortcode"))
                if via == "shortcode":
                    shortcode_hits += 1
                else:
                    name_hits += 1
                if args.write_shortcode_map and sc and m.amfi and m.score >= args.cutoff:
                    key = f"{amc_id}::{sc}"
                    prev = shortcode_map.get(key)
                    if prev and str(prev.get("canonical_amfi_code")) != str(m.amfi["canonical_amfi_code"]):
                        # Do not auto-overwrite a conflicting prior binding
                        if (prev.get("confidence") or "") in {"confirmed", "manual", "seeded"}:
                            continue
                    shortcode_map[key] = {
                        "amc_id": amc_id,
                        "shortcode": sc,
                        "aliases": list((prev or {}).get("aliases") or []),
                        "canonical_amfi_code": str(m.amfi["canonical_amfi_code"]),
                        "amfi_base_name": m.amfi["base_name"],
                        "disclosure_label": m.disclosure.get("label"),
                        "confidence": "seeded",
                        "source": "shortcode" if via == "shortcode" else "name_match",
                        "first_seen_period": (prev or {}).get("first_seen_period") or period,
                        "last_verified_period": period,
                    }

            payload = {
                "period": period,
                "amc_id": amc_id,
                "amc_map": meta,
                "cutoff": args.cutoff,
                "counts": {
                    "disclosure_schemes": len(disclosures),
                    "amfi_funds": len(amfi_funds),
                    "matched_disclosures": len(matched),
                    "matched_amfi": len(matched_amfi_meta),
                    "matched": len(matched_amfi_meta),  # coverage: AMFI rows mapped
                    "unmatched_disclosure": len(unmatched_disc),
                    "unmatched_amfi": len(unmatched_amfi),
                    "unmatched_amfi_primary": len(unmatched_core),
                    "unmatched_amfi_secondary": len(unmatched_variant),
                    "shortcode_bound": sum(
                        1 for mmeta in matched_amfi_meta if mmeta.get("via") == "shortcode"
                    ),
                },
                "matched_disclosures": [
                    {
                        "disclosure_label": m.disclosure["label"],
                        "disclosure_base": m.disclosure["base_name"],
                        "shortcode": _persist_key(m.disclosure.get("shortcode"))
                        or m.disclosure.get("shortcode"),
                        "source": m.disclosure.get("source"),
                        "amfi_base_name": m.amfi["base_name"] if m.amfi else None,
                        "canonical_amfi_code": m.amfi["canonical_amfi_code"] if m.amfi else None,
                        "amfi_codes": m.amfi["amfi_codes"] if m.amfi else None,
                        "score": m.score,
                        "via": _disc_via(m),
                    }
                    for m in sorted(matched, key=lambda x: -x.score)
                ],
                # Back-compat alias used by older coverage rebuilders
                "matched": [
                    {
                        "disclosure_label": m.disclosure["label"],
                        "disclosure_base": m.disclosure["base_name"],
                        "shortcode": _persist_key(m.disclosure.get("shortcode"))
                        or m.disclosure.get("shortcode"),
                        "source": m.disclosure.get("source"),
                        "amfi_base_name": m.amfi["base_name"] if m.amfi else None,
                        "canonical_amfi_code": m.amfi["canonical_amfi_code"] if m.amfi else None,
                        "amfi_codes": m.amfi["amfi_codes"] if m.amfi else None,
                        "score": m.score,
                        "via": _disc_via(m),
                    }
                    for m in sorted(matched, key=lambda x: -x.score)
                ]
                + [
                    {
                        "disclosure_label": None,
                        "disclosure_base": None,
                        "shortcode": None,
                        "source": mmeta.get("via") or "sibling",
                        "amfi_base_name": mmeta.get("base_name"),
                        "canonical_amfi_code": mmeta.get("canonical_amfi_code"),
                        "amfi_codes": None,
                        "score": mmeta.get("score"),
                        "via": mmeta.get("via"),
                        "sibling": mmeta.get("sibling"),
                    }
                    for mmeta in matched_amfi_meta
                    if mmeta.get("via") and mmeta.get("via") not in {"direct", "shortcode"}
                ],
                "matched_amfi": matched_amfi_meta,
                "unmatched_disclosure": [
                    {
                        "disclosure_label": m.disclosure["label"],
                        "disclosure_base": m.disclosure["base_name"],
                        "shortcode": _persist_key(m.disclosure.get("shortcode"))
                        or m.disclosure.get("shortcode"),
                        "source": m.disclosure.get("source"),
                        "best_amfi_guess": m.amfi["base_name"] if m.amfi else None,
                        "best_score": m.score,
                    }
                    for m in sorted(unmatched_disc, key=lambda x: -x.score)
                ],
                "unmatched_amfi": [
                    {
                        "base_name": f["base_name"],
                        "canonical_amfi_code": f["canonical_amfi_code"],
                        "plan_count": f["plan_count"],
                        "variant": is_secondary_amfi_fund(f["base_name"]),
                        "secondary": is_secondary_amfi_fund(f["base_name"]),  # back-compat tag
                    }
                    for f in sorted(
                        unmatched_amfi,
                        key=lambda x: (is_secondary_amfi_fund(x["base_name"]), x["base_name"]),
                    )
                ],
            }
            out_path = out_root / f"{amc_id}.json"
            out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

            # CSV matrix-lite
            import csv

            csv_path = out_root / f"{amc_id}.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(
                    [
                        "status",
                        "disclosure_label",
                        "disclosure_shortcode",
                        "amfi_base_name",
                        "canonical_amfi_code",
                        "score",
                        "source",
                        "via",
                    ]
                )
                for m in matched:
                    w.writerow(
                        [
                            "matched",
                            m.disclosure["label"],
                            _persist_key(m.disclosure.get("shortcode"))
                            or m.disclosure.get("shortcode")
                            or "",
                            m.amfi["base_name"],
                            m.amfi["canonical_amfi_code"],
                            f"{m.score:.1f}",
                            m.disclosure.get("source"),
                            _disc_via(m),
                        ]
                    )
                for m in unmatched_disc:
                    w.writerow(
                        [
                            "unmatched_disclosure",
                            m.disclosure["label"],
                            _persist_key(m.disclosure.get("shortcode"))
                            or m.disclosure.get("shortcode")
                            or "",
                            m.amfi["base_name"] if m.amfi else "",
                            "",
                            f"{m.score:.1f}",
                            m.disclosure.get("source"),
                            "",
                        ]
                    )
                for fnd in unmatched_amfi:
                    w.writerow(
                        [
                            "unmatched_amfi",
                            "",
                            "",
                            fnd["base_name"],
                            fnd["canonical_amfi_code"],
                            "",
                            "",
                            "",
                        ]
                    )

            c = payload["counts"]
            summary_rows.append(
                {
                    "amc_id": amc_id,
                    "amfi_amc_name": amfi_name,
                    "amc_match_score": meta.get("score"),
                    **c,
                }
            )
            print(
                f"{amc_id:40s} disc={c['disclosure_schemes']:4d} amfi={c['amfi_funds']:4d} "
                f"mapped_amfi={c['matched_amfi']:4d} shortcode={c['shortcode_bound']:3d} "
                f"un_disc={c['unmatched_disclosure']:4d} un_amfi={c['unmatched_amfi']:4d}"
            )
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    if args.write_shortcode_map:
        save_shortcode_map(shortcode_map_path, shortcode_map)
        print(
            f"Wrote durable shortcode map: {shortcode_map_path} "
            f"({len(json.loads(shortcode_map_path.read_text()).get('entries') or [])} entries)"
        )

    summary = {
        "period": period,
        "cutoff": args.cutoff,
        "amcs": len(summary_rows),
        "totals": {
            "disclosure_schemes": sum(r["disclosure_schemes"] for r in summary_rows),
            "amfi_funds": sum(r["amfi_funds"] for r in summary_rows),
            "matched": sum(r["matched"] for r in summary_rows),
            "unmatched_disclosure": sum(r["unmatched_disclosure"] for r in summary_rows),
            "unmatched_amfi": sum(r["unmatched_amfi"] for r in summary_rows),
            "unmatched_amfi_primary": sum(r.get("unmatched_amfi_primary", 0) for r in summary_rows),
            "unmatched_amfi_secondary": sum(r.get("unmatched_amfi_secondary", 0) for r in summary_rows),
            "shortcode_bound_amfi": sum(r.get("shortcode_bound", 0) for r in summary_rows),
        },
        "amc_map": amc_map,
        "results": summary_rows,
    }
    (out_root / "_summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print("\nSummary:", json.dumps(summary["totals"], indent=2))
    print(f"Wrote {out_root}/_summary.json and per-AMC json/csv matrices")


if __name__ == "__main__":
    main()
