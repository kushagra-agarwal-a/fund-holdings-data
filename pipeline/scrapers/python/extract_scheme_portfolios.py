#!/usr/bin/env python3
"""
Extract scheme-wise portfolio rows from downloaded AMC monthly Excel files.

Input layout (already created by fetchers):
  amcs/<amc-slug>/<YYYY-MM>/*.xlsx
  amcs/<amc-slug>/<YYYY-MM>/manifest.json

Output layout:
  amcs/<amc-slug>/<YYYY-MM>/<scheme-name>/portfolio.csv
  amcs/<amc-slug>/<YYYY-MM>/<scheme-name>/portfolio.json
  amcs/<amc-slug>/<YYYY-MM>/schemes.json
  amcs/<amc-slug>/schemes_global.json
  amcs/schemes_global.json
"""
from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as e:  # pragma: no cover
    raise SystemExit(
        "Missing dependency openpyxl.\n"
        "Install with: python3 -m pip install openpyxl\n"
        f"Original error: {e}"
    ) from e


HEADER_HINTS = (
    "scheme",
    "security",
    "instrument",
    "isin",
    "issuer",
    "quantity",
    "market value",
    "face value",
    "nav",
)
SCHEME_COL_HINTS = ("scheme", "scheme name", "name of scheme", "scheme_name")


@dataclass
class ExtractSummary:
    amc: str
    month: str
    files_seen: int = 0
    files_parsed: int = 0
    rows_extracted: int = 0
    schemes: int = 0


def safe_name(s: str) -> str:
    out = re.sub(r"[^\w.\-() ]+", "_", (s or "").strip())
    out = re.sub(r"\s+", " ", out).strip(" ._")
    return out[:180] or "unknown"


def norm_cell(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def looks_like_header(values: list[str]) -> bool:
    text = " | ".join(v.lower() for v in values if v)
    if not text:
        return False
    score = sum(1 for h in HEADER_HINTS if h in text)
    return score >= 2


def choose_header_row(rows: list[list[str]]) -> tuple[int, list[str]] | None:
    for idx, row in enumerate(rows[:40]):
        if looks_like_header(row):
            return idx, row
    return None


def scheme_col_index(headers: list[str]) -> int | None:
    lowered = [h.lower().strip() for h in headers]
    for i, h in enumerate(lowered):
        if any(k in h for k in SCHEME_COL_HINTS):
            return i
    return None


def non_empty_row(row: list[str]) -> bool:
    return any(c.strip() for c in row)


def sheet_rows(sheet) -> list[list[str]]:
    out: list[list[str]] = []
    max_col = min(max(sheet.max_column, 1), 120)
    for r in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=max_col, values_only=True):
        out.append([norm_cell(v) for v in r])
    return out


def parse_workbook(path: Path) -> dict[str, list[dict[str, str]]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    scheme_map: dict[str, list[dict[str, str]]] = {}
    for ws in wb.worksheets:
        rows = sheet_rows(ws)
        picked = choose_header_row(rows)
        if not picked:
            continue
        hidx, headers = picked
        headers = [h if h else f"col_{i+1}" for i, h in enumerate(headers)]
        sc_idx = scheme_col_index(headers)

        default_scheme = safe_name(path.stem)
        for raw in rows[hidx + 1 :]:
            if not non_empty_row(raw):
                continue
            if all((c or "").strip("-_ ").lower() in {"", "total", "totals"} for c in raw):
                continue
            row_obj: dict[str, str] = {}
            for i, head in enumerate(headers):
                row_obj[head] = raw[i] if i < len(raw) else ""

            if sc_idx is not None and sc_idx < len(raw):
                scheme_name = safe_name(raw[sc_idx] or default_scheme)
            else:
                scheme_name = default_scheme

            # Skip trivial lines that are effectively section labels only.
            filled = [v for v in row_obj.values() if (v or "").strip()]
            if len(filled) <= 1:
                continue

            scheme_map.setdefault(scheme_name, []).append(row_obj)
    return scheme_map


def write_scheme_outputs(month_dir: Path, scheme: str, rows: list[dict[str, str]]) -> None:
    scheme_dir = month_dir / safe_name(scheme)
    scheme_dir.mkdir(parents=True, exist_ok=True)

    json_path = scheme_dir / "portfolio.json"
    json_path.write_text(json.dumps(rows, indent=2, ensure_ascii=True), encoding="utf-8")

    # CSV uses union of row keys preserving first-seen order.
    fields: list[str] = []
    seen: set[str] = set()
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                fields.append(k)
    csv_path = scheme_dir / "portfolio.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def load_manifest_files(month_dir: Path) -> list[Path]:
    man = month_dir / "manifest.json"
    if not man.is_file():
        return []
    try:
        data = json.loads(man.read_text(encoding="utf-8"))
    except Exception:
        return []
    out: list[Path] = []
    if not isinstance(data, list):
        return out
    for rec in data:
        if not isinstance(rec, dict):
            continue
        saved_as = str(rec.get("saved_as") or "").strip()
        if not saved_as:
            continue
        p = month_dir / saved_as
        if p.is_file():
            out.append(p)
    return out


def update_json_list(path: Path, values: set[str]) -> None:
    path.write_text(json.dumps(sorted(values), indent=2, ensure_ascii=True) + "\n", encoding="utf-8")


def process_amc_month(repo_root: Path, amc: str, month: str) -> ExtractSummary:
    month_dir = repo_root / "amcs" / amc / month
    summary = ExtractSummary(amc=amc, month=month)
    if not month_dir.is_dir():
        return summary

    files = load_manifest_files(month_dir)
    if not files:
        files = [p for p in month_dir.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx"]
    summary.files_seen = len(files)

    merged: dict[str, list[dict[str, str]]] = {}
    for f in files:
        if f.suffix.lower() != ".xlsx":
            continue
        try:
            per_scheme = parse_workbook(f)
        except Exception:
            continue
        summary.files_parsed += 1
        for scheme, rows in per_scheme.items():
            for r in rows:
                r.setdefault("__source_file", f.name)
            merged.setdefault(scheme, []).extend(rows)

    for scheme, rows in merged.items():
        write_scheme_outputs(month_dir, scheme, rows)
    schemes = set(merged.keys())
    summary.schemes = len(schemes)
    summary.rows_extracted = sum(len(v) for v in merged.values())

    update_json_list(month_dir / "schemes.json", schemes)
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract scheme-wise portfolio data from monthly Excel files")
    parser.add_argument("--months", nargs="+", required=True, help="One or more YYYY-MM")
    parser.add_argument("--amcs", nargs="*", help="Optional AMC slugs subset")
    parser.add_argument(
        "--root",
        type=Path,
        default=Path(__file__).resolve().parent.parent,
        help="mf-monthly-holdings root",
    )
    args = parser.parse_args()

    amcs_root = args.root / "amcs"
    all_amcs = sorted(
        p.name for p in amcs_root.iterdir() if p.is_dir() and (p / "README.md").is_file()
    )
    selected = sorted(set(args.amcs)) if args.amcs else all_amcs

    global_all_schemes: set[str] = set()
    amc_global_schemes: dict[str, set[str]] = {a: set() for a in selected}
    summaries: list[ExtractSummary] = []

    for amc in selected:
        for month in args.months:
            s = process_amc_month(args.root, amc, month)
            summaries.append(s)
            month_scheme_file = args.root / "amcs" / amc / month / "schemes.json"
            if month_scheme_file.is_file():
                try:
                    month_schemes = set(json.loads(month_scheme_file.read_text(encoding="utf-8")))
                except Exception:
                    month_schemes = set()
                amc_global_schemes[amc].update(month_schemes)
                global_all_schemes.update({f"{amc}::{x}" for x in month_schemes})

    for amc, vals in amc_global_schemes.items():
        update_json_list(args.root / "amcs" / amc / "schemes_global.json", vals)
    update_json_list(args.root / "amcs" / "schemes_global.json", global_all_schemes)

    print("=== Scheme extraction summary ===")
    total_rows = 0
    total_files = 0
    for s in summaries:
        total_rows += s.rows_extracted
        total_files += s.files_parsed
        print(
            f"{s.amc} {s.month}: files_seen={s.files_seen}, files_parsed={s.files_parsed}, "
            f"schemes={s.schemes}, rows={s.rows_extracted}"
        )
    print(f"TOTAL: files_parsed={total_files}, rows={total_rows}")


if __name__ == "__main__":
    main()
