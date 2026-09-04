"""Shared helpers for AMC-wise portfolio disclosure parsers."""
from __future__ import annotations

import calendar
import csv
import json
import re
import shutil
import tempfile
from dataclasses import asdict, dataclass, field, replace
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except Exception as e:  # pragma: no cover
    raise SystemExit("Missing openpyxl. Install into .venv") from e

ISIN_RE = re.compile(r"\b([A-Z]{2}[A-Z0-9]{9}\d)\b")
NIL_RE = re.compile(r"(?i)^\s*(nil|n\.?a\.?|na|-|—|–|\|)\s*$")
SECTION_RE = re.compile(
    r"(?i)^\s*(equity|debt|money\s*market|treasury|government|reverse\s*repo|"
    r"treps|tri[- ]?party|cblo|commercial\s*paper|certificate\s*of\s*deposit|"
    r"fixed\s*deposit|cash|others?|derivatives?|futures?|options?|"
    r"arbitrage|commodity|"
    r"preference|mutual\s*fund|investment\s+in\s+mutual\s*fund|etf|gold|silver|units?\s+of|"
    r"a\)|b\)|c\)|listed|unlisted|privately|securiti[sz]ed|"
    r"foreign\s+securit|overseas|sub\s*total|total|grand[\s_]*total|"
    r"net\s+receivables|net\s+current|portfolio\s+turnover|"
    r"notes?\s*&|^\s*\(?[a-z]\)\s)",
)
SKIP_HOLDING_NAME_RE = re.compile(
    # NJ prints GRAND_TOTAL (underscore); others use "GRAND TOTAL".
    r"(?i)^(sub\s*total|grand[\s_]*total|total)\b|"
    r".*\b(?:sub\s*)?totals?\s*$|"  # Tata etc. "EQUITY & EQUITY RELATED TOTAL"
    r"^(portfolio\s+total|net\s+assets)\b|"
    r"^(name\s+of\s+(the\s+)?(instrument|security))\b|"  # stray re-bound header rows
    r"^(net\s+receivables|nil|others?(\s*\(.*\))?)$"
)
STOP_TABLE_RE = re.compile(
    r"(?i)^(?:"
    r"grand[\s_]*total\b|total\s+number\s+of\s+contracts|"
    r"underlying\s+current\s+(?:price|option)|"
    r"portfolio\s+turnover|^\s*~\s*ytm|"
    r"notes?\s*&?\s*symbols|"
    # Annex / side tables after the main SEBI portfolio (UTI etc.)
    r".*default\s+beyond\s+maturity|"
    r".*non[\s\-]?traded\s+securit|"
    r".*exposure\s+to\s+credit\s+default|"
    r".*total\s+amt\.?\s*due|"
    r".*value\s+as\s+per\s+nca|"
    r"name\s+of\s+the\s+security\b|"
    r"details\s+of\s+default|"
    r"a1\)\s*exposure|"
    # End after cash/NCA. Do not stop on PORTFOLIO TOTAL — Tata prints cash after it.
    r"net\s+assets\b"
    r")"
)

# Word-ish patterns — avoid substring traps like "instrument" inside "instruments"
# (UTI section banner "MONEY MARKET INSTRUMENTS" used to trip the mid-table header stop).
HEADER_HINT_RES = (
    re.compile(r"(?i)\bisin\b"),
    re.compile(r"(?i)\binstrument\b"),
    re.compile(r"(?i)\bsecurity\b"),
    re.compile(r"(?i)\bquantity\b"),
    re.compile(r"(?i)\bmarket\b"),
    re.compile(r"(?i)%\s*to"),
    re.compile(r"(?i)\bnav\b"),
    re.compile(r"(?i)\baum\b"),
    re.compile(r"(?i)\bindustry\b"),
    re.compile(r"(?i)\brating\b"),
    re.compile(r"(?i)\bytm\b"),
)

COLUMN_ALIASES: dict[str, re.Pattern[str]] = {
    "security_code": re.compile(
        r"(?i)^\s*(security\s*)?(code|no\.?)\s*$|^\s*sr\.?\s*no|^\s*scrip\s*code|^\s*sl\s*no"
    ),
    "instrument": re.compile(
        r"(?i)name\s+of\s+(the\s+)?(instrument|security|issuer)|"
        r"company\s*/?\s*issuer\s*/?\s*instrument|"
        r"instrument\s*/?\s*issuer|security\s+name|^(security|instrument|scrip)\b"
    ),
    "isin": re.compile(r"(?i)^\s*isin\b|isin\s*code"),
    "coupon": re.compile(r"(?i)\bcoupon\b"),
    "ytm": re.compile(r"(?i)\bytm\b|yield\s*to\s*maturity"),
    "ytc": re.compile(r"(?i)\bytc\b|yield\s*to\s*call|at1|tier\s*2"),
    "instrument_yield": re.compile(r"(?i)\byield\b"),
    "residual_maturity": re.compile(r"(?i)residual|remaining\s*maturity|\btenor\b"),
    "maturity_date": re.compile(
        r"(?i)maturity\s*date|date\s*of\s*maturity|redemption\s*date|(?<!to )\bmaturity\b"
    ),
    "face_value": re.compile(r"(?i)face\s*value|par\s*value"),
    "put_call_date": re.compile(r"(?i)put\s*/?\s*call|call\s*date|put\s*date"),
    "industry_rating": re.compile(
        r"(?i)industry\s*[\^+\*]?\s*/\s*rating|rating\s*/\s*industry|"
        r"industry\s+classification\s*/\s*rating"
    ),
    "rating_agency": re.compile(r"(?i)rating\s*agency"),
    "rating": re.compile(r"(?i)credit\s*rating|^\s*rating\b|conservative\s+rating"),
    "industry": re.compile(r"(?i)\bindustry\b|\bsector\b"),
    "quantity": re.compile(
        r"(?i)^\s*quantity\b|\bqty\b|no\.?\s*of\s*(shares|units)|hedged\s+quantity"
    ),
    "futures_price": re.compile(
        r"(?i)(futures?|option|contract)\s+price|current\s+price\s+of\s+the\s+contract"
    ),
    "market_value": re.compile(
        r"(?i)market\s*/?\s*fair\s*value|market[\s\-]*value|mkt\.?\s*[\-]?\s*val(?:ue)?|"
        r"exposure\s*/\s*market|value\s*\(.*?(rs|inr|lakh)|"
        r"value\s+recognised\s+in\s+nav"
    ),
    "pct_nav": re.compile(
        r"(?i)%\s*(to|of)?\s*(n\.?a\.?v|aum|net\s*assets?)|"
        r"percent(age)?\s*(to|of)?\s*(n\.?a\.?v|aum|net\s*assets?)"
    ),
    "macaulay_duration": re.compile(r"(?i)macaulay"),
    "modified_duration": re.compile(r"(?i)modified\s*duration"),
    "duration": re.compile(r"(?i)\bduration\b"),
    "listed_status": re.compile(r"(?i)listed\s*status|\blisted\b|\bunlisted\b"),
    "accrued_interest": re.compile(r"(?i)accrued\s*interest|interest\s*accrued"),
    "position_side": re.compile(r"(?i)long\s*/\s*\(?\s*short"),
    "margin": re.compile(r"(?i)^\s*margin\b|margin\s+maintained"),
    "market_cap": re.compile(r"(?i)market\s*capital"),
    "underlying": re.compile(r"(?i)^\s*underlying\b"),
    "asset_class": re.compile(r"(?i)asset\s*(class|type)|type of security"),
}

EXTRA_HOLDING_FIELDS = (
    "coupon",
    "maturity_date",
    "residual_maturity",
    "put_call_date",
    "instrument_yield",
    "industry_rating",
    "rating",
    "rating_agency",
    "face_value",
    "listed_status",
    "macaulay_duration",
    "modified_duration",
    "duration",
    "accrued_interest",
    "futures_price",
    "position_side",
    "margin",
    "market_cap",
    "underlying",
    "asset_class",
)


@dataclass
class Holding:
    instrument: str
    isin: str = ""
    industry: str = ""
    quantity: str = ""
    market_value: str = ""
    pct_nav: str = ""
    ytm: str = ""
    ytc: str = ""
    security_code: str = ""
    coupon: str = ""
    maturity_date: str = ""
    residual_maturity: str = ""
    put_call_date: str = ""
    instrument_yield: str = ""
    industry_rating: str = ""
    rating: str = ""
    rating_agency: str = ""
    face_value: str = ""
    listed_status: str = ""
    macaulay_duration: str = ""
    modified_duration: str = ""
    duration: str = ""
    accrued_interest: str = ""
    futures_price: str = ""
    position_side: str = ""
    margin: str = ""
    market_cap: str = ""
    underlying: str = ""
    asset_class: str = ""
    section: str = ""
    raw: dict[str, str] = field(default_factory=dict)


def holding_from_dict(h: dict[str, Any] | Holding) -> Holding:
    """Rebuild a Holding from portfolio.json row (keep industry/qty/MV for futures QC)."""
    if isinstance(h, Holding):
        return h
    raw_in = h.get("raw") if isinstance(h.get("raw"), dict) else None
    raw = {
        str(k): str(v)
        for k, v in (raw_in or h).items()
        if v is not None and k != "raw"
    }
    extras = {k: str(h.get(k) or "") for k in EXTRA_HOLDING_FIELDS}
    return Holding(
        instrument=str(h.get("instrument") or h.get("name") or ""),
        isin=str(h.get("isin") or ""),
        industry=str(h.get("industry") or extras.get("industry_rating") or ""),
        quantity=str(h.get("quantity") or ""),
        market_value=str(h.get("market_value") or ""),
        pct_nav=str(
            h.get("pct_nav") if h.get("pct_nav") is not None else h.get("weight") or ""
        ),
        ytm=str(h.get("ytm") or ""),
        ytc=str(h.get("ytc") or ""),
        security_code=str(h.get("security_code") or ""),
        section=str(h.get("section") or ""),
        raw=raw,
        **extras,
    )


@dataclass
class SchemePortfolio:
    amc_id: str
    disclosure_type: str  # monthly | fortnightly
    period: str
    scheme_name: str
    shortcode: str | None
    as_of: str | None
    source_file: str
    sheet_name: str
    holdings: list[Holding] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    def to_meta(self) -> dict[str, Any]:
        return {
            "amc_id": self.amc_id,
            "disclosure_type": self.disclosure_type,
            "period": self.period,
            "scheme_name": self.scheme_name,
            "shortcode": self.shortcode,
            "as_of": self.as_of,
            "source_file": self.source_file,
            "sheet_name": self.sheet_name,
            "holding_count": len(self.holdings),
            "notes": self.notes,
        }


def safe_name(s: str) -> str:
    out = re.sub(r"[^\w.\-() ]+", "_", (s or "").strip())
    out = re.sub(r"\s+", " ", out).strip(" ._")
    return out[:180] or "unknown"


def norm_cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def file_kind(path: Path) -> str:
    try:
        head = path.read_bytes()[:8]
    except OSError:
        return "other"
    if head.startswith(b"PK"):
        return "zip" if path.suffix.lower() == ".zip" else "ooxml"
    if head.startswith(b"\xd0\xcf\x11\xe0"):
        return "ole"
    return "other"


def prepare_workbook_path(path: Path, tmp_dir: Path) -> Path:
    kind = file_kind(path)
    suffix = path.suffix.lower()
    if kind == "ooxml" and suffix in {".xls", ".xlsm", ""}:
        dest = tmp_dir / (safe_name(path.stem) + ".xlsx")
        shutil.copyfile(path, dest)
        return dest
    if kind == "ole" and suffix in {".xlsx", ".xlsm"}:
        dest = tmp_dir / (safe_name(path.stem) + ".xls")
        shutil.copyfile(path, dest)
        return dest
    return path


def sheet_rows_openpyxl(path: Path) -> list[tuple[str, list[list[str]]]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets: list[tuple[str, list[list[str]]]] = []
    for ws in wb.worksheets:
        rows: list[list[str]] = []
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i >= 20000:
                break
            cells = list(r[:80]) if r is not None else []
            rows.append([norm_cell(v) for v in cells])
        sheets.append((ws.title or "", rows))
    try:
        wb.close()
    except Exception:
        pass
    return sheets


def sheet_rows_xlrd(path: Path) -> list[tuple[str, list[list[str]]]]:
    import xlrd

    book = xlrd.open_workbook(path)
    sheets: list[tuple[str, list[list[str]]]] = []
    for ws in book.sheets():
        rows: list[list[str]] = []
        cols = min(ws.ncols, 80)
        for r in range(min(ws.nrows, 20000)):
            rows.append([norm_cell(ws.cell_value(r, c)) for c in range(cols)])
        sheets.append((ws.name or "", rows))
    return sheets


def load_sheets(path: Path) -> list[tuple[str, list[list[str]]]]:
    with tempfile.TemporaryDirectory() as td:
        prep = prepare_workbook_path(path, Path(td))
        kind = file_kind(prep)
        errors: list[Exception] = []
        order = ["ole", "ooxml"] if kind == "ole" else ["ooxml", "ole"]
        for mode in order:
            try:
                if mode == "ooxml":
                    return sheet_rows_openpyxl(prep)
                return sheet_rows_xlrd(prep)
            except Exception as e:
                errors.append(e)
        raise RuntimeError(
            f"could not read {path.name}: " + " | ".join(str(e) for e in errors)
        )


def _parse_number(val: str) -> float | None:
    try:
        return float(str(val).replace(",", "").replace("%", "").strip())
    except Exception:
        return None


FUTURES_NAME_RE = re.compile(
    r"(?i)"
    # UTI / index futures & options: "NAME-28-Jul-2026", "BANKNIFTY 28-Jul-2026"
    r"((?:-|\s)\d{2}-[A-Za-z]{3}-\d{4})"
    # Edelweiss stock future: date glued after letter/dot — "HDFC Bank Ltd.28/07/2026"
    # (not GSEC/CD maturities like "MAT - 24/07/2037" or "CD - BANK - 28/01/2027")
    r"|((?<=[A-Za-z.])\d{2}/\d{2}/\d{4})\s*$"
    # Sundaram stock future: "HDFC Bank Ltd JUL-2026"
    r"|(\b(?:JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)-20\d{2}\b)"
    # DSP stock/index futures: "HDFC Bank Limited Jul26", "NIFTY Jul26"
    r"|(\b(?:JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\d{2}\s*$)"
    # Kotak: "ITC Ltd.-OCT2026", "Britannia Industries Ltd.-SEP2026", "…-AUG2026"
    r"|(-(?:JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)20\d{2}\b)"
    # Kotak: "BLUE STAR LTD. - August 2026 - Future" / "…-AUG2026"
    r"|(\b(?:January|February|March|April|May|June|July|August|September|October|November|December)"
    r"\s+20\d{2}\s*-\s*Futures?\b)"
    r"|(\bFutures?\s*$)"
    # Commodity futures — not T-Bills "02OCT2026" / GOI "12JUN2031"
    r"|((?:GOLD|SILVER|CRUDE|COPPER|NICKEL|ZINC|ALUMINIUM|ALUMINUM|NATURAL\s*GAS)"
    r"[A-Z0-9]*-?\d{2}[A-Za-z]{3}\d{4}"
    r"|\d{2}[A-Za-z]{3}\d{4}-(?:MCX|NSE|BSE))"
)


def is_futures_holding(
    h: Holding,
    *,
    irs_as_futures: bool = True,
    options_as_futures: bool = True,
) -> bool:
    """True for futures/options/swap hedge rows (incl. UTI dated futures, ICICI $$ legs).

    Covered calls written on equity (often suffixed $$ in ICICI sheets) stay in the
    equity/cash book — they are not treated as futures for exclude_futures totals.

    Interest-rate swaps: excluded from cash book by default (ICICI). Pass
    ``irs_as_futures=False`` for AMCs that keep IRS in the cash book (Nippon).

    Index/stock options: excluded by default. Pass ``options_as_futures=False`` for AMCs
    where option premiums sit inside GRAND TOTAL (Edelweiss).
    """
    name = h.instrument or ""
    industry = (h.industry or "").strip()
    if re.search(r"(?i)covered\s+call", name):
        return False
    # IRS checked before section banners — Nippon lists them under "Derivatives".
    if re.search(r"(?i)interest\s+rate\s+swaps?|\birs\b", name):
        return bool(irs_as_futures)
    sec = (h.section or "").strip()
    sec_u = sec.upper()
    is_option_sec = bool(
        re.search(r"(?i)\boptions?\b", sec) or re.search(r"(?i)\boptions?\b", industry)
    )
    # Parent "Derivatives" / "f) Derivative" (Sundaram) and Index/Stock Future count.
    # DSP also labels legs in the industry/rating column: "Stock Futures".
    is_future_sec = bool(
        (sec_u in {"FUTURES"} or sec_u.startswith("FUTURE"))
        or re.search(r"(?i)\bfutures?\b", sec)
        or re.search(r"(?i)\bderivatives?\b", sec)
        or re.search(r"(?i)commodity\s+derivative|exchange\s+traded\s+commodity", sec)
        or re.search(r"(?i)\b(?:stock|index)\s+futures?\b", industry)
    )
    if is_option_sec or re.search(r"(?i)^\s*(put|call)\b", name) or re.search(
        r"(?i)\b(?:put|call)\b", name
    ):
        return bool(options_as_futures)
    if is_future_sec:
        # Sticky futures section can leak onto Mutual Fund unit rows (ISIN INF…).
        # Kotak still prints the *underlying* equity ISIN on futures legs — keep
        # those when the name looks like a future or the leg is short (neg qty/MV).
        qty_n = _parse_number(h.quantity)
        mv_n = _parse_number(h.market_value)
        looks_future_name = bool(FUTURES_NAME_RE.search(name))
        looks_short_leg = (qty_n is not None and qty_n < 0) or (mv_n is not None and mv_n < 0)
        if h.isin and not looks_future_name and not looks_short_leg:
            pass
        else:
            return True
    if FUTURES_NAME_RE.search(name):
        # Dated cash bonds / TREPS (Quant/UTI) look like "GOI 18-Aug-2055" / "TREPS 03-Aug-2026".
        if re.search(
            r"(?i)\b(?:treps?|tri[\s\-]?party|reverse\s+repo|cblo|goi|gsec|sdl|"
            r"government|t[\s\-]?bill|treasury|cd\b|cp\b|certificate\s+of\s+deposit|"
            r"commercial\s+paper)\b",
            name,
        ):
            pass
        elif re.search(r"(?i)^\s*\d+(\.\d+)?\s*%", name):  # coupon bonds
            pass
        else:
            return True
    # Quant stock futures: "Infosys Limited 25/08/2026". Bonds like
    # "REC LTD - 7.58% - 31/05/2029" have ISINs / coupons — leave those alone.
    if (
        re.search(r"(?<![A-Za-z0-9])\d{2}/\d{2}/\d{4}\s*$", name)
        and not (h.isin or "").strip()
        and "%" not in name
        and not re.search(
            r"(?i)\b(?:goi|gsec|sdl|government|t[\s\-]?bill|maturity|mat\b|"
            r"red(?:emption)?|cd\b|cp\b|ncd|debenture)\b",
            name,
        )
    ):
        return True
    # ICICI equity-arbitrage / hedge legs often suffix "$$".
    # Tata debt footnotes use "$$" / "# $$" on bonds/T-bills marked for IRS —
    # usually prefixed with "**" or ending in "# $$" (not bare equity names).
    if re.search(r"\$\$\s*$", name):
        if re.search(
            r"(?i)^\s*\*+\s*|#\s*\$\$\s*$|^\s*\d+(\.\d+)?\s*%|"
            r"treasury\s+bill|t[\s\-]?bill|\bgoi\b\s*-|\birs\b",
            name,
        ):
            return False
        return True
    # Tata equity / commodity futures hedges: "ICICI BANK LTD^", "SILVER … COMMODITY^"
    # Do not treat "^^" structured-obligation footnotes as futures.
    # Sundaram sometimes footnotes cash as "Cash and Other Net Current Assets^".
    if re.search(r"(?<!\^)\^\s*$", name):
        if re.search(
            r"(?i)cash|net\s+current|net\s+receivables|nca\b|treps?|tri[\s\-]?party|"
            r"reverse\s+repo|\brepo\b|cblo",
            name,
        ):
            return False
        return True
    return False


def normalize_fractional_pct_nav(holdings: list[Holding]) -> tuple[list[Holding], dict[str, Any]]:
    """If %NAV looks like fraction-of-1.0 (ICICI etc.), store as percent (×100).

    Hedged books (DSP Dynamic AAF / Arbitrage) can sum to ≪1.0 on the full book while
    the positive sleeve (or cash book after dropping futures) still sits near 1.0.
    """
    meta: dict[str, Any] = {"fractional_pct_scaled": False}
    weights = [_parse_number(h.pct_nav) for h in holdings]
    present = [w for w in weights if w is not None]
    if not present:
        return holdings, meta
    total = sum(present)
    pos_total = sum(w for w in present if w >= 0)
    compact = max(abs(x) for x in present) <= 1.5
    near_one = 0.90 <= abs(total) <= 1.10 or 0.90 <= pos_total <= 1.10
    if not (compact and near_one):
        return holdings, meta
    out: list[Holding] = []
    for h in holdings:
        w = _parse_number(h.pct_nav)
        if w is None:
            out.append(h)
            continue
        pct = f"{w * 100.0:.6f}".rstrip("0").rstrip(".")
        raw = dict(h.raw or {})
        raw["pct_was_fraction"] = "1"
        out.append(replace(h, pct_nav=pct, raw=raw))
    meta["fractional_pct_scaled"] = True
    return out, meta


def allocation_totals(
    holdings: Iterable[Holding],
    *,
    irs_as_futures: bool = True,
    options_as_futures: bool = True,
) -> dict[str, float | None]:
    """Three portfolio %NAV sums used for QC / comparison.

    - ``all``: every numeric weight (cash + futures/options, positives + negatives)
    - ``exclude_negatives``: drop any holding with weight < 0
    - ``exclude_futures``: cash-market book only (drop futures/options; keep cash negatives e.g. NCA)
    - ``exclude_short_futures``: drop only *short* futures/options (keep long futures). Matches
      GRAND TOTAL when sleeve Sub Totals count cash-long arb legs only (DSP Value / Multi Asset).

    Some AMCs (e.g. ICICI) print `% to NAV` as a fraction of 1.0; those are scaled to percent.
    Set ``irs_as_futures=False`` so interest-rate swaps stay in ``exclude_futures`` (Nippon).
    Set ``options_as_futures=False`` so index/stock options stay in cash (Edelweiss).
    """
    all_w: list[float] = []
    pos_w: list[float] = []
    ex_fut_w: list[float] = []
    ex_short_fut_w: list[float] = []
    for h in holdings:
        w = _parse_number(h.pct_nav)
        if w is None:
            continue
        all_w.append(w)
        if w >= 0:
            pos_w.append(w)
        is_fut = is_futures_holding(
            h, irs_as_futures=irs_as_futures, options_as_futures=options_as_futures
        )
        if not is_fut:
            ex_fut_w.append(w)
        if not (is_fut and w < 0):
            ex_short_fut_w.append(w)

    # One scale decision for the whole book (avoid exclude_negatives×100 while all stays
    # fractional — common for DSP hedged portfolios where sum(all) ≪ 1).
    scale = False
    if all_w and max(abs(x) for x in all_w) <= 1.5:
        all_sum = sum(all_w)
        pos_sum = sum(x for x in all_w if x >= 0)
        ex_sum = sum(ex_fut_w) if ex_fut_w else 0.0
        ex_short_sum = sum(ex_short_fut_w) if ex_short_fut_w else 0.0
        if (
            0.90 <= abs(all_sum) <= 1.10
            or 0.90 <= pos_sum <= 1.10
            or 0.90 <= abs(ex_sum) <= 1.10
            or 0.90 <= abs(ex_short_sum) <= 1.10
        ):
            scale = True

    def _as_percent(ws: list[float]) -> list[float]:
        if not ws or not scale:
            return ws
        return [x * 100.0 for x in ws]

    all_w = _as_percent(all_w)
    pos_w = _as_percent(pos_w)
    ex_fut_w = _as_percent(ex_fut_w)
    ex_short_fut_w = _as_percent(ex_short_fut_w)
    return {
        "all": sum(all_w) if all_w else None,
        "exclude_negatives": sum(pos_w) if pos_w else None,
        "exclude_futures": sum(ex_fut_w) if ex_fut_w else None,
        "exclude_short_futures": sum(ex_short_fut_w) if ex_short_fut_w else None,
    }


def allocation_policy_for_amc(amc_id: str) -> dict[str, bool]:
    """Per-AMC flags for allocation_totals (from amc_parser_families.json)."""
    try:
        _root = Path(__file__).resolve().parents[2]
        _reg = _root / "registry" / "amc_parser_families.json"
        _old = _root / "data" / "sources" / "amc_parser_families.json"
        cfg_path = _reg if _reg.exists() else _old
        cfg = json.loads(cfg_path.read_text(encoding="utf-8")).get(amc_id) or {}
    except Exception:
        cfg = {}
    return {
        # Default: IRS treated as futures (excluded from cash book). Nippon keeps IRS in cash.
        "irs_as_futures": not bool(cfg.get("include_irs_in_cash_book")),
        # Default: options excluded. Edelweiss keeps index/stock option premiums in cash.
        "options_as_futures": not bool(cfg.get("include_options_in_cash_book")),
    }


def allocation_totals_for_amc(amc_id: str, holdings: Iterable[Holding]) -> dict[str, float | None]:
    policy = allocation_policy_for_amc(amc_id)
    return allocation_totals(
        holdings,
        irs_as_futures=policy["irs_as_futures"],
        options_as_futures=policy["options_as_futures"],
    )


def meets_allocation_100(total: float | None, *, tol: float = 0.10) -> bool:
    """Default band 99.90–100.10 (tol=0.10)."""
    return total is not None and (100.0 - tol) <= total <= (100.0 + tol)


def meets_allocation_qc(
    totals: dict[str, float | None],
    *,
    tol: float = 0.10,
) -> bool:
    """True when the portfolio cash/NAV book looks complete.

    Prefer ``exclude_futures`` (arb / equity-savings: longs − shorts ≈ 0 on ``all``).
    Also accept ``all`` when stock/index futures sit *inside* GRAND TOTAL (e.g. Sundaram
    multi-cap long hedge) so dropping them undershoots 100%.
    Also accept ``exclude_short_futures`` when GRAND TOTAL / sleeve totals count arb
    cash-longs only while still listing short futures rows (DSP Value / Multi Asset).
    """
    if meets_allocation_100(totals.get("exclude_futures"), tol=tol):
        return True
    if meets_allocation_100(totals.get("all"), tol=tol):
        return True
    return meets_allocation_100(totals.get("exclude_short_futures"), tol=tol)


# Segregated / fully-written-down junk portfolios: AUM and MV are reported as 0.
# Use a nominal 0.01 (same units as the disclosure MV column) per unit of quantity
# so relative weights can still be computed until recovery marks real values.
ZERO_AUM_JUNK_MV_PER_QTY = 0.01


def apply_zero_aum_junk_weights(
    holdings: list[Holding],
    portfolio_total: float | None,
) -> tuple[list[Holding], dict[str, Any]]:
    """If the whole book is zero-valued junk, set MV = 0.01 * qty and recompute %NAV."""
    meta: dict[str, Any] = {"zero_aum_junk_reweight": False}
    if not holdings:
        return holdings, meta
    if portfolio_total is not None and portfolio_total > 0:
        return holdings, meta
    if any((_parse_number(h.market_value) or 0) > 0 for h in holdings):
        return holdings, meta

    effective: list[tuple[Holding, float, bool]] = []
    for h in holdings:
        mv_n = _parse_number(h.market_value)
        qty_n = _parse_number(h.quantity)
        if (mv_n is None or mv_n == 0) and qty_n is not None and qty_n != 0:
            # Nominal valuation for unrecovered junk; use abs(qty) so shorts stay weighted.
            effective.append((h, abs(qty_n) * ZERO_AUM_JUNK_MV_PER_QTY, True))
        elif mv_n is not None and mv_n != 0:
            effective.append((h, mv_n, False))
        else:
            effective.append((h, 0.0, False))

    total = sum(v for _, v, _ in effective)
    if total <= 0:
        return holdings, meta

    out: list[Holding] = []
    for h, eff_mv, synthetic in effective:
        pct = f"{(eff_mv / total) * 100.0:.6f}".rstrip("0").rstrip(".")
        raw = dict(h.raw or {})
        if synthetic:
            raw["junk_synthetic_mv"] = "1"
            raw["junk_mv_per_qty"] = str(ZERO_AUM_JUNK_MV_PER_QTY)
            mv_out = f"{eff_mv:.6f}".rstrip("0").rstrip(".")
        else:
            mv_out = h.market_value
        out.append(replace(h, market_value=mv_out, pct_nav=pct, raw=raw))
    meta["zero_aum_junk_reweight"] = True
    meta["zero_aum_junk_synthetic_total_mv"] = total
    return out, meta


def extract_portfolio_total_mv(rows: list[list[str]]) -> float | None:
    """Best-effort scheme AUM from GRAND TOTAL / Net Assets / CAMS ``TOTAL :`` rows."""
    best: tuple[int, float] | None = None
    for row in rows:
        cells = [(c or "").strip() for c in row]
        if not any(cells):
            continue
        joined = " | ".join(c for c in cells if c)
        first = next((c for c in cells if c), "")
        score = 0
        label = ""
        m = re.match(r"(?i)^\s*total\s*:\s*(.*)$", joined)
        if m:
            label = m.group(1).split("|")[0].strip()
            # Section subtotals — not scheme AUM
            if re.match(
                r"(?i)^(others?|equity|debt|money\s*market|gold|silver|securiti|"
                r"listed|unlisted|short\s*term|corporate\s+debt|instruments?).*$",
                label,
            ) and not re.search(r"(?i)\b(fund|etf|plan|scheme)\b", label):
                continue
            if not label:
                continue
            if re.search(r"(?i)\b(fund|etf|plan|scheme)\b", label):
                score += 10
            if re.search(r"(?i)\buti\b", label):
                score += 5
        elif re.match(r"(?i)^\s*grand[\s_]*totals?\b", first):
            score += 20
            label = first
        elif re.match(r"(?i)^\s*total\s+net\s+assets?\b", first):
            score += 18
            label = first
        elif re.match(r"(?i)^\s*net\s+assets?\s*(as\s+on\b|/|\(|$)", first):
            score += 12
            label = first
        else:
            continue
        nums = [_parse_number(c) for c in cells]
        nums = [n for n in nums if n is not None]
        if not nums:
            continue
        # Prefer the market-value-sized figure (not the trailing 1.0 / 100 %NAV).
        mv_candidates = [n for n in nums if abs(n) > 1.5]
        mv = mv_candidates[0] if mv_candidates else nums[0]
        if mv > 0:
            score += 2
        if best is None or score > best[0] or (score == best[0] and abs(mv) > abs(best[1])):
            best = (score, mv)
    return best[1] if best else None


def absorb_rounding_residual(
    holdings: list[Holding],
) -> tuple[list[Holding], dict[str, Any]]:
    """Attach tiny %NAV residue (Axis/Tata rounding) onto the cash / NCA line."""
    meta: dict[str, Any] = {"rounding_residual_absorbed": 0.0}
    if not holdings:
        return holdings, meta
    weights = [(_parse_number(h.pct_nav), i) for i, h in enumerate(holdings)]
    present = [(w, i) for w, i in weights if w is not None]
    if not present:
        return holdings, meta
    total = sum(w for w, _ in present)
    # Fraction books stay for normalize_fractional; only polish percent books.
    if max(abs(w) for w, _ in present) <= 1.5:
        return holdings, meta
    if not (99.50 <= total < 99.95):
        return holdings, meta
    residual = 100.0 - total
    cash_i = None
    for i in range(len(holdings) - 1, -1, -1):
        name = (holdings[i].instrument or "").strip()
        if re.search(
            r"(?i)^(?:net\s+(?:current\s+(?:assets?|liabilit)|receivables)|nca\b|"
            r"cash\s*/\s*net\s+current|cash\b)",
            name,
        ):
            cash_i = i
            break
    if cash_i is None:
        return holdings, meta
    h = holdings[cash_i]
    cur = _parse_number(h.pct_nav) or 0.0
    new_pct = f"{(cur + residual):.6f}".rstrip("0").rstrip(".")
    raw = dict(h.raw or {})
    raw["pct_rounding_residual"] = f"{residual:.6f}".rstrip("0").rstrip(".")
    holdings = list(holdings)
    holdings[cash_i] = replace(h, pct_nav=new_pct, raw=raw)
    meta["rounding_residual_absorbed"] = residual
    return holdings, meta


def rebase_pct_nav_from_market_value(
    holdings: list[Holding],
    portfolio_total: float | None,
) -> tuple[list[Holding], dict[str, Any]]:
    """Rebuild %NAV from market value / GRAND when disclosed weights are unusable.

    Cases:
    - Axis-style ``$0.00%`` / missing % while MVs still sum to Grand.
    - DSP-style books where printed ``% to Net Assets`` is vs investments
      (ex–net receivables) or otherwise disagrees with ``MV / Grand``, so the
      cash/NCA line is ``*`` or double-counts once implied — keep NCA via MV.
    Arb disclosures often set Grand = sum of positive MVs only (shorts listed
    separately); treat either net-MV or positive-MV reconciliation as OK.
    """
    meta: dict[str, Any] = {"pct_rebased_from_mv": False}
    if not holdings or not portfolio_total or portfolio_total <= 0:
        return holdings, meta
    mvs: list[tuple[Holding, float]] = []
    for h in holdings:
        mv = _parse_number(h.market_value)
        if mv is None:
            return holdings, meta  # incomplete MV book — don't rebase
        mvs.append((h, mv))
    mv_sum = sum(mv for _, mv in mvs)
    mv_pos = sum(mv for _, mv in mvs if mv >= 0)
    if mv_sum == 0 and mv_pos == 0:
        return holdings, meta
    rel = 0.0005
    reconciled = (
        abs(mv_sum - portfolio_total) / portfolio_total <= rel
        or abs(mv_pos - portfolio_total) / portfolio_total <= rel
    )
    if not reconciled:
        return holdings, meta

    has_broken_pct = any(
        (h.pct_nav or "").strip().startswith("$") or _parse_number(h.pct_nav) is None
        for h, _ in mvs
    )
    weights = [_parse_number(h.pct_nav) for h, _ in mvs]
    present = [w for w in weights if w is not None]
    total_pct = None
    if present:
        total = sum(present)
        total_pct = total * 100.0 if max(abs(x) for x in present) <= 1.5 else total

    qc_ok = meets_allocation_qc(allocation_totals(holdings))
    if has_broken_pct:
        # Axis: only polish broken markers when the printed book is still short.
        if total_pct is not None and 99.90 <= abs(total_pct) <= 100.10:
            return holdings, meta
    elif qc_ok:
        return holdings, meta

    out: list[Holding] = []
    for h, mv in mvs:
        pct = f"{(mv / portfolio_total) * 100.0:.6f}".rstrip("0").rstrip(".")
        raw = dict(h.raw or {})
        raw["pct_rebased_from_mv"] = "1"
        out.append(replace(h, pct_nav=pct, raw=raw))
    meta["pct_rebased_from_mv"] = True
    return out, meta


def _cell_has_numeric(val: str) -> bool:
    if not val or NIL_RE.match(val):
        return False
    return bool(re.search(r"\d", val))


def _row_has_holding_metrics(
    cells: list[str],
    *,
    idx_qty: int | None,
    idx_mv: int | None,
    idx_pct: int | None,
) -> bool:
    """True when a row carries qty / market value / %NAV (not a section title)."""
    qty = _cell(cells, idx_qty)
    mv = _cell(cells, idx_mv)
    pct = _cell(cells, idx_pct)
    if _cell_has_numeric(qty) or _cell_has_numeric(mv):
        return True
    if pct and (pct == "*" or _cell_has_numeric(pct)):
        return True
    # Compact cash rows: Name | mv | pct with no aligned columns
    trailing = 0
    for c in cells[1:]:
        if c and re.fullmatch(r"[-+]?\d[\d,]*\.?\d*", c.replace(",", "")):
            trailing += 1
    return trailing >= 1


def looks_like_header(values: list[str]) -> bool:
    text = " | ".join(v for v in values if v)
    if not text:
        return False
    return sum(1 for rx in HEADER_HINT_RES if rx.search(text)) >= 2


def looks_like_new_table_header(values: list[str]) -> bool:
    """True only for a real column header row (e.g. UTI annex), not section banners."""
    nonempty = [v.strip() for v in values if (v or "").strip()]
    if len(nonempty) < 3:
        return False
    if not looks_like_header(values):
        return False
    joined = " | ".join(nonempty)
    has_name = bool(
        re.search(r"(?i)name\s+of\s+(the\s+)?(instrument|security)|^\s*(instrument|security)\b", joined)
    )
    has_isin = bool(re.search(r"(?i)\bisin\b", joined))
    has_pct = bool(re.search(r"(?i)%\s*(to|of|age)|percent", joined))
    return has_name and (has_isin or has_pct)


def header_row_score(values: list[str]) -> int:
    """Rank candidate header rows — prefer ISIN/%NAV columns over marketing blurbs.

    Shriram debt factsheets put "money market instrument" / "market" in product
    suitability bullets above the real ``Name of Instrument | ISIN | …`` header.
    """
    text = " | ".join(v for v in values if (v or "").strip())
    if not text or not looks_like_header(values):
        return -1
    score = sum(1 for rx in HEADER_HINT_RES if rx.search(text))
    if re.search(r"(?i)\bisin\b", text):
        score += 8
    if re.search(r"(?i)%\s*(to|of)|percent|%\s*nav|%\s*to\s*net", text):
        score += 6
    if re.search(r"(?i)name\s+of\s+(the\s+)?(instrument|security)", text):
        score += 4
    if re.search(r"(?i)\bquantity\b", text):
        score += 2
    # Marketing / suitability noise that still trips instrument+market hints.
    if re.search(
        r"(?i)(?:^[•\-\*]|\bseeking\b|investors?\s+who|consult\s+their|"
        r"suitable\s+for|riskometer|potential\s+risk\s+class)",
        text,
    ):
        score -= 12
    return score


def choose_header_row(rows: list[list[str]]) -> tuple[int, list[str]] | None:
    """Pick the first strong holdings header (ISIN/%NAV), not marketing blurbs.

    Falls back to the highest-scoring weak match so older templates still parse.
    """
    best: tuple[int, int, list[str]] | None = None  # score, idx, row
    for idx, row in enumerate(rows[:120]):
        score = header_row_score(row)
        if score < 2:
            continue
        # Prefer the first "real" SEBI header (ISIN / %NAV enriched → score ≥ 10).
        if score >= 10:
            return idx, row
        if best is None or score > best[0]:
            best = (score, idx, row)
    if not best:
        return None
    return best[1], best[2]


def alias_headers(headers: list[str]) -> list[str]:
    used: set[str] = set()
    out: list[str] = []
    for h in headers:
        key = h.strip() or "col"
        canon = None
        for name, rx in COLUMN_ALIASES.items():
            if name in used:
                continue
            if rx.search(key):
                canon = name
                break
        if canon:
            used.add(canon)
            out.append(canon)
        else:
            base = key
            n = 2
            while key in out:
                key = f"{base}_{n}"
                n += 1
            out.append(key)
    return out


def _cell(row: list[str], idx: int | None) -> str:
    if idx is None or idx < 0 or idx >= len(row):
        return ""
    return (row[idx] or "").strip()


def _find(aliases: list[str], name: str) -> int | None:
    try:
        return aliases.index(name)
    except ValueError:
        return None


_MONTH_NUM = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
    "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9, "oct": 10,
    "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}
_MONTH_RE = (
    r"jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|"
    r"jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?"
)
# Prefer labeled portfolio dates over inception / dividend / holding maturities.
_AS_OF_LABEL_RE = re.compile(
    r"(?i)(?:as\s*on|as\s*of|as\s*at|month\s+ended|period\s+ended|"
    r"(?:monthly|fortnightly|half[\s-]?yearly)?\s*portfolio\s+statement|"
    r"portfolio\s+(?:disclosure\s+)?as\s+o[nf]|holdings?\s+as\s+o[nf])"
)


def _iso_date(year: int, month: int, day: int) -> str | None:
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return None


def _expand_year(y: str) -> int:
    n = int(y)
    if n < 100:
        return 2000 + n
    return n


def parse_as_of(text: str) -> str | None:
    """Parse a portfolio as-of date to ISO YYYY-MM-DD (month-only → last day)."""
    s = (text or "").replace("_", " ")
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"(?i)(\d)(?:st|nd|rd|th)\b", r"\1", s)

    m = re.search(r"\b(20\d{2})-(\d{2})-(\d{2})\b", s)
    if m:
        return _iso_date(int(m.group(1)), int(m.group(2)), int(m.group(3)))

    m = re.search(rf"\b(\d{{1,2}})[-/ .]({_MONTH_RE})[-/ . ]?(20\d{{2}}|\d{{2}})\b", s, re.I)
    if m:
        mon = _MONTH_NUM.get(m.group(2).lower())
        if mon:
            return _iso_date(_expand_year(m.group(3)), mon, int(m.group(1)))

    m = re.search(rf"\b({_MONTH_RE})[-/ .]+(\d{{1,2}})[-/ .,]+(20\d{{2}}|\d{{2}})\b", s, re.I)
    if m:
        mon = _MONTH_NUM.get(m.group(1).lower())
        if mon:
            return _iso_date(_expand_year(m.group(3)), mon, int(m.group(2)))

    m = re.search(r"\b(\d{1,2})[./-](\d{1,2})[./-]((?:20)?\d{2})\b", s)
    if m:
        a, b, y = int(m.group(1)), int(m.group(2)), _expand_year(m.group(3))
        if a > 12 and 1 <= b <= 12:
            return _iso_date(y, b, a)
        if b > 12 and 1 <= a <= 12:
            return _iso_date(y, a, b)
        if 1 <= a <= 12 and 1 <= b <= 12:
            return _iso_date(y, b, a)  # India DMY
        return None

    m = re.search(rf"\b({_MONTH_RE})\s+(20\d{{2}})\b", s, re.I)
    if m:
        mon = _MONTH_NUM.get(m.group(1).lower())
        if mon:
            y = int(m.group(2))
            return _iso_date(y, mon, calendar.monthrange(y, mon)[1])
    return None


def extract_as_of_from_rows(rows: list[list[str]]) -> str | None:
    """Read as-of from banner rows (as on / as of / month ended), not holding dates."""
    for row in rows[:40]:
        joined = " | ".join(x for x in row if x)
        if not joined or not _AS_OF_LABEL_RE.search(joined):
            continue
        got = parse_as_of(joined)
        if got:
            return got
        for cell in row:
            got = parse_as_of(cell or "")
            if got:
                return got
    return None


def extract_as_of(
    rows: list[list[str]],
    *,
    filename: str | None = None,
) -> str | None:
    """Sheet banner first, then filename (double-check when the date is in the file name)."""
    got = extract_as_of_from_rows(rows)
    if got:
        return got
    if filename:
        return parse_as_of(filename)
    return None


def extract_scheme_name_cams(rows: list[list[str]]) -> str | None:
    for row in rows[:25]:
        joined = " | ".join(x for x in row if x)
        if re.search(r"(?i)scheme\s*name\s*:", joined):
            for i, cell in enumerate(row):
                if re.search(r"(?i)scheme\s*name", cell or ""):
                    # value in next non-empty cell
                    for j in range(i + 1, len(row)):
                        if row[j] and not re.search(r"(?i)scheme\s*name", row[j]):
                            return row[j].strip()
            # fallback: after colon in joined
            m = re.search(r"(?i)scheme\s*name\s*:\s*(.+)$", joined)
            if m:
                return m.group(1).split("|")[0].strip()
        # Choice / CAMS: "Name of the scheme" label in col, fund name in next col
        for i, cell in enumerate(row):
            if re.search(r"(?i)^name\s+of\s+the\s+scheme$", (cell or "").strip()):
                for j in range(i + 1, len(row)):
                    val = (row[j] or "").strip()
                    if val and not re.search(r"(?i)^name\s+of", val):
                        return val
    return None


def extract_title_scheme(rows: list[list[str]], sheet_name: str) -> str:
    """NJ / Angel / Abakkus / Navi / Union: fund name above the holdings header."""
    for row in rows[:12]:
        for cell in row:
            t = (cell or "").strip()
            if not t or len(t) < 6:
                continue
            # Union: "MONTHLY PORTFOLIO STATEMENT OF {FUND} AS ON …"
            m = re.search(
                r"(?i)portfolio\s+statement\s+of\s+(.+?)\s+as\s+on\b",
                t,
            )
            if m:
                title = m.group(1).strip(" -–—")
                if title and not re.search(r"(?i)registration\s+no", title):
                    return title
            if re.search(
                r"(?i)registration\s+no|registered\s+office|toll\s+free|"
                r"mutual\s+fund$|open\s+ended|name\s+of\s+the\s+instrument|"
                r"portfolio\s+(statement|as\s+on)",
                t,
            ):
                continue
            # Include FOF / index-style titles that omit the word "Fund" (Navi, etc.)
            if re.search(
                r"(?i)\b(fund|etf|fof|index|scheme|liquid|overnight|arbitrage|"
                r"momentum|quality|nifty|sensex|nasdaq)\b",
                t,
            ):
                if not looks_like_header([t]):
                    return t
    return sheet_name.strip() or "unknown"


def parse_holdings_table(
    rows: list[list[str]],
    *,
    prefer_leading_code: bool = False,
) -> tuple[list[Holding], dict[str, Any]]:
    """Parse SEBI-style holdings table under a detected header row."""
    picked = choose_header_row(rows)
    meta: dict[str, Any] = {"header_row": None}
    if not picked:
        return [], meta
    hidx, headers_raw = picked
    meta["header_row"] = hidx
    headers_raw = [h if h else f"col_{i+1}" for i, h in enumerate(headers_raw)]
    aliases = alias_headers(headers_raw)
    meta["headers"] = aliases

    idx_instr = _find(aliases, "instrument")
    idx_isin = _find(aliases, "isin")
    idx_ind = _find(aliases, "industry")
    idx_qty = _find(aliases, "quantity")
    idx_mv = _find(aliases, "market_value")
    idx_pct = _find(aliases, "pct_nav")
    idx_ytm = _find(aliases, "ytm")
    idx_ytc = _find(aliases, "ytc")
    idx_code = _find(aliases, "security_code")
    extra_idx = {k: _find(aliases, k) for k in EXTRA_HOLDING_FIELDS}
    if idx_ind is None:
        idx_ind = extra_idx.get("industry_rating")

    # CAMS/Choice/SBI/Abakkus often put issuer code in col0 with no header label.
    header_shifted = False
    if idx_code is None and prefer_leading_code:
        if idx_instr == 0 and idx_isin == 2:
            idx_code, idx_instr = 0, 1
            header_shifted = True
        elif idx_instr is None and idx_isin == 2:
            idx_code, idx_instr = 0, 1
            header_shifted = True
        elif idx_instr == 1 and idx_isin == 2:
            idx_code = 0

    holdings: list[Holding] = []
    section = ""
    portfolio_total = extract_portfolio_total_mv(rows)
    meta["portfolio_total_mv"] = portfolio_total
    implied_pct_count = 0
    data_rows = rows[hidx + 1 :]
    row_i = 0
    while row_i < len(data_rows):
        row = data_rows[row_i]
        row_i += 1
        cells = [(c or "").strip() for c in row]
        if not any(cells):
            continue
        joined = " | ".join(c for c in cells if c)
        first = next((c for c in cells if c), "")

        # Annex / legend tables after the portfolio end — stop.
        if STOP_TABLE_RE.search(first) or STOP_TABLE_RE.search(joined[:120]):
            break

        # Second column header: Tata debt-after-equity (holdings-like) → rebind.
        # FUTURES/OPTIONS hedge table → rebind. Other annex tables → stop.
        if holdings and looks_like_new_table_header(cells):
            headers_raw = [h if h else f"col_{i+1}" for i, h in enumerate(cells)]
            new_aliases = alias_headers(headers_raw)
            holdings_like = (
                _find(new_aliases, "pct_nav") is not None
                and _find(new_aliases, "instrument") is not None
                and (
                    _find(new_aliases, "isin") is not None
                    or _find(new_aliases, "market_value") is not None
                )
            )
            deriv = False
            if not holdings_like:
                for look in data_rows[row_i : row_i + 8]:
                    look_first = next(((c or "").strip() for c in look if (c or "").strip()), "")
                    if not look_first:
                        continue
                    if re.match(r"(?i)^(futures?|options?)\s*$", look_first):
                        deriv = True
                        break
                    if looks_like_new_table_header([(c or "").strip() for c in look]):
                        break
                    if STOP_TABLE_RE.search(look_first) or re.search(r"(?i)^total\s*:", look_first):
                        break
            if holdings_like or deriv:
                aliases = new_aliases
                idx_instr = _find(aliases, "instrument")
                idx_isin = _find(aliases, "isin")
                idx_ind = _find(aliases, "industry")
                idx_qty = _find(aliases, "quantity")
                idx_mv = _find(aliases, "market_value")
                idx_pct = _find(aliases, "pct_nav")
                idx_ytm = _find(aliases, "ytm")
                idx_ytc = _find(aliases, "ytc")
                idx_code = _find(aliases, "security_code")
                extra_idx = {k: _find(aliases, k) for k in EXTRA_HOLDING_FIELDS}
                if idx_ind is None:
                    idx_ind = extra_idx.get("industry_rating")
                header_shifted = False
                if idx_code is None and prefer_leading_code:
                    if idx_instr == 0 and idx_isin == 2:
                        idx_code, idx_instr = 0, 1
                        header_shifted = True
                    elif idx_instr is None and idx_isin == 2:
                        idx_code, idx_instr = 0, 1
                        header_shifted = True
                    elif idx_instr == 1 and idx_isin == 2:
                        idx_code = 0
                section = ""
                continue
            break

        # Section banners (no ISIN). Holding rows that merely *start with* a section
        # keyword (e.g. "GOLD BULLION", "SILVER BULLION") must keep their weights.
        is_cash_weight_row = bool(
            re.search(
                r"(?i)^(?:net\s+current\s+(?:assets?|liabilit)|net\s+receivables|"
                r"cash\s*/\s*net\s+current)",
                first,
            )
            and _row_has_holding_metrics(cells, idx_qty=idx_qty, idx_mv=idx_mv, idx_pct=idx_pct)
        )
        has_metrics = _row_has_holding_metrics(
            cells, idx_qty=idx_qty, idx_mv=idx_mv, idx_pct=idx_pct
        )
        if (
            not ISIN_RE.search(joined)
            and SECTION_RE.search(first)
            and not is_cash_weight_row
            and not has_metrics
        ):
            if not SKIP_HOLDING_NAME_RE.match(first):
                section = first
            continue
        if (
            not ISIN_RE.search(joined)
            and SECTION_RE.search(joined)
            and len([c for c in cells if c]) <= 2
            and not is_cash_weight_row
            and not has_metrics
        ):
            section = first
            continue

        isin = _cell(cells, idx_isin)
        m = ISIN_RE.search(isin) or ISIN_RE.search(joined)
        if m:
            isin = m.group(1)
        else:
            isin = ""

        instrument = _cell(cells, idx_instr)
        code = _cell(cells, idx_code)
        # Unlabeled leading code immediately before the instrument column.
        if prefer_leading_code and not code and idx_instr and idx_instr > 0:
            prev = cells[idx_instr - 1] if idx_instr - 1 < len(cells) else ""
            if prev and re.fullmatch(r"[A-Za-z0-9./\-]{1,16}", prev) and not ISIN_RE.search(prev):
                code = prev

        industry = _cell(cells, idx_ind)
        qty = _cell(cells, idx_qty)
        mv = _cell(cells, idx_mv)
        pct = _cell(cells, idx_pct)
        ytm = _cell(cells, idx_ytm)
        ytc = _cell(cells, idx_ytc)
        extras = {k: _cell(cells, extra_idx.get(k)) for k in EXTRA_HOLDING_FIELDS}
        if not industry:
            industry = extras.get("industry_rating") or ""

        # Compact SEBI "Others" line: Name | Market Value | % to NAV (no ISIN / qty cols).
        # Edelweiss etc. put Yield in a trailing column (Name|…|MV|%|Yield) — prefer the
        # header-aligned MV/% when both already parse; otherwise take the last two nums.
        if (
            not isin
            and re.search(r"(?i)^net\s+current\s+assets|^net\s+receivables", instrument or first)
        ):
            nums: list[str] = []
            for c in cells[1:]:
                if not c or NIL_RE.match(c):
                    continue
                if re.fullmatch(r"[-+]?\d[\d,]*\.?\d*", c.replace(",", "")):
                    nums.append(c.replace(",", ""))
            hdr_mv_ok = _parse_number(mv) is not None
            hdr_pct_ok = _parse_number(pct) is not None
            if len(nums) >= 2 and not (hdr_mv_ok and hdr_pct_ok):
                mv, pct = nums[-2], nums[-1]
                industry, qty = "", ""
            elif len(nums) == 1 and not pct:
                pct = nums[0]
                industry, qty = "", ""
            if not instrument:
                instrument = first
        # Unlabeled leading code: "100006 | HDFC Bank | INE…" while header says Name|ISIN|…
        # Do NOT remap when the labeled ISIN column already holds an ISIN (ICICI/Nippon style
        # sheets where short issuer names like "NABARD" falsely match the code pattern).
        if (
            prefer_leading_code
            and not header_shifted
            and instrument
            and re.fullmatch(r"[A-Za-z0-9./\-]{1,16}", instrument)
            and not ISIN_RE.search(instrument)
            and len(cells) > 2
            and ISIN_RE.search(joined)
            and not (idx_isin is not None and ISIN_RE.search(_cell(cells, idx_isin)))
        ):
            maybe_name = cells[1] if len(cells) > 1 else ""
            maybe_isin_cell = cells[2] if len(cells) > 2 else ""
            if maybe_name and (ISIN_RE.search(maybe_isin_cell) or ISIN_RE.search(joined)):
                code = instrument
                instrument = maybe_name
                if not isin:
                    mm = ISIN_RE.search(maybe_isin_cell) or ISIN_RE.search(joined)
                    isin = mm.group(1) if mm else ""
                industry = cells[3] if len(cells) > 3 else industry
                qty = cells[4] if len(cells) > 4 else qty
                mv = cells[5] if len(cells) > 5 else mv
                pct = cells[6] if len(cells) > 6 else pct
                ytm = cells[7] if len(cells) > 7 else ytm
                ytc = cells[8] if len(cells) > 8 else ytc
        elif not instrument:
            for c in cells:
                if c and not ISIN_RE.fullmatch(c) and not re.fullmatch(r"[\d.]+", c):
                    if not NIL_RE.match(c):
                        instrument = c
                        break

        if not instrument or SKIP_HOLDING_NAME_RE.match(instrument) or NIL_RE.match(instrument):
            # Quant arb: weighted Sub Total under Index/Stock Futures with no child legs.
            # Trust: weighted Sub Total under CBLO/TREPS/Reverse Repo with no child legs.
            if (
                instrument
                and re.match(r"(?i)^sub\s*total\b", instrument)
                and re.search(
                    r"(?i)\bfutures?\b|treps?|tri[\s\-]?party|reverse\s+repo|\bcblo\b",
                    section or "",
                )
                and _row_has_holding_metrics(cells, idx_qty=idx_qty, idx_mv=idx_mv, idx_pct=idx_pct)
                and not any(
                    (h.section or "") == section
                    and not re.match(r"(?i)^(?:sub\s*)?totals?\b", (h.instrument or "").strip())
                    for h in holdings
                )
            ):
                instrument = (section or instrument).strip()
            else:
                continue
        if (
            SECTION_RE.match(instrument)
            and not isin
            and not re.search(
                r"(?i)^(?:\(?[a-z0-9]+\)\s+)?(?:net\s+current|net\s+receivables|nca\b|"
                r"treps|tri[- ]?party|reverse\s+repo|cblo|cash)\b"
                r"|^\(?[a-z0-9]+\)\s+repo\b",
                instrument,
            )
            and not has_metrics
        ):
            section = instrument
            continue
        # ICICI-style category / sleeve subtotals: "Certificate of Deposits | 2219971 | 0.64".
        # Bare "Reverse Repo | mv | %" is an ICICI *sleeve total* only when dated
        # "Reverse Repo (…)" children follow; otherwise (Nippon) keep it as a holding.
        if (
            not isin
            and not qty
            and _parse_number(mv) is not None
            and (_parse_number(pct) is not None or (pct or "").strip() == "*")
            and not re.search(
                # Keep ICICI "Reverse Repo" on the sleeve path (drop when dated
                # children follow). Only exempt bare cash/TREPS and Tata "J) REPO".
                r"(?i)^(?:\(?[a-z0-9]+\)\s+)?(?:"
                r"treps|tri[- ]?party|cblo|cash|"
                r"net\s+current|net\s+receivables|nca\b"
                r")\b"
                r"|^\(?[a-z0-9]+\)\s+repo\b",
                instrument,
            )
        ):
            if re.match(r"(?i)^reverse\s+repo\s*$", instrument):
                has_dated_repo = False
                for look in data_rows[row_i : row_i + 30]:
                    look_cells = [(c or "").strip() for c in look]
                    if not any(look_cells):
                        continue
                    look_instr = _cell(look_cells, idx_instr) if idx_instr is not None else ""
                    look_name = look_instr or next((c for c in look_cells if c), "")
                    if re.match(r"(?i)^reverse\s+repo\s*\(", look_name):
                        has_dated_repo = True
                        break
                    if re.match(
                        r"(?i)^(total|sub\s*total|grand|triparty|treps|"
                        r"corporate\s+bond|money\s+market|others?|net\s+current|"
                        r"cash\s+margin|units\s+of)\b",
                        look_name,
                    ):
                        break
                if has_dated_repo:
                    section = instrument
                    continue
                # else fall through — Nippon-style aggregate Reverse Repo holding
            elif re.search(r"(?i)\bfutures?\b|\boptions?\b|derivative", instrument):
                # Quant sleeve: Sub Total promoted to "(a) Index / Stock Futures"
                pass
            elif (
                (
                    # Keep HSBC/Nippon-style "Reverse Repo(s) | MV | %" as holdings.
                    # Singular "Reverse Repo" already fell through above when no dated kids.
                    SECTION_RE.search(instrument)
                    and not re.match(r"(?i)^reverse\s+repos?\b", instrument)
                )
                or (
                    # Category banners like "Debt Securities" — but not coupon lines /
                    # Invesco "5.35% Repo in Corporate Debt Securities".
                    re.search(
                        r"(?i)instruments?$|exchanges?$|securities$|bills$|bonds?$|"
                        r"papers?$|deposits?$|debentures?$|\bfunds?\b|\baif\b",
                        instrument,
                    )
                    and not re.search(
                        r"(?i)^\d+(\.\d+)?\s*%|\brepo\s+in\b",
                        instrument,
                    )
                )
            ):
                section = instrument
                continue
        if STOP_TABLE_RE.search(instrument):
            break
        if isin and NIL_RE.match(isin):
            continue
        # Drop empty NIL placeholder rows (common in CAMS templates)
        if not isin and (NIL_RE.match(mv or "") or not mv) and (NIL_RE.match(pct or "") or not pct) and not qty:
            continue
        if not isin and not qty and (NIL_RE.match(mv or "") or NIL_RE.match(pct or "")):
            continue
        if not isin and not (qty or (mv and not NIL_RE.match(mv)) or (pct and not NIL_RE.match(pct))):
            continue
        # Annex-like margin tables under a receivables *section* — but keep real NCA /
        # TREPS / reverse-repo cash lines (Trust puts TREPS under "CBLO / Reverse Repo / TREPS").
        if (
            not isin
            and section
            and re.search(r"(?i)receivable|margin|contract|reverse\s+repo|treps?|cblo", section)
            and not re.search(
                r"(?i)^(?:nca\b|net\s+current|net\s+receivables|other\s+receivables|"
                r"cash(?:\s*/\s*net\s+current)?\b|"
                r"(?:\(?[a-z0-9]+\)\s+)?(?:treps?\b|tri[\s\-]?party|reverse\s+repo|cblo)|"
                r"\(?[a-z0-9]+\)\s+repo\b|"
                r"clearing)\b",
                instrument,
            )
            and not re.search(r"(?i)\bfutures?\b|\boptions?\b|derivative", section)
            and not re.search(r"(?i)treps?|tri[\s\-]?party|reverse\s+repo|\bcblo\b", section)
        ):
            continue

        # UTI etc. often print "*" for tiny OR negative NCA weights; recover from MV / AUM.
        # DSP fractional sheets need fraction-of-1 (normalize ×100 later); percent sheets need ×100.
        if (not pct or pct.strip() == "*") and portfolio_total:
            mv_n = _parse_number(mv or "")
            if mv_n is not None and portfolio_total != 0:
                prior = [_parse_number(h.pct_nav) for h in holdings]
                prior = [w for w in prior if w is not None]
                sheet_is_fraction = bool(prior) and max(abs(w) for w in prior) <= 1.5
                implied = mv_n / portfolio_total * (1.0 if sheet_is_fraction else 100.0)
                pct = f"{implied:.10f}".rstrip("0").rstrip(".")
                implied_pct_count += 1

        raw = {
            aliases[i] if i < len(aliases) else f"col_{i+1}": cells[i]
            for i in range(len(cells))
            if cells[i]
        }
        holdings.append(
            Holding(
                instrument=instrument,
                isin=isin,
                industry=industry,
                quantity=qty,
                market_value="" if NIL_RE.match(mv or "") else mv,
                pct_nav="" if NIL_RE.match(pct or "") else pct,
                ytm=ytm,
                ytc=ytc,
                security_code=code,
                section=section,
                raw=raw,
                **extras,
            )
        )
    holdings, nca_meta = collapse_duplicate_net_receivables(holdings)
    holdings, margin_meta = drop_ccil_margin_when_nca_present(holdings)
    holdings, nca_rollup_meta = drop_net_current_assets_rollup(holdings)
    holdings, junk_meta = apply_zero_aum_junk_weights(holdings, portfolio_total)
    holdings, scale_meta = normalize_fractional_pct_nav(holdings)
    holdings, rebase_meta = rebase_pct_nav_from_market_value(holdings, portfolio_total)
    holdings, round_meta = absorb_rounding_residual(holdings)
    holdings, cash_meta = normalize_cash_position_names(holdings)
    meta["holding_count"] = len(holdings)
    meta["implied_pct_from_mv"] = implied_pct_count
    meta.update(nca_meta)
    meta.update(margin_meta)
    meta.update(nca_rollup_meta)
    meta.update(junk_meta)
    meta.update(scale_meta)
    meta.update(rebase_meta)
    meta.update(round_meta)
    meta.update(cash_meta)
    return holdings, meta


# TREPS / CCIL / reverse repo / NCA / cash margin — one instrument name: Cash.
_CASH_POSITION_RE = re.compile(
    r"(?i)treps?|tri[\s\-]?party|"
    r"reverse\s+repos?|\bcblo\b|"
    r"clearing\s+corporation|\bccil\b|"
    r"amc\s+repo\s+clearing|"
    r"net\s+current\s+assets?|\bnca\b|"
    r"net\s+receivables?|net\s+payables?|"
    r"receivable\s*/\s*\(?\s*payable|payables?\s*/\s*\(?\s*receivable|"
    r"cash\s+margin|margin\s+money|cash\s*/\s*bank|cash\s+and\s+other|"
    r"^\s*cash\s*$|^\s*cash\s*/|"
    r"\brepos?\b|"
    r"^\s*trp[_-]|^\s*rep\d+"
)
_NOT_CASH_POSITION_RE = re.compile(
    r"(?i)interest\s+rate\s+swaps?|\birs\b|(?<![A-Za-z])ois(?![A-Za-z])|"
    r"t[\s\-]?bill|treasury\s+bill|"
    r"commercial\s+paper|certificate\s+of\s+deposit|"
    r"\bdebenture\b|\bncd\b"
)


def is_cash_position(h: Holding) -> bool:
    """True for TREPS, CCIL, reverse repo, cash, and net current assets."""
    name = (h.instrument or "").strip()
    sec = (h.section or "").strip()
    if re.search(r"(?i)interest\s+rate\s+swaps?|\birs\b|(?<![A-Za-z])ois(?![A-Za-z])", name):
        return False
    if _NOT_CASH_POSITION_RE.search(name) and not _CASH_POSITION_RE.search(name):
        return False
    if _CASH_POSITION_RE.search(name):
        return True
    if _CASH_POSITION_RE.search(sec) and not (h.isin or "").strip():
        if re.search(r"(?i)\d+(?:\.\d+)?\s*%", name) and not re.search(r"(?i)repo", name):
            return False
        return True
    return False


def normalize_cash_position_names(
    holdings: list[Holding],
) -> tuple[list[Holding], dict[str, Any]]:
    """Rename cash-equivalent legs to instrument=Cash, section=Cash."""
    meta: dict[str, Any] = {"cash_positions_renamed": 0}
    out: list[Holding] = []
    for h in holdings:
        if is_cash_position(h) and (
            (h.instrument or "").strip() != "Cash" or (h.section or "").strip() != "Cash"
        ):
            meta["cash_positions_renamed"] += 1
            h = replace(h, instrument="Cash", section="Cash")
        out.append(h)
    return out, meta


def collapse_duplicate_net_receivables(holdings: list[Holding]) -> tuple[list[Holding], dict[str, Any]]:
    """Mahindra prints Net Receivables twice (pre/post margin) with ~same %NAV — keep last."""
    meta: dict[str, Any] = {"duplicate_nca_collapsed": 0}
    if not holdings:
        return holdings, meta
    out: list[Holding] = []
    for h in holdings:
        if (
            out
            and re.search(r"(?i)^net\s+receivables", (h.instrument or "").strip())
            and re.search(r"(?i)^net\s+receivables", (out[-1].instrument or "").strip())
        ):
            out[-1] = h
            meta["duplicate_nca_collapsed"] += 1
            continue
        out.append(h)
    return out, meta


def drop_ccil_margin_when_nca_present(holdings: list[Holding]) -> tuple[list[Holding], dict[str, Any]]:
    """Mahindra lists CCIL/ARCL margin separately even though NCA already includes it."""
    meta: dict[str, Any] = {"ccil_margin_dropped": 0}
    has_nca = any(re.search(r"(?i)^net\s+receivables", (h.instrument or "").strip()) for h in holdings)
    if not has_nca:
        return holdings, meta
    out: list[Holding] = []
    for h in holdings:
        if re.search(r"(?i)^margin\s+placed\s+with\s+(?:ccil|arcl|nse|bse|exchange)", (h.instrument or "").strip()):
            meta["ccil_margin_dropped"] += 1
            continue
        out.append(h)
    return out, meta


def drop_net_current_assets_rollup(holdings: list[Holding]) -> tuple[list[Holding], dict[str, Any]]:
    """Drop Bandhan-style ``Net Current Assets`` when it rolls up cash children already listed.

    Bandhan OTHERS block::
      Cash Margin - CCIL / Derivatives, Cash / Bank Balance, Net Receivables/Payables,
      then Net Current Assets (= sum of those), then GRAND TOTAL.
    ICICI also prints Cash Margin + Net Current Assets, but NCA is a residual (not a
    rollup) — only drop when NCA weight ≈ sum of cash-component weights.
    """
    meta: dict[str, Any] = {"net_current_assets_rollup_dropped": 0}
    if not holdings:
        return holdings, meta

    cash_child = re.compile(
        r"(?i)^(?:cash\s+margin|cash\s*/\s*bank|cash\s+/\s+bank|"
        r"net\s+receivables|bank\s+balance)"
    )
    nca_re = re.compile(r"(?i)^net\s+current\s+assets?\b")
    nr_re = re.compile(r"(?i)^net\s+receivables")

    child_w: list[float] = []
    nca_indices: list[int] = []
    has_nr = False
    child_kinds = 0
    seen_kinds: set[str] = set()
    for i, h in enumerate(holdings):
        name = (h.instrument or "").strip()
        if nca_re.match(name):
            nca_indices.append(i)
            continue
        if nr_re.match(name):
            has_nr = True
        m = cash_child.match(name)
        if m:
            w = _parse_number(h.pct_nav)
            child_w.append(0.0 if w is None else w)
            # Distinct Bandhan child kinds (margin / bank / NR).
            if re.match(r"(?i)^cash\s+margin", name):
                seen_kinds.add("margin")
            elif re.match(r"(?i)^(?:cash\s*/\s*bank|cash\s+/\s+bank|bank\s+balance)", name):
                seen_kinds.add("bank")
            elif nr_re.match(name):
                seen_kinds.add("nr")

    # Bandhan rollup always lists NR (even at 0) plus margins/bank under OTHERS.
    # ICICI often has only Cash Margin + residual NCA with similar magnitude — don't drop.
    if not nca_indices or not child_w or not has_nr or len(seen_kinds) < 2:
        return holdings, meta

    child_sum = sum(child_w)
    # Fraction books still unscaled: both sides near 0–1.
    scale_tol = 0.0005 if max(abs(x) for x in child_w + [child_sum]) <= 1.5 else 0.05

    out: list[Holding] = []
    for i, h in enumerate(holdings):
        if i in nca_indices:
            nca_w = _parse_number(h.pct_nav)
            if nca_w is not None and abs(nca_w - child_sum) <= scale_tol:
                meta["net_current_assets_rollup_dropped"] += 1
                continue
        out.append(h)
    return out, meta


def write_scheme_portfolio(out_root: Path, portfolio: SchemePortfolio) -> Path:
    """Write portfolio.{json,csv} under existing layout; return scheme dir."""
    folder = safe_name(portfolio.shortcode or portfolio.scheme_name)
    dest = out_root / folder
    dest.mkdir(parents=True, exist_ok=True)

    portfolio.holdings, _ = normalize_cash_position_names(portfolio.holdings)
    rows = [asdict(h) for h in portfolio.holdings]
    # Flatten: keep canonical cols; drop nested raw from csv
    csv_fields = [
        "instrument",
        "isin",
        "industry",
        "industry_rating",
        "rating",
        "rating_agency",
        "coupon",
        "maturity_date",
        "residual_maturity",
        "put_call_date",
        "quantity",
        "face_value",
        "market_value",
        "pct_nav",
        "ytm",
        "ytc",
        "instrument_yield",
        "duration",
        "macaulay_duration",
        "modified_duration",
        "accrued_interest",
        "listed_status",
        "futures_price",
        "position_side",
        "margin",
        "market_cap",
        "underlying",
        "asset_class",
        "security_code",
        "section",
    ]
    with (dest / "portfolio.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=csv_fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)

    payload = {
        "meta": portfolio.to_meta(),
        "holdings": rows,
    }
    (dest / "portfolio.json").write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return dest


def write_amc_schemes_index(out_root: Path, portfolios: Iterable[SchemePortfolio]) -> None:
    items = []
    for p in portfolios:
        items.append(
            {
                "scheme": p.scheme_name,
                "shortcode": p.shortcode,
                "folder": safe_name(p.shortcode or p.scheme_name),
                "rows": len(p.holdings),
                "as_of": p.as_of,
                "source_file": p.source_file,
                "sheet_name": p.sheet_name,
            }
        )
    out_root.mkdir(parents=True, exist_ok=True)
    (out_root / "schemes.json").write_text(json.dumps(items, indent=2) + "\n", encoding="utf-8")


def period_from_path(path: Path) -> str:
    # .../monthly/2026-07/amc/file.xlsx or .../latest/...
    parts = path.parts
    for i, p in enumerate(parts):
        if p in {"monthly", "fortnightly"} and i + 1 < len(parts):
            return parts[i + 1]
    return "unknown"


def disclosure_type_from_path(path: Path) -> str:
    parts = set(path.parts)
    if "fortnightly" in parts:
        return "fortnightly"
    return "monthly"
