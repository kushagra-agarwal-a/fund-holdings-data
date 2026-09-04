#!/usr/bin/env python3
"""
Extract scheme-wise portfolio holdings from downloaded AMC monthly disclosures.

Input:
  data/disclosures/monthly/{YYYY-MM}/{amc-id}/*.{xlsx,xls,xlsb,zip}

Output:
  data/parsed/monthly/{YYYY-MM}/{amc-id}/{scheme}/portfolio.csv
  data/parsed/monthly/{YYYY-MM}/{amc-id}/{scheme}/portfolio.json
  data/parsed/monthly/{YYYY-MM}/{amc-id}/schemes.json
  data/parsed/monthly/{YYYY-MM}/_report.json

Coverage-first: light header aliasing, robust sheet/file → scheme mapping.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import tempfile
import zipfile
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except Exception as e:  # pragma: no cover
    raise SystemExit(
        "Missing openpyxl. Install:  .venv/bin/pip install openpyxl xlrd\n" + str(e)
    ) from e

HEADER_HINTS = (
    "scheme",
    "security",
    "instrument",
    "isin",
    "issuer",
    "quantity",
    "qty",
    "market value",
    "market/fair",
    "market_value",
    "face value",
    "nav",
    "% to nav",
    "% to net",
    "percent",
    "holding",
    "rating",
    "industry",
    "sector",
    "coupon",
    "yield",
    "maturity",
    "ytm",
)

SCHEME_COL_HINTS = (
    "scheme",
    "scheme name",
    "name of the scheme",
    "name of scheme",
    "scheme_name",
    "fund name",
    "fund",
)
SKIP_SHEET_RE = re.compile(
    r"^(index|contents|cover|notes?|disclaimer|summary|risk.?o.?meter|"
    r"annexure|instruction|read.?me|legend|glossary|overview)$",
    re.I,
)
SKIP_SHEET_CONTAINS = re.compile(
    r"risk.?o.?meter|disclaimer|notes?\s*to|important\s+information|"
    r"sebi\s+circular|annexure\s*[a-z0-9]*\s*-\s*risk",
    re.I,
)
SKIP_FILE_RE = re.compile(
    r"viewfile\.php|all[-_\s]?schemes|consolidated|combined[-_\s]?portfolio|"
    r"risk[-_\s]?o[-_\s]?meter|exposure\s+to\s+top|"
    r"top\s*\d+\s*holdings(?:\s+by\s+issuer)?|holdings\s+by\s+issuer|"
    r"fortnight",
    re.I,
)
MEGA_FILE_RE = re.compile(r"all[-_\s]?schemes|consolidated|combined", re.I)

# Light canonical aliases (coverage-first; keep original headers too via rename)
COLUMN_ALIASES = {
    "isin": re.compile(r"^\s*isin\b|isin\s*code", re.I),
    "instrument": re.compile(
        r"name\s+of\s+(the\s+)?(instrument|security)|security\s+name|instrument\s+name|"
        r"^(security|instrument|scrip|stock)\b",
        re.I,
    ),
    "quantity": re.compile(r"quantity|qty|no\.?\s*of\s*(shares|units)|units?\b", re.I),
    "market_value": re.compile(
        r"market\s*value|mkt\.?\s*value|value\s*\(.*?(rs|inr|lac|lakh)",
        re.I,
    ),
    "pct_nav": re.compile(
        r"%\s*(to|of)?\s*n\.?a\.?v|percent(age)?\s*(to|of)?\s*nav|%nav|"
        r"%\s*(to|of)\s*net(\s*assets)?",
        re.I,
    ),
    "industry": re.compile(r"industry|sector|rating|coupon", re.I),
}


@dataclass
class AmcResult:
    amc: str
    files_seen: int = 0
    files_parsed: int = 0
    files_skipped: int = 0
    schemes: int = 0
    rows: int = 0
    errors: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)


def safe_name(s: str) -> str:
    out = re.sub(r"[^\w.\-() ]+", "_", (s or "").strip())
    out = re.sub(r"\s+", " ", out).strip(" ._")
    return out[:180] or "unknown"


def norm_cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def looks_like_header(values: list[str]) -> bool:
    text = " | ".join(v.lower() for v in values if v)
    if not text:
        return False
    score = sum(1 for h in HEADER_HINTS if h in text)
    return score >= 2


def choose_header_row(rows: list[list[str]]) -> tuple[int, list[str]] | None:
    for idx, row in enumerate(rows[:60]):
        if looks_like_header(row):
            return idx, row
    return None


def scheme_col_index(headers: list[str]) -> int | None:
    lowered = [h.lower().strip() for h in headers]
    for i, h in enumerate(lowered):
        if any(k == h or k in h for k in SCHEME_COL_HINTS):
            # avoid matching "scheme code" alone when richer name exists later — still ok
            return i
    return None


def alias_headers(headers: list[str]) -> list[str]:
    used: set[str] = set()
    out: list[str] = []
    for h in headers:
        key = h.strip() or "col"
        canon = None
        for name, rx in COLUMN_ALIASES.items():
            if name in used:
                continue
            if rx.search(key):
                canon = name
                break
        if canon:
            used.add(canon)
            out.append(canon)
        else:
            # de-dupe raw headers
            base = key
            n = 2
            while key in out:
                key = f"{base}_{n}"
                n += 1
            out.append(key)
    return out


def is_junk_sheet(name: str) -> bool:
    n = (name or "").strip()
    if not n:
        return True
    if SKIP_SHEET_RE.match(n):
        return True
    if SKIP_SHEET_CONTAINS.search(n) and not re.search(r"portfolio|holding|isin", n, re.I):
        return True
    return False


def file_kind(path: Path) -> str:
    """Return ooxml|ole|zip|other based on magic bytes."""
    try:
        head = path.read_bytes()[:8]
    except OSError:
        return "other"
    if head.startswith(b"PK"):
        # zip or xlsx/xlsb
        if path.suffix.lower() == ".zip":
            return "zip"
        return "ooxml"
    if head.startswith(b"\xd0\xcf\x11\xe0"):
        return "ole"
    return "other"


def list_source_files(amc_dir: Path) -> list[Path]:
    files: list[Path] = []
    for p in sorted(amc_dir.iterdir()):
        if not p.is_file():
            continue
        if p.name in {"manifest.json"}:
            continue
        if p.suffix.lower() not in {".xlsx", ".xls", ".xlsb", ".zip", ".xlsm"}:
            # allow misnamed with spreadsheet magic later
            if p.suffix.lower() in {".php", ".html", ".json", ".txt", ".pdf"}:
                continue
            continue
        files.append(p)

    # Prefer per-scheme files: if many workbooks exist, skip mega "all-schemes" files
    workbooks = [p for p in files if p.suffix.lower() != ".zip"]
    if len(workbooks) >= 10:
        filtered = [p for p in files if not MEGA_FILE_RE.search(p.name)]
        if filtered:
            files = filtered
    return files


def expand_zip(path: Path, dest: Path) -> list[Path]:
    dest.mkdir(parents=True, exist_ok=True)
    out: list[Path] = []
    with zipfile.ZipFile(path) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            name = Path(info.filename).name
            if name.startswith("."):
                continue
            if not re.search(r"\.(xlsx|xls|xlsb|xlsm)$", name, re.I):
                continue
            if SKIP_FILE_RE.search(name) and not re.search(r"portfolio|holding", name, re.I):
                continue
            target = dest / safe_name(name)
            with zf.open(info) as src, target.open("wb") as dst:
                shutil.copyfileobj(src, dst)
            out.append(target)
    return out


def sheet_rows_openpyxl(path: Path) -> list[tuple[str, list[list[str]]]]:
    # read_only: max_row/max_column are often None/1 until fully scanned — do not
    # use them as hard limits or we drop almost every row (seen on NJ workbooks).
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets: list[tuple[str, list[list[str]]]] = []
    for ws in wb.worksheets:
        name = ws.title or ""
        if is_junk_sheet(name):
            continue
        rows: list[list[str]] = []
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i >= 20000:
                break
            cells = list(r[:80]) if r is not None else []
            rows.append([norm_cell(v) for v in cells])
        sheets.append((name, rows))
    try:
        wb.close()
    except Exception:
        pass
    return sheets


def sheet_rows_xlrd(path: Path) -> list[tuple[str, list[list[str]]]]:
    import xlrd

    book = xlrd.open_workbook(path)
    sheets: list[tuple[str, list[list[str]]]] = []
    for ws in book.sheets():
        name = ws.name or ""
        if is_junk_sheet(name):
            continue
        rows: list[list[str]] = []
        cols = min(ws.ncols, 80)
        for r in range(min(ws.nrows, 20000)):
            rows.append([norm_cell(ws.cell_value(r, c)) for c in range(cols)])
        sheets.append((name, rows))
    return sheets


def prepare_workbook_path(path: Path, tmp_dir: Path) -> Path:
    """
    openpyxl decides format partly by extension. OOXML files mislabeled .xls
    must be opened as .xlsx.
    """
    kind = file_kind(path)
    suffix = path.suffix.lower()
    if kind == "ooxml" and suffix in {".xls", ".xlsm", ""}:
        dest = tmp_dir / (safe_name(path.stem) + ".xlsx")
        shutil.copyfile(path, dest)
        return dest
    if kind == "ole" and suffix in {".xlsx", ".xlsm"}:
        dest = tmp_dir / (safe_name(path.stem) + ".xls")
        shutil.copyfile(path, dest)
        return dest
    return path


def load_sheets(path: Path) -> list[tuple[str, list[list[str]]]]:
    kind = file_kind(path)
    errors: list[Exception] = []
    order = ["ole", "ooxml"] if kind == "ole" else ["ooxml", "ole"]
    for mode in order:
        try:
            if mode == "ooxml":
                return sheet_rows_openpyxl(path)
            return sheet_rows_xlrd(path)
        except Exception as e:
            errors.append(e)
    raise RuntimeError(
        f"could not read {path.name} as xlsx/xls: " + " | ".join(str(e) for e in errors)
    )


def rows_to_scheme_map(
    sheets: list[tuple[str, list[list[str]]]],
    *,
    file_stem: str,
    multi_sheet: bool,
) -> dict[str, list[dict[str, str]]]:
    scheme_map: dict[str, list[dict[str, str]]] = defaultdict(list)
    for sheet_name, rows in sheets:
        picked = choose_header_row(rows)
        if not picked:
            continue
        hidx, headers_raw = picked
        headers_raw = [h if h else f"col_{i+1}" for i, h in enumerate(headers_raw)]
        headers = alias_headers(headers_raw)
        sc_idx = scheme_col_index(headers_raw)

        if sc_idx is None:
            if multi_sheet:
                default_scheme = safe_name(sheet_name)
            else:
                default_scheme = safe_name(file_stem)
        else:
            default_scheme = safe_name(file_stem)

        for raw in rows[hidx + 1 :]:
            if not any((c or "").strip() for c in raw):
                continue
            lowered = [(c or "").strip().lower() for c in raw]
            if all(c in {"", "-", "--", "total", "totals", "grand total", "sub total", "subtotal"} for c in lowered):
                continue
            # stop on trailing totals block
            if any(c.startswith("total") for c in lowered[:3]) and sum(1 for c in lowered if c) <= 3:
                continue

            row_obj: dict[str, str] = {}
            for i, head in enumerate(headers):
                row_obj[head] = raw[i] if i < len(raw) else ""

            if sc_idx is not None and sc_idx < len(raw) and (raw[sc_idx] or "").strip():
                scheme_name = safe_name(raw[sc_idx])
            else:
                scheme_name = default_scheme

            filled = [v for v in row_obj.values() if (v or "").strip()]
            if len(filled) <= 1:
                continue
            # Prefer rows that look like holdings (isin or quantity/value)
            has_signal = bool(
                row_obj.get("isin")
                or row_obj.get("quantity")
                or row_obj.get("market_value")
                or row_obj.get("pct_nav")
                or any(re.match(r"[A-Z]{2}[A-Z0-9]{9}[0-9]$", v) for v in filled)
            )
            if not has_signal and len(filled) < 3:
                continue

            scheme_map[scheme_name].append(row_obj)
    return scheme_map


def write_scheme_outputs(out_amc: Path, scheme: str, rows: list[dict[str, str]]) -> None:
    scheme_dir = out_amc / safe_name(scheme)
    scheme_dir.mkdir(parents=True, exist_ok=True)
    (scheme_dir / "portfolio.json").write_text(
        json.dumps(rows, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    fields: list[str] = []
    seen: set[str] = set()
    for r in rows:
        for k in r:
            if k not in seen:
                seen.add(k)
                fields.append(k)
    with (scheme_dir / "portfolio.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def process_workbook(path: Path, *, tmp_dir: Path | None = None) -> dict[str, list[dict[str, str]]]:
    work = path
    if tmp_dir is not None:
        work = prepare_workbook_path(path, tmp_dir)
    sheets = load_sheets(work)
    multi = len(sheets) > 1
    return rows_to_scheme_map(sheets, file_stem=path.stem, multi_sheet=multi)


def process_amc(amc_dir: Path, out_amc: Path, *, disclosure_type: str = "monthly") -> AmcResult:
    amc = amc_dir.name
    result = AmcResult(amc=amc)
    if out_amc.exists():
        shutil.rmtree(out_amc)
    out_amc.mkdir(parents=True, exist_ok=True)

    sources = list_source_files(amc_dir)
    result.files_seen = len(sources)
    if not sources:
        result.notes.append("no spreadsheet/zip files")
        return result

    merged: dict[str, list[dict[str, str]]] = defaultdict(list)
    tmp_root = Path(tempfile.mkdtemp(prefix=f"parse-{amc}-"))
    try:
        work_files: list[Path] = []
        for src in sources:
            # Skip non-portfolio noise always
            if re.search(r"risk.?o.?meter|viewfile\.php", src.name, re.I):
                result.files_skipped += 1
                continue
            # Fortnight packs only when parsing fortnightly
            if re.search(r"fortnight", src.name, re.I) and disclosure_type != "fortnightly":
                result.files_skipped += 1
                continue
            # Mega/consolidated: skip on monthly; keep for fortnightly debt workbooks
            if disclosure_type != "fortnightly" and SKIP_FILE_RE.search(src.name) and src.suffix.lower() != ".zip":
                result.files_skipped += 1
                continue
            kind = file_kind(src)
            if kind == "zip" or src.suffix.lower() == ".zip":
                try:
                    extracted = expand_zip(src, tmp_root / src.stem)
                    if not extracted:
                        result.notes.append(f"zip empty/non-portfolio: {src.name}")
                        result.files_skipped += 1
                        continue
                    work_files.extend(extracted)
                    result.files_parsed += 1  # zip counted as parsed container
                except Exception as e:
                    result.errors.append(f"zip {src.name}: {e}")
                continue
            if SKIP_FILE_RE.search(src.name) and MEGA_FILE_RE.search(src.name):
                result.files_skipped += 1
                result.notes.append(f"skipped mega file: {src.name}")
                continue
            work_files.append(src)

        for path in work_files:
            try:
                scheme_map = process_workbook(path, tmp_dir=tmp_root)
                if not scheme_map:
                    result.notes.append(f"no holdings rows: {path.name}")
                    continue
                if path.parent == amc_dir or path.suffix.lower() != ".zip":
                    # count leaf workbooks (copied .xls→.xlsx still count once)
                    if path.parent == amc_dir or (tmp_root in path.parents and path.suffix.lower() in {".xlsx", ".xls", ".xlsb", ".xlsm"}):
                        result.files_parsed += 1
                for scheme, rows in scheme_map.items():
                    merged[scheme].extend(rows)
            except Exception as e:
                result.errors.append(f"{path.name}: {e}")

        for scheme, rows in sorted(merged.items()):
            write_scheme_outputs(out_amc, scheme, rows)
            result.rows += len(rows)
        result.schemes = len(merged)
        schemes_meta = [
            {"scheme": s, "rows": len(rows), "folder": safe_name(s)}
            for s, rows in sorted(merged.items())
        ]
        (out_amc / "schemes.json").write_text(
            json.dumps(schemes_meta, indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )
    finally:
        shutil.rmtree(tmp_root, ignore_errors=True)
    return result


def main() -> None:
    ap = argparse.ArgumentParser(description="Parse monthly MF portfolio disclosures → per-scheme holdings")
    ap.add_argument("--period", required=True, help="YYYY-MM")
    ap.add_argument(
        "--type",
        default="monthly",
        choices=("monthly", "fortnightly"),
        help="Disclosure cadence (default monthly)",
    )
    ap.add_argument("--root", type=Path, default=Path(__file__).resolve().parent.parent)
    ap.add_argument("--amc", action="append", default=[], help="Limit to AMC id (repeatable)")
    ap.add_argument("--limit-amcs", type=int, default=0, help="Process only first N AMCs (debug)")
    args = ap.parse_args()

    disc_root = args.root / "data" / "disclosures" / args.type / args.period
    out_root = args.root / "data" / "parsed" / args.type / args.period
    if not disc_root.is_dir():
        raise SystemExit(f"Missing disclosures dir: {disc_root}")

    amc_dirs = sorted([p for p in disc_root.iterdir() if p.is_dir()])
    if args.amc:
        wanted = set(args.amc)
        amc_dirs = [p for p in amc_dirs if p.name in wanted]
    if args.limit_amcs > 0:
        amc_dirs = amc_dirs[: args.limit_amcs]

    out_root.mkdir(parents=True, exist_ok=True)
    results: list[AmcResult] = []
    print(f"Parsing {len(amc_dirs)} AMC(s) for {args.period}", flush=True)
    print(f"  in:  {disc_root}", flush=True)
    print(f"  out: {out_root}", flush=True)

    for amc_dir in amc_dirs:
        print(f"\n→ {amc_dir.name}", flush=True)
        res = process_amc(amc_dir, out_root / amc_dir.name, disclosure_type=args.type)
        results.append(res)
        status = "ok" if res.schemes else ("error" if res.errors else "empty")
        print(
            f"  {status}: schemes={res.schemes} rows={res.rows} "
            f"files={res.files_parsed}/{res.files_seen} skipped={res.files_skipped} "
            f"errors={len(res.errors)}",
            flush=True,
        )
        for e in res.errors[:5]:
            print(f"    ERR {e}", flush=True)

    report = {
        "type": args.type,
        "period": args.period,
        "amcs": len(results),
        "with_schemes": sum(1 for r in results if r.schemes > 0),
        "empty": sum(1 for r in results if r.schemes == 0),
        "total_schemes": sum(r.schemes for r in results),
        "total_rows": sum(r.rows for r in results),
        "results": [asdict(r) for r in results],
    }
    report_path = out_root / "_report.json"
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(
        f"\nDone. amcs_ok={report['with_schemes']}/{report['amcs']} "
        f"schemes={report['total_schemes']} rows={report['total_rows']}",
        flush=True,
    )
    print(f"Report: {report_path}", flush=True)


if __name__ == "__main__":
    main()
