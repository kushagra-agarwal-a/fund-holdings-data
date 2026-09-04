#!/usr/bin/env python3
"""Inventory original holdings-table headers across AMC disclosures.

Walks disclosure workbooks, collects unique column titles per AMC, and maps
them onto a universal field list (ISIN, maturity, YTM, coupon, …).
"""
from __future__ import annotations

import csv
import json
import re
import shutil
import sys
import tempfile
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "parsers"))

from amc_parsers.common import (  # noqa: E402
    choose_header_row,
    file_kind,
    norm_cell,
    prepare_workbook_path,
    safe_name,
)
from amc_parsers.family import (  # noqa: E402
    MEGA_NAME,
    SKIP_FILE_RE,
    SKIP_SHEET,
    SKIP_SHEET_CONTAINS,
)

OUT = ROOT / "data" / "parsed" / "holding_columns"
PERIODS = [
    ("monthly", "2026-06"),
    ("fortnightly", "2026-07-15"),
    ("monthly", "latest"),
    ("fortnightly", "latest"),
]
EXTS = {".xlsx", ".xls", ".xlsm", ".xlsb", ".zip"}
ZIP_INNER_CAP = 24
HEADER_ROWS = 80

# More specific patterns first.
UNIVERSAL_FIELDS: list[tuple[str, str, re.Pattern[str]]] = [
    (
        "security_code",
        "Issuer / security / serial code",
        re.compile(r"(?i)^\s*(security\s*)?(code|no\.?)\s*$|^\s*sr\.?\s*no|^\s*scrip\s*code|^\s*sl\s*no|^\s*sr\s*$"),
    ),
    (
        "instrument",
        "Name of instrument / issuer / security",
        re.compile(
            r"(?i)name\s+of\s+(the\s+)?(instrument|security|issuer)|"
            r"company\s*/?\s*issuer\s*/?\s*instrument|"
            r"instrument\s*/?\s*issuer|security\s+name|"
            r"^(security|instrument|scrip|issuer|stock\s*name)\b"
        ),
    ),
    (
        "isin",
        "ISIN",
        re.compile(r"(?i)\bisin\b"),
    ),
    (
        "coupon",
        "Coupon rate",
        re.compile(r"(?i)\bcoupon\b"),
    ),
    (
        "ytm",
        "Yield to maturity",
        re.compile(r"(?i)\bytm\b|yield\s*to\s*maturity"),
    ),
    (
        "ytc",
        "Yield to call (incl. AT1 / Tier 2 ~YTC)",
        re.compile(r"(?i)\bytc\b|yield\s*to\s*call|at1|tier\s*2"),
    ),
    (
        "yield",
        "Yield of the instrument (when not labelled YTM/YTC)",
        re.compile(r"(?i)\byield\b"),
    ),
    (
        "residual_maturity",
        "Residual / remaining maturity or tenor",
        re.compile(r"(?i)residual|remaining\s*maturity|\btenor\b|term\s*to\s*maturity"),
    ),
    (
        "maturity_date",
        "Maturity date",
        re.compile(r"(?i)maturity\s*date|date\s*of\s*maturity|redemption\s*date|(?<!to )\bmaturity\b"),
    ),
    (
        "face_value",
        "Face / par / principal value",
        re.compile(r"(?i)face\s*value|par\s*value|\bprincipal\b"),
    ),
    (
        "put_call_date",
        "Put / call date",
        re.compile(r"(?i)put\s*/?\s*call|call\s*date|put\s*date"),
    ),
    (
        "industry_rating",
        "Combined industry / rating column",
        re.compile(
            r"(?i)industry\s*[\^+\*]?\s*/\s*rating|rating\s*/\s*industry|"
            r"industry\s+classification\s*/\s*rating"
        ),
    ),
    (
        "rating",
        "Credit rating",
        re.compile(r"(?i)credit\s*rating|^\s*rating\b|conservative\s+rating"),
    ),
    (
        "rating_agency",
        "Name of rating agency",
        re.compile(r"(?i)rating\s*agency|\bicra\b|\bcrisil\b|\bcare\b|\bind[- ]?ra\b"),
    ),
    (
        "industry",
        "Industry / sector",
        re.compile(r"(?i)\bindustry\b|\bsector\b"),
    ),
    (
        "quantity",
        "Quantity / units / shares",
        re.compile(
            r"(?i)^\s*quantity\b|\bqty\b|no\.?\s*of\s*(shares|units)|hedged\s+quantity|"
            r"^\s*units\s*$"
        ),
    ),
    (
        "futures_price",
        "Futures / option / contract price",
        re.compile(r"(?i)(futures?|option|contract)\s+price|current\s+price\s+of\s+the\s+contract"),
    ),
    (
        "market_value",
        "Market / fair value (usually Rs lakh)",
        re.compile(
            r"(?i)market\s*/?\s*fair\s*value|market[\s\-]*value|mkt\.?\s*[\-]?\s*val|"
            r"exposure\s*/\s*market|value\s*\(.*?(rs|inr|lakh|lac)|"
            r"value\s+recognised\s+in\s+nav|total\s+security\s+value"
        ),
    ),
    (
        "pct_nav",
        "% of NAV / AUM",
        re.compile(
            r"(?i)%\s*(to|of)?\s*(n\.?a\.?v|aum|net\s*assets?)|"
            r"percent(age)?\s*(to|of)?\s*(n\.?a\.?v|aum|net\s*assets?)|"
            r"%\s*to\s*nav|%\s*nav|^\s*percent\s*$"
        ),
    ),
    (
        "macaulay_duration",
        "Macaulay duration",
        re.compile(r"(?i)macaulay"),
    ),
    (
        "modified_duration",
        "Modified duration",
        re.compile(r"(?i)modified\s*duration"),
    ),
    (
        "duration",
        "Duration (unspecified)",
        re.compile(r"(?i)\bduration\b"),
    ),
    (
        "listed_status",
        "Listed / unlisted",
        re.compile(r"(?i)\blisted\b|\bunlisted\b"),
    ),
    (
        "exchange",
        "Exchange",
        re.compile(r"(?i)\bexchange\b"),
    ),
    (
        "ticker",
        "Ticker / NSE / BSE code",
        re.compile(r"(?i)\bticker\b|\bsymbol\b|nse\s*code|bse\s*code"),
    ),
    (
        "accrued_interest",
        "Accrued interest",
        re.compile(r"(?i)accrued\s*interest|interest\s*accrued"),
    ),
    (
        "clean_price",
        "Clean / dirty price",
        re.compile(r"(?i)\bclean\s*price\b|\bdirty\s*price\b|\bprice\b"),
    ),
    (
        "asset_class",
        "Asset class / type",
        re.compile(r"(?i)asset\s*(class|type)|instrument\s*type|type of security"),
    ),
    (
        "scheme_name",
        "Scheme / mutual fund name (mega-sheet column, not a holding field)",
        re.compile(r"(?i)scheme\s*name|mutual\s+fund\s+name|type of scheme|^\s*scheme\s*$"),
    ),
    (
        "position_side",
        "Long / short indicator",
        re.compile(r"(?i)long\s*/\s*\(?\s*short|outstanding position in derivative"),
    ),
    (
        "margin",
        "Margin amount",
        re.compile(r"(?i)\bmargin\b"),
    ),
    (
        "market_cap",
        "Market capitalization",
        re.compile(r"(?i)market\s*capital"),
    ),
    (
        "underlying",
        "Underlying of a derivative",
        re.compile(r"(?i)^\s*underlying\b"),
    ),
    (
        "trade_date",
        "Trade date (transaction annex)",
        re.compile(r"(?i)trade\s*date"),
    ),
    (
        "settlement_date",
        "Settlement date (transaction annex)",
        re.compile(r"(?i)settlement\s*date"),
    ),
    (
        "remarks",
        "Notes / remarks / footnotes",
        re.compile(r"(?i)\bremarks?\b|\bnotes?\b|\bfootnote"),
    ),
]

PARSED_TODAY = {
    "security_code",
    "instrument",
    "isin",
    "industry",
    "quantity",
    "market_value",
    "pct_nav",
    "ytm",
    "ytc",
}


def is_junk_sheet(name: str) -> bool:
    n = (name or "").strip()
    if not n or SKIP_SHEET.match(n):
        return True
    if SKIP_SHEET_CONTAINS.search(n) and not re.search(r"(?i)portfolio|holding|isin", n):
        return True
    return False


def collapse_header(s: str) -> str:
    t = (s or "").replace("\r", " ").replace("\n", " ")
    t = t.replace("_x000d_", " ").replace("_x000a_", " ")
    t = re.sub(r"\s+", " ", t).strip()
    t = t.strip(".:*|")
    return t


IGNORE_RE = re.compile(
    r"(?i)\bbrsr\b|\besg\b|risk-?o-?meter|prc\s*matrix|potential\s+risk\s+class|"
    r"assurance on|link to |"
    r"interim repayment|top\s*\d+\s*issuers|"
    r"transaction type|type of trade|value of the trade|"
    r"price at which traded|at which traded|invit units of |"
    r"^null$"
)


def map_header(original: str) -> str:
    h = collapse_header(original)
    if not h or re.fullmatch(r"col(_\d+)?", h, re.I) or h.lower() == "null":
        return "blank"
    if IGNORE_RE.search(h):
        return "ignore"
    for uid, _desc, rx in UNIVERSAL_FIELDS:
        if rx.search(h):
            return uid
    return "unmapped"


def sheet_rows_header_only(path: Path) -> list[tuple[str, list[list[str]]]]:
    """First HEADER_ROWS of each sheet — enough to find the SEBI header."""
    kind = file_kind(path)
    if kind == "ole":
        import xlrd

        book = xlrd.open_workbook(path)
        out = []
        for ws in book.sheets():
            rows = []
            cols = min(ws.ncols, 80)
            for r in range(min(ws.nrows, HEADER_ROWS)):
                rows.append([norm_cell(ws.cell_value(r, c)) for c in range(cols)])
            out.append((ws.name or "", rows))
        return out

    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    out = []
    try:
        for ws in wb.worksheets:
            rows = []
            for i, r in enumerate(ws.iter_rows(max_row=HEADER_ROWS, values_only=True)):
                cells = list(r[:80]) if r is not None else []
                rows.append([norm_cell(v) for v in cells])
            out.append((ws.title or "", rows))
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return out


def load_header_sheets(path: Path) -> list[tuple[str, list[list[str]]]]:
    with tempfile.TemporaryDirectory() as td:
        prep = prepare_workbook_path(path, Path(td))
        try:
            return sheet_rows_header_only(prep)
        except Exception:
            # fallback to full loader
            from amc_parsers.common import load_sheets

            return [(n, rows[:HEADER_ROWS]) for n, rows in load_sheets(path)]


def expand_zip(path: Path, dest: Path) -> list[Path]:
    out: list[Path] = []
    with zipfile.ZipFile(path) as zf:
        members = [
            i
            for i in zf.infolist()
            if not i.is_dir()
            and re.search(r"\.(xlsx|xls|xlsm|xlsb)$", Path(i.filename).name, re.I)
            and not Path(i.filename).name.startswith(".")
            and "__" not in Path(i.filename).name
        ]
        if len(members) > ZIP_INNER_CAP:
            step = max(1, len(members) // ZIP_INNER_CAP)
            members = members[::step][:ZIP_INNER_CAP]
        for info in members:
            name = Path(info.filename).name
            target = dest / safe_name(name)
            target.parent.mkdir(parents=True, exist_ok=True)
            if target.exists():
                target = dest / f"{safe_name(Path(name).stem)}_{len(out)}{Path(name).suffix.lower()}"
            with zf.open(info) as src, target.open("wb") as dst:
                shutil.copyfileobj(src, dst)
            out.append(target)
    return out


def iter_workbooks(amc_dir: Path, tmp: Path) -> list[Path]:
    files = [
        p
        for p in sorted(amc_dir.iterdir())
        if p.is_file() and p.suffix.lower() in EXTS and not SKIP_FILE_RE.search(p.name)
    ]
    out: list[Path] = []
    for p in files:
        if p.suffix.lower() == ".zip":
            inner = expand_zip(p, tmp / safe_name(p.stem))
            inner = [x for x in inner if not (MEGA_NAME.search(x.name) and len(inner) > 1)]
            out.extend(inner)
        else:
            out.append(p)
    return out


def collect_amc(amc_id: str, amc_dir: Path, cadence: str, period: str) -> dict[str, Any]:
    seen: dict[str, dict[str, Any]] = {}
    files_ok = 0
    files_err = 0
    sheets_with_header = 0
    errors: list[str] = []
    with tempfile.TemporaryDirectory(prefix="hdrzip_") as td:
        workbooks = iter_workbooks(amc_dir, Path(td))
        for path in workbooks:
            try:
                sheets = load_header_sheets(path)
            except Exception as e:
                files_err += 1
                if len(errors) < 8:
                    errors.append(f"{path.name}: {e}"[:200])
                continue
            files_ok += 1
            for sheet_name, rows in sheets:
                if is_junk_sheet(sheet_name):
                    continue
                picked = choose_header_row(rows)
                if not picked:
                    continue
                sheets_with_header += 1
                _idx, headers = picked
                originals = [collapse_header(h) for h in headers]
                originals = [h for h in originals if h]
                for orig in originals:
                    rec = seen.setdefault(
                        orig,
                        {
                            "original": orig,
                            "universal": map_header(orig),
                            "files": set(),
                            "sheets": set(),
                            "periods": set(),
                        },
                    )
                    rec["files"].add(path.name[:80])
                    rec["sheets"].add(sheet_name[:60])
                    rec["periods"].add(f"{cadence}/{period}")

    columns = []
    for orig, rec in sorted(seen.items(), key=lambda kv: (kv[1]["universal"], kv[0].lower())):
        columns.append(
            {
                "original": orig,
                "universal": rec["universal"],
                "file_count": len(rec["files"]),
                "sheet_count": len(rec["sheets"]),
                "periods": sorted(rec["periods"]),
                "sample_files": sorted(rec["files"])[:3],
                "sample_sheets": sorted(rec["sheets"])[:3],
            }
        )
    return {
        "amc_id": amc_id,
        "files_ok": files_ok,
        "files_err": files_err,
        "sheets_with_header": sheets_with_header,
        "unique_headers": len(columns),
        "unmapped": sum(1 for c in columns if c["universal"] == "unmapped"),
        "errors": errors,
        "columns": columns,
    }


def write_outputs(amcs: list[dict[str, Any]]) -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    payload = {
        "universal_fields": [
            {
                "universal": uid,
                "description": desc,
                "parsed_today": uid in PARSED_TODAY or uid == "industry_rating",
            }
            for uid, desc, _ in UNIVERSAL_FIELDS
        ]
        + [
            {
                "universal": "blank",
                "description": "Empty / unnamed leading column (usually security code)",
                "parsed_today": False,
            },
            {
                "universal": "ignore",
                "description": "Annex / ESG / riskometer / trade blotter — not a holdings field",
                "parsed_today": False,
            },
            {
                "universal": "unmapped",
                "description": "Header not classified — needs a rule",
                "parsed_today": False,
            },
        ],
        "amcs": amcs,
    }
    (OUT / "by_amc.json").write_text(json.dumps(payload, indent=2) + "\n")

    master_rows = []
    for a in amcs:
        for c in a["columns"]:
            master_rows.append(
                {
                    "amc_id": a["amc_id"],
                    "original_header": c["original"],
                    "universal": c["universal"],
                    "parsed_today": c["universal"] in PARSED_TODAY
                    or c["universal"] == "industry_rating",
                    "file_count": c["file_count"],
                    "sheet_count": c["sheet_count"],
                    "periods": "|".join(c["periods"]),
                    "sample_file": (c["sample_files"] or [""])[0],
                    "sample_sheet": (c["sample_sheets"] or [""])[0],
                }
            )

    fields = [
        "amc_id",
        "original_header",
        "universal",
        "parsed_today",
        "file_count",
        "sheet_count",
        "periods",
        "sample_file",
        "sample_sheet",
    ]
    with (OUT / "master_mapping.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(master_rows)

    # Unique original → universal (collapse across AMCs)
    by_orig: dict[str, dict[str, Any]] = {}
    for r in master_rows:
        rec = by_orig.setdefault(
            r["original_header"],
            {
                "original_header": r["original_header"],
                "universal": r["universal"],
                "amcs": set(),
            },
        )
        rec["amcs"].add(r["amc_id"])
        if rec["universal"] == "unmapped" and r["universal"] != "unmapped":
            rec["universal"] = r["universal"]

    uniq_path = OUT / "unique_headers.csv"
    with uniq_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f, fieldnames=["original_header", "universal", "amc_count", "amcs"]
        )
        w.writeheader()
        for orig, rec in sorted(by_orig.items(), key=lambda kv: (kv[1]["universal"], kv[0].lower())):
            w.writerow(
                {
                    "original_header": orig,
                    "universal": rec["universal"],
                    "amc_count": len(rec["amcs"]),
                    "amcs": "|".join(sorted(rec["amcs"])),
                }
            )

    # compact per-AMC unique list
    with (OUT / "by_amc_unique.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["amc_id", "original_header", "universal"])
        w.writeheader()
        for a in amcs:
            for c in a["columns"]:
                w.writerow(
                    {
                        "amc_id": a["amc_id"],
                        "original_header": c["original"],
                        "universal": c["universal"],
                    }
                )

    unmapped = [r for r in master_rows if r["universal"] == "unmapped"]
    summary = {
        "amcs": len(amcs),
        "mapping_rows": len(master_rows),
        "unique_original_headers": len(by_orig),
        "unmapped_rows": len(unmapped),
        "unmapped_unique": len({r["original_header"] for r in unmapped}),
        "universal_counts": {},
    }
    from collections import Counter

    summary["universal_counts"] = dict(Counter(r["universal"] for r in master_rows))
    (OUT / "summary.json").write_text(json.dumps(summary, indent=2) + "\n")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E79")

        def write_sheet(ws, headers, rows):
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            for row in rows:
                ws.append(list(row))
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"
            for i, h in enumerate(headers, 1):
                maxlen = len(h)
                for r in rows[:200]:
                    maxlen = max(maxlen, len(str(r[i - 1] or "")))
                ws.column_dimensions[get_column_letter(i)].width = min(48, max(12, maxlen + 2))

        ws = wb.active
        ws.title = "Universal fields"
        write_sheet(
            ws,
            ["universal", "description", "parsed_today"],
            [
                (u["universal"], u["description"], u["parsed_today"])
                for u in payload["universal_fields"]
            ],
        )
        ws = wb.create_sheet("Master mapping")
        write_sheet(
            ws,
            [
                "amc_id",
                "original_header",
                "universal",
                "parsed_today",
                "file_count",
                "sample_file",
            ],
            [
                (
                    r["amc_id"],
                    r["original_header"],
                    r["universal"],
                    r["parsed_today"],
                    r["file_count"],
                    r["sample_file"],
                )
                for r in master_rows
                if r["universal"] not in {"ignore"}
            ],
        )
        ws = wb.create_sheet("Unique headers")
        write_sheet(
            ws,
            ["original_header", "universal", "amc_count", "amcs"],
            [
                (rec["original_header"], rec["universal"], len(rec["amcs"]), "|".join(sorted(rec["amcs"])))
                for rec in sorted(by_orig.values(), key=lambda x: (x["universal"], x["original_header"].lower()))
                if rec["universal"] not in {"ignore"}
            ],
        )
        xlsx = OUT / "holding_column_mapping.xlsx"
        wb.save(xlsx)
    except Exception as e:
        print(f"xlsx skipped: {e}")


def main() -> None:
    if "--remap" in sys.argv:
        payload = json.loads((OUT / "by_amc.json").read_text())
        amcs = payload["amcs"]
        for a in amcs:
            for c in a["columns"]:
                c["universal"] = map_header(c["original"])
            a["unmapped"] = sum(1 for c in a["columns"] if c["universal"] == "unmapped")
            a["ignored"] = sum(1 for c in a["columns"] if c["universal"] == "ignore")
        write_outputs(amcs)
        print(json.dumps(json.loads((OUT / "summary.json").read_text()), indent=2))
        print(f"remapped {OUT}")
        return
    disc = ROOT / "data" / "disclosures"
    by_amc: dict[str, list[tuple[str, str, Path]]] = defaultdict(list)
    for cadence, period in PERIODS:
        root = disc / cadence / period
        if not root.is_dir():
            continue
        for amc_dir in sorted(root.iterdir()):
            if not amc_dir.is_dir() or amc_dir.name.startswith("_"):
                continue
            by_amc[amc_dir.name].append((cadence, period, amc_dir))

    amcs_out = []
    n = len(by_amc)
    for i, amc_id in enumerate(sorted(by_amc), 1):
        # Merge all period dirs for this AMC into one inventory.
        merged_cols: dict[str, dict[str, Any]] = {}
        files_ok = files_err = sheets = 0
        errors: list[str] = []
        print(f"  {i}/{n} {amc_id}", flush=True)
        for cadence, period, amc_dir in by_amc[amc_id]:
            part = collect_amc(amc_id, amc_dir, cadence, period)
            files_ok += part["files_ok"]
            files_err += part["files_err"]
            sheets += part["sheets_with_header"]
            errors.extend(part["errors"])
            for c in part["columns"]:
                rec = merged_cols.setdefault(
                    c["original"],
                    {
                        "original": c["original"],
                        "universal": c["universal"],
                        "file_count": 0,
                        "sheet_count": 0,
                        "periods": set(),
                        "sample_files": [],
                        "sample_sheets": [],
                    },
                )
                rec["file_count"] += c["file_count"]
                rec["sheet_count"] += c["sheet_count"]
                rec["periods"].update(c["periods"])
                for s in c["sample_files"]:
                    if s not in rec["sample_files"]:
                        rec["sample_files"].append(s)
                for s in c["sample_sheets"]:
                    if s not in rec["sample_sheets"]:
                        rec["sample_sheets"].append(s)
        columns = []
        for orig, rec in sorted(
            merged_cols.items(), key=lambda kv: (kv[1]["universal"], kv[0].lower())
        ):
            columns.append(
                {
                    "original": orig,
                    "universal": rec["universal"],
                    "file_count": rec["file_count"],
                    "sheet_count": rec["sheet_count"],
                    "periods": sorted(rec["periods"]),
                    "sample_files": rec["sample_files"][:3],
                    "sample_sheets": rec["sample_sheets"][:3],
                }
            )
        amcs_out.append(
            {
                "amc_id": amc_id,
                "files_ok": files_ok,
                "files_err": files_err,
                "sheets_with_header": sheets,
                "unique_headers": len(columns),
                "unmapped": sum(1 for c in columns if c["universal"] == "unmapped"),
                "errors": errors[:8],
                "columns": columns,
            }
        )

    write_outputs(amcs_out)
    print(json.dumps(json.loads((OUT / "summary.json").read_text()), indent=2))
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
